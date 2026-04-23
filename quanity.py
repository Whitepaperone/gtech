import pandas as pd
from openpyxl import load_workbook
from tkinter import filedialog

INPUT_FILE = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
OUTPUT_FILE = "./output.xlsx"

EXCLUDE_PROCESS = ["BRAZING", "Leak Test", "가열로", "출하-액세사리"]

def get_merged_value(ws, cell):
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col).value
    return cell.value


def extract():
    wb = load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active

    data = []

    # 👉 R열까지 = 18열
    MAX_COL = 18

    # 👉 보통 데이터 시작행 (필요시 조정)
    START_ROW = 3

    # 👉 컬럼 위치 (여기만 맞추면 끝)
    COL_PRODUCT = 10   # B열이면 1, C열이면 2
    COL_PROCESS = 6   # 공정명 위치
    COL_QTY = 17       # 실적수량 위치

    for row in ws.iter_rows(min_row=START_ROW, max_col=MAX_COL):
        values = [get_merged_value(ws, cell) for cell in row]

        product = values[COL_PRODUCT]
        process = values[COL_PROCESS]
        qty = values[COL_QTY]

        # 공정 필터
        if process and process in EXCLUDE_PROCESS:
            continue

        if product is None or qty is None:
            continue

        data.append([product, process, qty])

    df = pd.DataFrame(data, columns=["품번", "공정", "실적수량"])
    df.to_excel(OUTPUT_FILE, index=False)

    print("완료")


if __name__ == "__main__":
    extract()