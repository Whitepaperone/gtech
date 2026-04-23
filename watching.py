import os
import time
import shutil
import hashlib
import json
from datetime import datetime
from tkinter import filedialog
from typing import List, Tuple, Optional

from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side


# =========================
# 사용자 설정
# =========================
SOURCE_FILE = filedialog.askopenfilename(title="엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])                 # 감시할 원본 파일
if not SOURCE_FILE:
    raise SystemExit("엑셀 파일을 선택하지 않았습니다.")
WORK_DIR = r".\excel_watch"
STATE_FILE = os.path.join(WORK_DIR, "state.json")  # 상태 저장 파일 (anchor/last/result 경로 등)
ANCHOR_PREFIX = "생산계획 마지막_확인본"
LAST_PREFIX = "생산계획 최신본"
RESULT_PREFIX = "생산계획 변경이력 기록"

STATUS_FILE = os.path.join(WORK_DIR, "작업상태.txt")

def write_status(message: str) -> None:
    try:
        with open(STATUS_FILE, "w", encoding="utf-8") as f:
            f.write(message)
    except Exception:
        pass

def clear_status() -> None:
    try:
        if os.path.exists(STATUS_FILE):
            os.remove(STATUS_FILE)
    except Exception:
        pass

def source_mtime_str(path: str) -> str:
    return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y%m%d_%H%M%S")


def make_dated_path(prefix: str, source_path: str) -> str:
    ts = source_mtime_str(source_path)
    return os.path.abspath(os.path.join(WORK_DIR, f"{prefix}_{ts}.xlsx"))

# 비교 옵션
COMPARE_FORMULA_TEXT = False
# False: 셀의 표시값/값 기준 비교
# True : 수식 문자열까지 비교 (=SUM(A1:A3) 같은 식의 변경도 감지)

TARGET_SHEETS = ["완성공정(실적)", "TANK공정(실적)", "CORE공정(실적)"]


COMPARE_FORMULA_TEXT = False   # False: 결과값 비교 / True: 수식 문자열 비교
DEBOUNCE_SECONDS = 3.0
SAVE_WAIT_SECONDS = 2.5
COPY_RETRY_COUNT = 10
COPY_RETRY_DELAY = 0.5


# =========================
# 스타일
# =========================
# 변경 셀: 검정 배경 + 흰 글씨
FILL_CHANGED = PatternFill(fill_type="solid", fgColor="000000")
FONT_CHANGED = Font(color="FFFFFF", bold=True)

# 신규 시트/신규값도 같은 방식으로 눈에 띄게
FILL_NEW = PatternFill(fill_type="solid", fgColor="000000")
FONT_NEW = Font(color="FFFFFF", bold=True)

# 로그 시트 헤더
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# =========================
# 타입
# =========================
ChangeRecord = Tuple[str, str, str, str, str]
# (변경시각, 시트명, 셀주소, 이전값, 현재값)


# =========================
# 공용 함수
# =========================
def load_state() -> dict:
    if not os.path.exists(STATE_FILE):
        return {}

    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_state(state: dict) -> None:
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def ensure_dirs() -> None:
    os.makedirs(WORK_DIR, exist_ok=True)


def safe_copy(src: str, dst: str) -> None:
    last_error = None
    for _ in range(COPY_RETRY_COUNT):
        try:
            shutil.copy2(src, dst)
            return
        except Exception as e:
            last_error = e
            time.sleep(COPY_RETRY_DELAY)
    raise RuntimeError(f"파일 복사 실패: {src} -> {dst}, 오류: {last_error}")


def workbook_load(path: str):
    last_error = None
    for _ in range(COPY_RETRY_COUNT):
        try:
            return load_workbook(path, data_only=not COMPARE_FORMULA_TEXT)
        except Exception as e:
            last_error = e
            time.sleep(COPY_RETRY_DELAY)
    raise RuntimeError(f"엑셀 열기 실패: {path}, 오류: {last_error}")


def file_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_value(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def get_target_sheet_names(base_sheet_names: set, new_sheet_names: set) -> List[str]:
    if TARGET_SHEETS:
        return TARGET_SHEETS
    return sorted(base_sheet_names | new_sheet_names)


def apply_changed_style(cell) -> None:
    cell.fill = FILL_CHANGED
    cell.font = FONT_CHANGED


def set_log_header_style(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER


def set_log_body_style(cell) -> None:
    cell.border = THIN_BORDER


def autosize_log_columns(ws) -> None:
    widths = {
        "A": 22,
        "B": 20,
        "C": 12,
        "D": 35,
        "E": 35,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def summarize_changes(changes: List[ChangeRecord], sample_limit: int = 5):
    """
    시트별로 변경건수를 요약하고, 예시 셀 몇 개만 남긴다.
    반환:
    [
        (변경시각, 시트명, 변경건수, 예시셀문자열, 비고),
        ...
    ]
    """
    summary = {}
    order = []

    for changed_at, sheet_name, cell_addr, old_val, new_val in changes:
        if sheet_name not in summary:
            summary[sheet_name] = {
                "changed_at": changed_at,
                "count": 0,
                "samples": []
            }
            order.append(sheet_name)

        summary[sheet_name]["count"] += 1

        if len(summary[sheet_name]["samples"]) < sample_limit:
            summary[sheet_name]["samples"].append(cell_addr)

    rows = []
    for sheet_name in order:
        item = summary[sheet_name]
        count = item["count"]
        samples = ", ".join(item["samples"])
        note = ""
        if count > sample_limit:
            note = f"외 {count - sample_limit}건"

        rows.append(
            (
                item["changed_at"],
                sheet_name,
                count,
                samples,
                note
            )
        )

    return rows
# =========================
# 변경 비교
# =========================
def collect_changes(base_path: str, new_path: str) -> List[ChangeRecord]:
    wb_base = workbook_load(base_path)
    wb_new = workbook_load(new_path)

    base_sheet_names = set(wb_base.sheetnames)
    new_sheet_names = set(wb_new.sheetnames)

    target_sheet_names = get_target_sheet_names(base_sheet_names, new_sheet_names)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    changes: List[ChangeRecord] = []

    print(f"[비교] 대상 시트 수: {len(target_sheet_names)}")

    for idx, sheet_name in enumerate(target_sheet_names, start=1):
        print(f"[비교] 시트 시작 ({idx}/{len(target_sheet_names)}): {sheet_name}")

        in_base = sheet_name in base_sheet_names
        in_new = sheet_name in new_sheet_names

        if not in_base and not in_new:
            print(f"[비교] 시트 없음, 건너뜀: {sheet_name}")
            continue

        if in_new and not in_base:
            ws_new = wb_new[sheet_name]
            print(f"[비교] 신규 시트: {sheet_name}, 행 {ws_new.max_row}, 열 {ws_new.max_column}")
            for row in range(1, ws_new.max_row + 1):
                if row % 200 == 0:
                    print(f"[비교중] {sheet_name}: {row}/{ws_new.max_row} 행")
                for col in range(1, ws_new.max_column + 1):
                    val = ws_new.cell(row=row, column=col).value
                    if val not in (None, ""):
                        changes.append((now_str, sheet_name, ws_new.cell(row=row, column=col).coordinate, "", normalize_value(val)))
            print(f"[비교] 시트 완료: {sheet_name}")
            continue

        if in_base and not in_new:
            ws_base = wb_base[sheet_name]
            print(f"[비교] 삭제 시트: {sheet_name}, 행 {ws_base.max_row}, 열 {ws_base.max_column}")
            for row in range(1, ws_base.max_row + 1):
                if row % 200 == 0:
                    print(f"[비교중] {sheet_name}: {row}/{ws_base.max_row} 행")
                for col in range(1, ws_base.max_column + 1):
                    old_val = ws_base.cell(row=row, column=col).value
                    if old_val not in (None, ""):
                        changes.append((now_str, sheet_name, ws_base.cell(row=row, column=col).coordinate, normalize_value(old_val), "[시트삭제]"))
            print(f"[비교] 시트 완료: {sheet_name}")
            continue

        ws_base = wb_base[sheet_name]
        ws_new = wb_new[sheet_name]

        max_row = max(ws_base.max_row, ws_new.max_row)
        max_col = max(ws_base.max_column, ws_new.max_column)

        print(f"[비교] 기존 시트: {sheet_name}, 행 {max_row}, 열 {max_col}")

        for row in range(1, max_row + 1):
            if row % 200 == 0:
                print(f"[비교중] {sheet_name}: {row}/{max_row} 행")
            for col in range(1, max_col + 1):
                old_val = normalize_value(ws_base.cell(row=row, column=col).value)
                new_val = normalize_value(ws_new.cell(row=row, column=col).value)

                if old_val != new_val:
                    changes.append((now_str, sheet_name, ws_new.cell(row=row, column=col).coordinate, old_val, new_val))

        print(f"[비교] 시트 완료: {sheet_name}")

    print(f"[비교 완료] 총 변경 건수: {len(changes)}")
    return changes


# =========================
# 결과 파일 생성/갱신
# =========================
def replace_with_timestamped_copy(src: str, old_path: Optional[str], prefix: str) -> str:
    new_path = make_dated_path(prefix, src)

    if old_path and os.path.abspath(old_path) == os.path.abspath(new_path) and os.path.exists(old_path):
        return old_path

    safe_copy(src, new_path)

    if old_path and os.path.exists(old_path) and os.path.abspath(old_path) != os.path.abspath(new_path):
        try:
            os.remove(old_path)
        except Exception as e:
            print(f"[삭제 실패] {old_path} / {e}")
            pass

    return new_path


def replace_with_timestamped_result(source_path: str, old_result_path: Optional[str], changes: List[ChangeRecord]) -> str:
    new_result_path = make_dated_path(RESULT_PREFIX, source_path)
    print(f"[결과경로] 기존: {old_result_path}")
    print(f"[결과경로] 신규: {new_result_path}")
    if old_result_path and os.path.exists(old_result_path) and os.path.abspath(old_result_path) == os.path.abspath(new_result_path):
        print("[결과경로] 같은 파일명 사용")
        create_or_update_result_file(source_path, old_result_path, changes)
        return old_result_path

    print("[결과경로] 새 파일 생성")
    create_or_update_result_file(source_path, new_result_path, changes)

    if old_result_path and os.path.exists(old_result_path) and os.path.abspath(old_result_path) != os.path.abspath(new_result_path):
        try:
            os.remove(old_result_path)
            print(f"[결과경로] 이전 결과 삭제: {old_result_path}")
        except Exception:
            print(f"[결과경로] 이전 결과 삭제: {old_result_path}")
            pass

    return new_result_path

def create_or_update_result_file(source_path: str, result_path: str, changes: List[ChangeRecord]) -> None:
    """
    변경이 있을 때만 호출
    최신 원본 파일을 바탕으로 RESULT_FILE 하나만 갱신
    변경 셀 스타일 적용
    변경내역 시트는 시트별 요약만 기록
    """
    # 기존 로그 읽기
    existing_logs = []
    if os.path.exists(result_path):
        try:
            wb_old = load_workbook(result_path)
            if "변경내역" in wb_old.sheetnames:
                ws_old_log = wb_old["변경내역"]
                for row in ws_old_log.iter_rows(min_row=2, values_only=True):
                    if not any(v is not None and str(v).strip() != "" for v in row):
                        continue
                    existing_logs.append(
                        (
                            "" if row[0] is None else str(row[0]),  # 변경시각
                            "" if row[1] is None else str(row[1]),  # 시트명
                            0 if row[2] is None else int(row[2]),   # 변경건수
                            "" if row[3] is None else str(row[3]),  # 예시셀
                            "" if row[4] is None else str(row[4]),  # 비고
                        )
                    )
        except Exception:
            existing_logs = []

    print(f"[결과파일] 열기 시작: {source_path}")
    wb_result = workbook_load(source_path)
    print(f"[결과파일] 열기 완료: {source_path}")

    # 변경 셀 스타일 반영
    for idx, (changed_at, sheet_name, cell_addr, old_val, new_val) in enumerate(changes, start=1):
        if sheet_name in wb_result.sheetnames:
            ws = wb_result[sheet_name]
            ws[cell_addr].fill = FILL_CHANGED
            ws[cell_addr].font = FONT_CHANGED

        if idx % 5000 == 0:
            print(f"[결과파일] 변경 셀 스타일 적용 중: {idx}/{len(changes)}")

    # 기존 변경내역 시트 제거
    if "변경내역" in wb_result.sheetnames:
        del wb_result["변경내역"]

    # 변경내역 시트 새로 생성
    ws_log = wb_result.create_sheet("변경내역", 0)
    headers = ["변경시각", "시트명", "변경 건수", "예시 셀", "비고"]
    ws_log.append(headers)

    for col in range(1, len(headers) + 1):
        set_log_header_style(ws_log.cell(row=1, column=col))

    # 신규 변경을 시트별 요약
    summary_rows = summarize_changes(changes, sample_limit=5)

    # 기존 요약 + 신규 요약 누적
    all_logs = existing_logs + summary_rows
    print(f"[결과파일] 변경내역 요약 기록: {len(all_logs)}행")

    for idx, record in enumerate(all_logs, start=1):
        ws_log.append(list(record))
        current_row = ws_log.max_row
        for col in range(1, 6):
            set_log_body_style(ws_log.cell(row=current_row, column=col))

        if idx % 100 == 0:
            print(f"[결과파일] 요약 기록 중: {idx}/{len(all_logs)}")

    ws_log.freeze_panes = "A2"
    ws_log.column_dimensions["A"].width = 22
    ws_log.column_dimensions["B"].width = 20
    ws_log.column_dimensions["C"].width = 12
    ws_log.column_dimensions["D"].width = 40
    ws_log.column_dimensions["E"].width = 15

    print(f"[결과파일] 저장 시작: {result_path}")

    temp_result_path = result_path + ".writing.xlsx"
    wb_result.save(temp_result_path)

    # 기존 최종 파일이 있으면 삭제
    if os.path.exists(result_path):
        try:
            os.remove(result_path)
        except Exception as e:
            print(f"[결과파일] 기존 파일 삭제 실패: {e}")

    os.replace(temp_result_path, result_path)

    print(f"[결과파일] 저장 완료: {result_path}")

def validate_source_for_anchor(anchor_path: str, new_source_path: str) -> bool:
    wb_anchor = workbook_load(anchor_path)
    wb_new = workbook_load(new_source_path)

    anchor_sheets = set(wb_anchor.sheetnames)
    new_sheets = set(wb_new.sheetnames)

    return any(sheet in new_sheets for sheet in TARGET_SHEETS if sheet in anchor_sheets)
# =========================
# 감시 클래스
# =========================
class ExcelWatcher:
    def __init__(self, source_file: str):
        self.source_file = os.path.normcase(os.path.abspath(source_file))

        self.anchor_file: Optional[str] = None
        self.last_file: Optional[str] = None
        self.result_file: Optional[str] = None

        self.last_event_time = 0.0
        self.last_processed_hash: Optional[str] = None
        self.prev_result_exists = False
        self.force_compare_once = False

    def save_current_state(self) -> None:
        state = {
            "source_file": self.source_file,
            "anchor_file": self.anchor_file,
            "last_file": self.last_file,
            "result_file": self.result_file,
            "last_processed_hash": self.last_processed_hash,
            "prev_result_exists": self.prev_result_exists,
            "force_compare_once": self.force_compare_once,
        }
        save_state(state)

    def initialize(self) -> None:
        ensure_dirs()
        if self.anchor_file and not validate_source_for_anchor(self.anchor_file, self.source_file):
            print("[경고] 새 원본 파일의 시트 구성이 앵커와 많이 다릅니다.")

        if not os.path.exists(self.source_file):
            raise FileNotFoundError(f"원본 파일이 없습니다: {self.source_file}")

        state = load_state()


        # 같은 원본 파일이면 이전 상태 복구
    
        self.anchor_file = state.get("anchor_file")
        self.last_file = state.get("last_file")
        self.result_file = state.get("result_file")
        self.last_processed_hash = state.get("last_processed_hash")
        self.prev_result_exists = state.get("prev_result_exists", False)

        self.force_compare_once = False
        old_source_file = state.get("source_file")



        # 파일이 실제 존재하는지 검증
        if self.anchor_file and not os.path.exists(self.anchor_file):
            self.anchor_file = None
        if self.last_file and not os.path.exists(self.last_file):
            self.last_file = None
        if self.result_file and not os.path.exists(self.result_file):
            self.result_file = None

        if self.anchor_file:
            print(f"[복구] anchor 사용: {self.anchor_file}")
        if self.last_file:
            print(f"[복구] last 사용: {self.last_file}")
        if self.result_file:
            print(f"[복구] result 사용: {self.result_file}")

        # 없으면 새로 생성
        if not self.anchor_file:
            self.anchor_file = replace_with_timestamped_copy(self.source_file, None, ANCHOR_PREFIX)
            print(f"[초기화] anchor 생성: {self.anchor_file}")

        if not self.last_file:
            self.last_file = replace_with_timestamped_copy(self.source_file, None, LAST_PREFIX)
            print(f"[초기화] last 생성: {self.last_file}")

        if old_source_file:
            old_source_file = os.path.normcase(os.path.abspath(old_source_file))

        current_source_file = os.path.normcase(os.path.abspath(self.source_file))

        # source 파일명이 바뀌었거나 state가 없으면 첫 비교는 강제로 수행
        if old_source_file != current_source_file or (self.anchor_file and not self.result_file):
            self.force_compare_once = True
        else:
            self.force_compare_once = False

        if not self.last_processed_hash:
            self.last_processed_hash = file_hash(self.source_file)


                # 기본값
        self.force_compare_once = False

        # 1) 파일명이 바뀌면 강제 비교
        if old_source_file != current_source_file:
            self.force_compare_once = True
            print("[초기화] 원본 파일 변경 감지 → 앵커 기준으로 1회 강제 비교")

        # 2) 결과본이 없는데 앵커가 있으면 강제 비교
        elif self.anchor_file and not self.result_file:
            self.force_compare_once = True
            print("[초기화] 결과본 없음 → 앵커 기준으로 1회 강제 비교")

        # 3) last_processed_hash가 아예 없을 때만 현재 해시 저장
        if not self.last_processed_hash:
            self.last_processed_hash = file_hash(self.source_file)

    def process_deleted_result_reset(self) -> None:
        current_result_exists = bool(self.result_file and os.path.exists(self.result_file))

        if self.prev_result_exists and not current_result_exists:
            print("[RESET] 변경 확인 완료로 판단 → 기준 재설정")
            try:
                self.anchor_file = replace_with_timestamped_copy(self.source_file, self.anchor_file, ANCHOR_PREFIX)
                self.last_file = replace_with_timestamped_copy(self.source_file, self.last_file, LAST_PREFIX)
                self.last_processed_hash = file_hash(self.source_file)
                self.result_file = None
                print(f"[RESET] anchor 갱신: {self.anchor_file}")
                print(f"[RESET] last 갱신: {self.last_file}")
                self.save_current_state()
            except Exception as e:
                print(f"[RESET 오류] {e}")

        self.prev_result_exists = current_result_exists

    def process_change(self) -> None:
        self.process_deleted_result_reset()

        current_hash = file_hash(self.source_file)

        if not self.force_compare_once:
            if self.last_processed_hash and current_hash == self.last_processed_hash:
                print("[건너뜀] 파일 내용 변경 없음")
                return

        print(f"[비교 시작] {self.source_file}")
        changes = collect_changes(self.anchor_file, self.source_file)

        if not changes:
            print("[변경 없음] 결과 파일 생성/갱신 안 함")
            self.last_file = replace_with_timestamped_copy(self.source_file, self.last_file, LAST_PREFIX)
            self.last_processed_hash = current_hash
            self.force_compare_once = False
            self.save_current_state()
            print(f"[last 갱신] {self.last_file}")
            return

        self.result_file = replace_with_timestamped_result(self.source_file, self.result_file, changes)
        print(f"[결과 저장] {self.result_file}")
        print(f"[변경 건수] {len(changes)}건")

        self.last_file = replace_with_timestamped_copy(self.source_file, self.last_file, LAST_PREFIX)
        self.last_processed_hash = current_hash
        self.force_compare_once = False
        self.prev_result_exists = bool(self.result_file and os.path.exists(self.result_file))
        self.save_current_state()

        print(f"[last 갱신] {self.last_file}")

class ExcelFileEventHandler(FileSystemEventHandler):

   
    def __init__(self, watcher: ExcelWatcher):
        self.watcher = watcher
       

    def on_moved(self, event):
        if event.is_directory:
            return

        dest_path = os.path.normcase(os.path.abspath(event.dest_path))
        if dest_path != self.watcher.source_file:
            return

        now = time.time()
        if now - self.watcher.last_event_time < DEBOUNCE_SECONDS:
            return

        self.watcher.last_event_time = now
        time.sleep(SAVE_WAIT_SECONDS)

        try:
            self.watcher.process_change()
        except Exception as e:
            print(f"[오류] 처리 실패: {e}")
    def on_modified(self, event):
        self._handle(event)

    def on_created(self, event):
        self._handle(event)

    def _handle(self, event):
        if event.is_directory:
            return

        event_path = os.path.normcase(os.path.abspath(event.src_path))
        source_path = self.watcher.source_file

        # 정확 일치 또는 파일명 일치 시 처리
        if event_path != source_path and os.path.basename(event_path) != os.path.basename(source_path):
            return

        now = time.time()
        if now - self.watcher.last_event_time < DEBOUNCE_SECONDS:
            return

        self.watcher.last_event_time = now
        time.sleep(SAVE_WAIT_SECONDS)

        try:
            self.watcher.process_change()
        except Exception as e:
            print(f"[오류] 처리 실패: {e}")


def main():
    watcher = ExcelWatcher(SOURCE_FILE)
    watcher.initialize()

    watch_dir = os.path.dirname(os.path.abspath(SOURCE_FILE))
    event_handler = ExcelFileEventHandler(watcher)
    observer = Observer()
    observer.schedule(event_handler, watch_dir, recursive=False)
    observer.start()

    print("====================================")
    print("엑셀 감시 시작")
    print(f"원본 파일 : {watcher.source_file}")
    print(f"기준본     : {watcher.anchor_file}")
    print(f"결과본     : {watcher.result_file}")
    print(f"대상 시트  : {TARGET_SHEETS if TARGET_SHEETS else '전체 시트'}")
    print("====================================")
    print("종료: Ctrl + C")

    try:
       while True:
            watcher.process_deleted_result_reset()

            current_hash = file_hash(watcher.source_file)

            if watcher.force_compare_once or current_hash != watcher.last_processed_hash:
                print("[루프] process_change 호출")
                watcher.process_change()

            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("감시 종료")

    observer.join()


if __name__ == "__main__":
    main()