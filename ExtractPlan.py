from typing import Optional, List, Dict
import tkinter as tk
from tkinter import filedialog, messagebox
from tkcalendar import Calendar
import pandas as pd
from openpyxl import load_workbook

from AppConstants import (
    DATA_START_ROW,
    DATE_HEADER_GAP_BREAK,
    DATE_SCAN_MIN_COL,
    HEADER_DATE_ROW,
    HEADER_KIND_ROW,
    MES_SHEET_CONFIG,
    PLAN_OUTPUT_FILE,
    PLAN_SHEET_NAMES,
    WORKSHOP_TO_PROCESS,
)
from CommonUtils import (
    build_merged_map,
    header_text,
    is_date_value,
    normalize_kind,
    normalize_part_no,
    normalize_process_token,
    safe_float,
)

# =========================
# 설정
# =========================
SHEET_NAME = PLAN_SHEET_NAMES
OUTPUT_FILE = PLAN_OUTPUT_FILE


# =========================
# 열 찾기
# =========================
def find_date_start_col(ws) -> Optional[int]:
    for c in range(DATE_SCAN_MIN_COL, ws.max_column + 1):
        if is_date_value(ws.cell(HEADER_DATE_ROW, c).value):
            return c
    return None


def find_col_by_keywords(ws, keywords, required: bool = False, date_start_col: Optional[int] = None) -> Optional[int]:
    keys = [k.upper().replace(" ", "") for k in keywords]
    date_start_col = date_start_col or find_date_start_col(ws) or (ws.max_column + 1)

    for c in range(1, date_start_col):
        txt = header_text(ws, HEADER_DATE_ROW, c)
        txt = txt.upper().replace(" ", "").replace("\n", "")

        if all(k in txt for k in keys):
            return c

    if required:
        raise RuntimeError(f"{ws.title}: 헤더를 찾지 못했습니다. keywords={keywords}")

    return None


def build_header_lookup(ws) -> Dict[str, Optional[int]]:
    date_start_col = find_date_start_col(ws) or (ws.max_column + 1)
    return {
        "model_col": find_col_by_keywords(ws, ["모델"], required=True, date_start_col=date_start_col),
        "customer_part_col": find_col_by_keywords(ws, ["고객사", "품번"], date_start_col=date_start_col),
        "core_part_col": find_col_by_keywords(ws, ["CORE", "품번"], date_start_col=date_start_col),
        "tank_part_col": find_col_by_keywords(ws, ["TANK", "품번"], date_start_col=date_start_col),
        "finish_part_col": find_col_by_keywords(ws, ["완성", "품번"], date_start_col=date_start_col),
        "accessory_part_col": find_col_by_keywords(ws, ["액세", "품번"], date_start_col=date_start_col),
        "process_print_col": find_col_by_keywords(ws, ["공정", "인쇄"], date_start_col=date_start_col),
    }


def build_plan_date_columns(ws) -> List[Dict]:
    """
    날짜 열 중에서 구분행이 '계획'인 열만 추출.
    미달 / 실적은 스킵.
    """
    cols = []
    started = False
    invalid_streak = 0

    start_col = find_date_start_col(ws)
    if start_col is None:
        raise RuntimeError(f"{ws.title}: 날짜 시작 열을 찾지 못했습니다.")

    for c in range(start_col, ws.max_column + 1):
        date_val = ws.cell(HEADER_DATE_ROW, c).value
        kind_val = normalize_process_token(ws.cell(HEADER_KIND_ROW, c).value).replace("\n", "")

        if is_date_value(date_val):
            started = True
            invalid_streak = 0

            # 핵심: 미달 / 실적은 스킵, 계획만 읽음
            if kind_val != "계획":
                continue

            cols.append({
                "col": c,
                "date": pd.to_datetime(date_val).normalize(),
                "kind": kind_val,
            })
            continue

        if started:
            invalid_streak += 1
            if invalid_streak >= DATE_HEADER_GAP_BREAK:
                break

    return cols


def get_sheet_columns(ws) -> Dict[str, Optional[int]]:
    cols = build_header_lookup(ws)
    if not cols["finish_part_col"]:
        raise RuntimeError(f"{ws.title}: 헤더를 찾지 못했습니다. keywords={['완성', '품번']}")
    return cols


# =========================
# 스킵 조건
# =========================


def is_skip_summary_row(*values) -> bool:
    text = " ".join([normalize_process_token(v) for v in values if normalize_process_token(v)])

    if not text:
        return True

    skip_keywords = ["합계", "소계", "TOTAL", "누계"]

    return any(k in text.upper() for k in skip_keywords)


def should_skip_row(part_no, process_name, row_text="", sheet_title: str = "", product_key=None) -> bool:
    part_txt = normalize_process_token(part_no)
    product_txt = normalize_process_token(product_key if product_key is not None else part_no)
    process_txt = normalize_process_token(process_name)
    row_txt = normalize_process_token(row_text)

    if part_txt.startswith("HH") or product_txt.startswith("HH"):
        return True

    if part_txt == "2":
        return True

    if sheet_title == "TANK공정(실적)" and "클린칭C" in process_txt:
        return True

    if "개발,기타" in process_txt or "개발,기타" in row_txt:
        return True

    return False


def starts_with_alpha_part_no(part_no) -> bool:
    part_txt = normalize_process_token(part_no)
    return bool(part_txt) and part_txt[0].isalpha() and part_txt[0].isascii()


# =========================
# 생산계획 추출
# =========================
def extract_finish_plan_sheet(ws) -> pd.DataFrame:
    return _extract_plan_sheet(ws, mode="finish")


def build_mes_date_columns(ws) -> List[Dict]:
    cols = []
    started = False
    invalid_streak = 0
    start_col = find_date_start_col(ws)

    if start_col is None:
        return []

    for c in range(start_col, ws.max_column + 1):
        date_val = ws.cell(HEADER_DATE_ROW, c).value
        kind_val = ws.cell(HEADER_KIND_ROW, c).value

        if is_date_value(date_val):
            started = True
            invalid_streak = 0
            kind = normalize_kind(kind_val) or "계획"
            cols.append({
                "col": c,
                "date": pd.to_datetime(date_val).normalize(),
                "kind": kind,
            })
            continue

        if started:
            invalid_streak += 1
            if invalid_streak >= DATE_HEADER_GAP_BREAK:
                break

    return cols


def get_mes_sheet_columns(ws) -> Dict[str, Optional[int]]:
    return build_header_lookup(ws)


def is_red_font_cell(ws, row: int, col: int) -> bool:
    cell = ws.cell(row, col)
    color = getattr(cell.font, "color", None)
    if color is None:
        return False

    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str):
        rgb = rgb.upper()
        if rgb in {"FFFF0000", "00FF0000", "FF0000"}:
            return True

    return False


def map_workcenter_and_team(process_name):
    p_raw = normalize_process_token(process_name)
    p = normalize_process_token(p_raw)

    if "AL" == p or "AL" in p:
        return "CORE조립", "CORE조립-AL"
    if "CU" == p or "CU" in p:
        return "CORE조립", "CORE조립-CU"
    if "수동" in p_raw:
        return "용접", "TANK용접 - 수동"
    if "로봇" in p:
        return "용접", "TANK용접 - ROBOT"
    if "클린칭" in p:
        return "CLINCHING", "CLINCHING"
    if "TANK조립" in p_raw:
        return "TANK조립&Leak Test", "TANK조립&Leak Test"
    if "한국" in p_raw:
        return "용접", "TANK용접 - 한국RAD"
    if "일반" in p_raw:
        return "완성조립", "완성조립공정 - 일반"
    if "G2" in p:
        return "완성조립", "완성조립공정 - G2"
    if "클라크" in p_raw:
        return "완성조립", "완성조립공정 - 클라크"
    if "특수품" in p_raw:
        return "완성조립", "완성조립공정 - 특수품"
    if "액세" in p_raw or "HEATSCREEN" in p:
        return "출하-액세사리", "출하-액세사리"
    return None, None


def map_process_from_workshop(name: str) -> Optional[str]:
    return WORKSHOP_TO_PROCESS.get(normalize_process_token(name))


def pick_plan_part_no(workshop_name: str, row_values: Dict[str, str]) -> str:
    w = normalize_process_token(workshop_name)

    if w == "CORE조립":
        return row_values.get("core_part_no", "")
    if w in ("용접", "CLINCHING", "TANK조립&LEAKTEST"):
        return row_values.get("tank_part_no", "")
    if w == "완성조립":
        return row_values.get("finish_part_no", "")
    if "액세" in w:
        return row_values.get("accessory_part_no", "") or row_values.get("finish_part_no", "")

    return (
        row_values.get("finish_part_no", "")
        or row_values.get("tank_part_no", "")
        or row_values.get("core_part_no", "")
    )


def make_compare_key(process: str, workshop: str, team: str, part_no: str, day) -> str:
    process = normalize_process_token(process)
    workshop = normalize_process_token(workshop)
    team = normalize_process_token(team)
    part_no = normalize_part_no(part_no)
    day = pd.to_datetime(day, errors="coerce").normalize()

    if process == "완성" and "액세" in workshop:
        return f"{process}|{part_no}|{day}"

    return f"{process}|{workshop}|{team}|{part_no}|{day}"


def extract_plan_sheet(ws, config: Dict, check_red_font: bool = False, progress=None) -> pd.DataFrame:
    return _extract_plan_sheet(ws, mode="mes", config=config, check_red_font=check_red_font, progress=progress)


def select_date_range(parent):
    win = tk.Toplevel(parent)
    win.title("기간 선택")
    win.geometry("620x330")
    win.resizable(False, False)

    selected = {"start": None, "end": None}

    tk.Label(win, text="시작일").grid(row=0, column=0, padx=10, pady=(10, 0))
    tk.Label(win, text="종료일").grid(row=0, column=1, padx=10, pady=(10, 0))

    cal_start = Calendar(win, selectmode="day", date_pattern="yyyy-mm-dd")
    cal_start.grid(row=1, column=0, padx=10, pady=10)

    cal_end = Calendar(win, selectmode="day", date_pattern="yyyy-mm-dd")
    cal_end.grid(row=1, column=1, padx=10, pady=10)

    def confirm():
        start = pd.to_datetime(cal_start.get_date()).normalize()
        end = pd.to_datetime(cal_end.get_date()).normalize()

        if start > end:
            messagebox.showerror("오류", "시작일이 종료일보다 늦습니다.", parent=win)
            return

        selected["start"] = start
        selected["end"] = end
        win.destroy()

    def no_filter():
        selected["start"] = None
        selected["end"] = None
        win.destroy()

    tk.Button(win, text="확인", width=15, command=confirm).grid(row=2, column=0, pady=10)
    tk.Button(win, text="기간 필터 없이 진행", width=20, command=no_filter).grid(row=2, column=1, pady=10)

    win.protocol("WM_DELETE_WINDOW", no_filter)
    win.update_idletasks()
    win.lift()
    win.attributes("-topmost", True)
    win.after(300, lambda: win.attributes("-topmost", False))
    win.focus_force()
    win.grab_set()

    parent.wait_window(win)
    return selected["start"], selected["end"]


def filter_work_order_period(df: pd.DataFrame, start_date=None, end_date=None) -> pd.DataFrame:
    if df.empty:
        return df

    date_col = df.columns[8]
    out = df.copy()
    out[date_col] = pd.to_datetime(out[date_col], errors="coerce").dt.normalize()

    if start_date is not None:
        out = out[out[date_col] >= start_date]
    if end_date is not None:
        out = out[out[date_col] <= end_date]

    return out.copy()


def extract_work_order_file(plan_file: str, progress=None) -> pd.DataFrame:
    if progress:
        progress(8, "생산계획 통합 문서를 여는 중...")

    wb = load_workbook(plan_file, data_only=True)
    frames = []
    sheet_items = [
        (sheet_name, config)
        for sheet_name, config in MES_SHEET_CONFIG.items()
        if sheet_name in wb.sheetnames
    ]
    total = max(len(sheet_items), 1)

    for i, (sheet_name, config) in enumerate(sheet_items, start=1):
        if progress:
            start = 10 + int((i - 1) / total * 65)
            progress(start, f"{sheet_name} 시트 추출 중...")

        def sheet_progress(percent, text=None, i=i, sheet_name=sheet_name):
            if progress:
                base = 10 + ((i - 1) / total * 65)
                span = 65 / total
                progress(base + (percent / 100 * span), text or f"{sheet_name} 시트 추출 중...")

        df_sheet = extract_plan_sheet(
            wb[sheet_name],
            config,
            check_red_font=True,
            progress=sheet_progress,
        )
        if not df_sheet.empty:
            frames.append(df_sheet)

        if progress:
            done = 10 + int(i / total * 65)
            progress(done, f"{sheet_name} 시트 추출 완료")

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


def build_work_order_upload_df(plan_df: pd.DataFrame) -> pd.DataFrame:
    output_columns = ["작업장명", "작업반명", "모델", "품번", "공정 인쇄"]

    if plan_df.empty:
        return pd.DataFrame(columns=output_columns)

    cols = list(plan_df.columns)
    workcenter_col = cols[3]
    team_col = cols[4]
    model_col = cols[5]
    part_col = cols[6]
    date_col = cols[8]
    kind_col = cols[9]
    qty_col = cols[10]
    process_print_col = cols[11]

    df = plan_df.copy()
    plan_kind = normalize_kind("계획") or "계획"
    df = df[df[kind_col].apply(normalize_kind).fillna(df[kind_col]) == plan_kind].copy()
    if df.empty:
        return pd.DataFrame(columns=output_columns)

    df[date_col] = pd.to_datetime(df[date_col], errors="coerce").dt.normalize()
    df[part_col] = df[part_col].apply(normalize_part_no)
    df = df[~df[part_col].isin(["", "NAN", "NONE"])].copy()
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)
    df = df[df[qty_col] != 0].copy()

    grouped = (
        df.groupby(
            [workcenter_col, team_col, model_col, part_col, process_print_col, date_col],
            as_index=False,
            dropna=False,
        )[qty_col]
        .sum()
    )

    wide = grouped.pivot_table(
        index=[workcenter_col, team_col, model_col, part_col, process_print_col],
        columns=date_col,
        values=qty_col,
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    date_columns = sorted([c for c in wide.columns if isinstance(c, pd.Timestamp)])
    wide = wide[[workcenter_col, team_col, model_col, part_col, process_print_col] + date_columns]
    wide = wide.rename(
        columns={
            workcenter_col: "작업장명",
            team_col: "작업반명",
            model_col: "모델",
            part_col: "품번",
            process_print_col: "공정 인쇄",
        }
    )
    wide = wide.rename(columns={c: c.strftime("%Y/%m/%d") for c in date_columns})

    qty_cols = [c for c in wide.columns if c not in output_columns]
    wide = wide[(wide[qty_cols] != 0).any(axis=1)].copy()
    return wide.reset_index(drop=True)


def _extract_plan_sheet(ws, mode: str, config: Optional[Dict] = None, check_red_font: bool = False, progress=None) -> pd.DataFrame:
    is_mes = mode == "mes"
    config = config or {}
    merged_map = build_merged_map(ws)
    date_cols = build_mes_date_columns(ws) if is_mes else build_plan_date_columns(ws)

    if not date_cols:
        raise RuntimeError(f"{ws.title}: 날짜/구분 컬럼을 찾지 못했습니다.")

    cols = get_mes_sheet_columns(ws) if is_mes else get_sheet_columns(ws)
    records = []
    current_model = ""
    accessory_mode = False
    max_col = ws.max_column
    model_col = cols["model_col"]
    customer_part_col = cols["customer_part_col"]
    core_part_col = cols["core_part_col"] if is_mes else None
    tank_part_col = cols["tank_part_col"] if is_mes else None
    finish_part_col = cols["finish_part_col"]
    accessory_part_col = cols["accessory_part_col"] if is_mes else None
    process_print_col = cols["process_print_col"]

    for row_cells in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, max_col=max_col):
        r = row_cells[0].row
        if is_mes and (r - DATA_START_ROW) % 200 == 0:
            percent = (r - DATA_START_ROW) / max(ws.max_row - DATA_START_ROW, 1) * 100
            if progress:
                progress(percent, f"{ws.title} 계획 추출 중... {percent:.1f}%")

        values = [cell.value for cell in row_cells]

        model_val = normalize_process_token(values[model_col - 1])
        if model_val:
            current_model = model_val
        model = current_model
        row_text = " ".join(
            text
            for text in (normalize_process_token(cell.value) for cell in row_cells)
            if text
        )

        row_values = {
            "customer_part_no": normalize_process_token(values[customer_part_col - 1]) if customer_part_col else "",
            "core_part_no": normalize_process_token(values[core_part_col - 1]) if core_part_col else "",
            "tank_part_no": normalize_process_token(values[tank_part_col - 1]) if tank_part_col else "",
            "finish_part_no": normalize_process_token(values[finish_part_col - 1]) if finish_part_col else "",
            "accessory_part_no": normalize_process_token(values[accessory_part_col - 1]) if accessory_part_col else "",
        }

        first_col_raw = merged_map.get((r, 1), values[0])
        first_col_text = normalize_process_token(first_col_raw)

        if is_mes and ws.title == "완성공정(실적)":
            if "용접C/M" in first_col_text or "코어C/M" in first_col_text or "선발주-용접C/M" in first_col_text:
                continue

        row_text_norm = normalize_process_token(row_text)
        if "액세" in row_text_norm and "HEATSCREEN" in row_text_norm:
            accessory_mode = True
            continue

        if accessory_mode:
            stop_prefix = "단품" if is_mes else "제외"
            if first_col_text.startswith(stop_prefix):
                break

        process_name_raw = ""
        if accessory_mode and ws.title == "완성공정(실적)":
            process_name_raw = "액세사리 & HEAT SCREEN"
        elif process_print_col:
            process_name_raw = normalize_process_token(values[process_print_col - 1])

        if is_mes:
            if not process_name_raw:
                process_name_raw = config.get("default_process", "")

            workshop_name, team_name = map_workcenter_and_team(process_name_raw)

            if workshop_name and "액세" in workshop_name:
                workshop_name = "출하-액세사리"
                team_name = "출하-액세사리"

            if not workshop_name:
                if config.get("default_process") == "CORE":
                    workshop_name, team_name = "CORE조립", ""
                elif config.get("default_process") == "TANK":
                    workshop_name, team_name = "용접", ""
                elif config.get("default_process") == "완성":
                    workshop_name, team_name = "완성조립", ""

            part_no = pick_plan_part_no(workshop_name, row_values)
            product_key = row_values.get("customer_part_no") or part_no
        else:
            part_no = row_values["finish_part_no"]
            product_key = row_values.get("customer_part_no") or part_no

        if should_skip_row(part_no, process_name_raw, row_text, sheet_title=ws.title, product_key=product_key):
            continue

        if is_skip_summary_row(model, part_no, row_text):
            continue

        if not model or not part_no:
            continue

        if not starts_with_alpha_part_no(part_no):
            continue

        for dc in date_cols:
            if is_mes and check_red_font and is_red_font_cell(ws, r, dc["col"]):
                continue

            qty = values[dc["col"] - 1]
            if qty is None or qty == "":
                continue

            qty_num = safe_float(qty)
            if qty_num == 0:
                continue

            if is_mes:
                process = map_process_from_workshop(workshop_name)
                workshop_raw = workshop_name or ""
                workshop = normalize_process_token(workshop_name)
                team = team_name or ""
                part_norm = normalize_part_no(part_no)

                records.append({
                    "비교키": make_compare_key(process, workshop, team, part_norm, dc["date"]),
                    "시트명": ws.title,
                    "공정": process,
                    "작업장명": workshop_raw,
                    "작업반명": team,
                    "모델": model,
                    "품번": part_norm,
                    "고객사품번": row_values.get("customer_part_no", ""),
                    "날짜": dc["date"],
                    "구분": dc["kind"],
                    "수량": qty_num,
                    "공정인쇄": process_name_raw,
                    "원본행": r,
                    "원본열": dc["col"],
                })
            else:
                records.append({
                    "시트명": ws.title,
                    "모델": model,
                    "품번": normalize_part_no(part_no),
                    "고객사품번": row_values.get("customer_part_no", ""),
                    "날짜": dc["date"],
                    "구분": dc["kind"],
                    "계획수량": qty_num,
                    "공정인쇄": process_name_raw,
                })

    df = pd.DataFrame(records)
    if df.empty or not is_mes:
        return df

    return df[df["공정"].notna()].copy()


def extract_plan_file(plan_file: str, progress=None) -> pd.DataFrame:
    wb = load_workbook(plan_file, data_only=True)

    frames = []

    total = len(SHEET_NAME)
    for i, sheet_name in enumerate(SHEET_NAME, start=1):
        if sheet_name not in wb.sheetnames:
            if progress:
                progress(10 + int(i / total * 70), f"{sheet_name} 시트 없음, 건너뜀")
            continue

        if progress:
            progress(10 + int((i - 1) / total * 70), f"{sheet_name} 계획 추출 중...")

        ws = wb[sheet_name]
        df = extract_finish_plan_sheet(ws)

        if progress:
            progress(10 + int(i / total * 70), f"{sheet_name} 계획 추출 완료: {len(df)}건")
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


# =========================
# 실행
# =========================
def main():
    root = tk.Tk()
    root.withdraw()

    plan_file = filedialog.askopenfilename(
        title="생산계획 엑셀 선택",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not plan_file:
        raise SystemExit("생산계획 엑셀을 선택하지 않았습니다.")

    plan_df = extract_plan_file(plan_file)

    if plan_df.empty:
        messagebox.showinfo("알림", "추출된 계획 수량이 없습니다.")
        return

    plan_df.to_excel(OUTPUT_FILE, index=False)
    messagebox.showinfo("완료", f"저장 완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
