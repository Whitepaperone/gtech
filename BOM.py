import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import Tk, filedialog, messagebox
from datetime import datetime

# =========================
# 설정값
# =========================

BOM_PART_COL = "품번"
BOM_LEVEL_COL = "레벨"

ADJ_PART_COL = "품번"
ADJ_LOT_TYPE_COL = "LOT유형"
ADJ_QTY_COL = "전산수량"

SUP_PART_COL = "품번(코드)"
SUP_NAME_COL = "거래처명"

PLAN_PART_COL_INDEX = 5     # F열
PLAN_QTY_COL_INDEX = 52     # BA열


# =========================
# 파일 선택
# =========================

from openpyxl import load_workbook



def select_file(title, parent):
    return filedialog.askopenfilename(
        parent=parent,
        title=title,
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )


def clean_part_no(value):
    if pd.isna(value):
        return ""
    return str(value).strip().replace("-", "").upper()


def read_plan_file(plan_file):

    df = pd.read_excel(
        plan_file,
        sheet_name="완성공정(실적)",
        header=None,
        usecols=[5, 52],   # F, BA
        skiprows=35,
        engine="openpyxl"
    )

    df.columns = ["품번", "계획수량"]

    df["품번정리"] = df["품번"].apply(clean_part_no)

    df["계획수량"] = (
        pd.to_numeric(df["계획수량"], errors="coerce")
        .fillna(0)
    )

    plan_sum = (
        df[df["품번정리"] != ""]
        .groupby("품번정리", as_index=False)["계획수량"]
        .sum()
    )

    plan_sum = plan_sum[
        plan_sum["계획수량"] > 0
    ]

    planned_parts = set(plan_sum["품번정리"])

    return planned_parts, plan_sum


def filter_bom_with_children(bom_df, planned_parts):
    """
    BOM 구조에서 계획 품번이 나오면,
    해당 행부터 다음 동일/상위 레벨이 나오기 전까지 하위레벨을 모두 유지
    """

    keep_indexes = []

    levels = pd.to_numeric(bom_df[BOM_LEVEL_COL], errors="coerce")
    parts = bom_df["품번정리"]

    i = 0
    while i < len(bom_df):
        part = parts.iloc[i]
        level = levels.iloc[i]

        if part in planned_parts and pd.notna(level):
            keep_indexes.append(i)

            j = i + 1
            while j < len(bom_df):
                next_level = levels.iloc[j]

                if pd.notna(next_level) and next_level <= level:
                    break

                keep_indexes.append(j)
                j += 1

            i = j
        else:
            i += 1

    return bom_df.iloc[sorted(set(keep_indexes))].copy()

def move_column_after(df, column_name, after_column):
    cols = list(df.columns)

    if column_name not in cols or after_column not in cols:
        return df

    cols.remove(column_name)
    idx = cols.index(after_column) + 1
    cols.insert(idx, column_name)

    return df[cols]

def format_result_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    # 레벨 컬럼 찾기
    level_col = None
    for cell in ws[1]:
        if cell.value == "레벨":
            level_col = cell.column
            break

    # 기본 스타일
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(
                vertical="center",
            )

    # 레벨 0 행 색칠
    if level_col:
        fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"  # 연노랑
        )

        for row in range(2, ws.max_row + 1):
            level_value = ws.cell(row=row, column=level_col).value

            try:
                is_level_zero = int(level_value) == 0
            except:
                is_level_zero = False

            if is_level_zero:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
                    ws.cell(row=row, column=col).font = Font(
                        name="맑은 고딕",
                        size=10,
                        bold=True
                    )

    wb.save(file_path)
    wb.close()


def main():
    root = Tk()
    root.withdraw()

    # 최상위로 올리기
    root.attributes("-topmost", True)
    root.lift()
    root.focus_force()

    bom_file = select_file("BOM 기준 재고 현황 파일 선택", root)
    if not bom_file:
        return

    adj_file = select_file("재고실사조정 파일 선택", root)
    if not adj_file:
        return

    supplier_file = select_file("협력사(기준정보액셀관리) 파일 선택", root)
    if not supplier_file:
        return

    plan_file = select_file("계획 수량 파일 선택", root)
    if not plan_file:
        return

    save_file = filedialog.asksaveasfilename(
        parent=root,
        title="결과 파일 저장",
        initialfile=f"BOM_결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not save_file:
        return

    # =========================
    # BOM 기준 재고 현황
    # =========================

    bom_df = pd.read_excel(bom_file)
    bom_df["품번정리"] = bom_df[BOM_PART_COL].apply(clean_part_no)

    # =========================
    # 재고실사조정 - 재공만 추출
    # =========================

    adj_df = pd.read_excel(adj_file)
    adj_df["품번정리"] = adj_df[ADJ_PART_COL].apply(clean_part_no)

    wip_df = adj_df[
        adj_df[ADJ_LOT_TYPE_COL].astype(str).str.strip() == "재공"
    ].copy()

    wip_sum = (
        wip_df
        .groupby("품번정리", as_index=False)[ADJ_QTY_COL]
        .sum()
        .rename(columns={ADJ_QTY_COL: "재공"})
    )

    bom_df = bom_df.merge(wip_sum, on="품번정리", how="left")
    bom_df["재공"] = bom_df["재공"].fillna(0)

    # =========================
    # 협력사 정보 삽입
    # =========================

    sup_df = pd.read_excel(supplier_file)
    sup_df["품번정리"] = sup_df[SUP_PART_COL].apply(clean_part_no)

    sup_df = sup_df[["품번정리", SUP_NAME_COL]].drop_duplicates("품번정리")
    sup_df = sup_df.rename(columns={SUP_NAME_COL: "협력사"})

    bom_df = bom_df.merge(sup_df, on="품번정리", how="left")

    # =========================
    # 계획 품번 읽기
    # =========================

    planned_parts, plan_qty_df = read_plan_file(plan_file)

    # =========================
    # 계획 품번 + 하위레벨 유지
    # =========================

    bom_df = bom_df.drop(columns=["계획수량"], errors="ignore")
    bom_df = bom_df.merge(plan_qty_df, on="품번정리", how="left")
    bom_df["계획수량"] = bom_df["계획수량"].fillna(0)

    result_df = filter_bom_with_children(bom_df, planned_parts)

    # 정리용 컬럼 제거
    result_df = result_df.drop(columns=["품번정리"], errors="ignore")

    result_df = move_column_after(result_df, "협력사", "품번")
    result_df = move_column_after(result_df, "계획수량", "소요량")
    result_df = move_column_after(result_df, "재공", "재고량")

    result_df.to_excel(save_file, index=False)

    format_result_excel(save_file)

    messagebox.showinfo("완료", f"결과 파일 생성 완료\n\n{save_file}")


if __name__ == "__main__":
    main()