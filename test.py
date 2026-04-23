# file: mes_hybrid.py

import pandas as pd
import xlwings as xw
from openpyxl import load_workbook

FILE_PATH = "./26년 3월 생산계획(기간별) 260319.xlsx"

DATE_ROW = 33
DATA_START_ROW = 35
PRODUCT_COL = "F"
DATA_START_COL = "BB"


# -----------------------------
# 1. 색 판별
# -----------------------------
def get_color(cell):
    try:
        color = cell.font.color

        if color is None:
            return "black"

        # RGB 직접 색상
        if color.type == "rgb":
            rgb = color.rgb.upper()

            if rgb is None:
                return "black"

            # 🔥 초록 (#00B050)
            if rgb.endswith("00B050"):
                return "green"

            # 🔥 보라 (#9966FF)
            if rgb.endswith("9966FF"):
                return "purple"

        return "black"

    except:
        return "black"


# -----------------------------
# 2. 데이터 추출 (xlwings)
# -----------------------------
def extract_values():
    wb = xw.Book(FILE_PATH)
    ws = wb.sheets[0]

    df = pd.read_excel(FILE_PATH, header=None)

    # 날짜
    dates = ws.range("BB33").expand("right").value
    dates = pd.to_datetime(dates, errors="coerce")

    # 🔥 핵심: products를 pandas에서 읽는다
    products = df.iloc[34:, 5].tolist()

    # 수량
    data = ws.range("BB35").expand("table").value
    df_data = pd.DataFrame(data)

    return df_data, dates, products


# -----------------------------
# 3. 색 추출 (openpyxl)
# -----------------------------
def extract_colors(n_rows, n_cols):
    wb = load_workbook(FILE_PATH, data_only=True)
    ws = wb.active

    color_map = {}

    for r in range(n_rows):
        for c in range(n_cols):
            excel_row = DATA_START_ROW + r
            excel_col = 54 + c  # BB = 54 (1-based)

            cell = ws.cell(row=excel_row, column=excel_col)
            color = get_color(cell)

            color_map[(r, c)] = color

    return color_map


# -----------------------------
# 4. 데이터 결합
# -----------------------------
def build_dataframe(df, dates, products, color_map):
    rows = []

    for r in range(df.shape[0]):
        product = str(products[r]).strip().upper()

        if not product.startswith("H"):
            continue

        for c in range(df.shape[1]):
            date = dates[c]

            if pd.isna(date):
                continue

            val = pd.to_numeric(df.iloc[r, c], errors="coerce")

           

            color = color_map.get((r, c), "black")

            rows.append({
                "product": product,
                "date": date,
                "qty": val,
                "type": color
            })

    return pd.DataFrame(rows)


# -----------------------------
# 5. 데이터 분리
# -----------------------------
def split_data(df):
    plan = df[df["type"] == "black"].copy()
    green = df[df["type"] == "green"].copy()
    purple = df[df["type"] == "purple"].copy()

    plan = plan.groupby(["product", "date"])["qty"].sum().reset_index()
    green = green.groupby(["product", "date"])["qty"].sum().reset_index()
    purple = purple.groupby(["product", "date"])["qty"].sum().reset_index()

    return plan, green, purple


# -----------------------------
# 6. MRP 계산
# -----------------------------
def calculate(plan, green):
    start_date, end_date = get_user_date_range()
    result = []

    for p in plan["product"].unique():
        p_plan = plan[plan["product"] == p].sort_values("date").copy()
        p_green = green[green["product"] == p].sort_values("date")

        stock = 0
        shortage = 0

        last = p_green.iloc[-1]["qty"]

        if last < 0:
            stock = abs(last)
        else:
                shortage = last

        # 🔥 2. 기간 필터
        p_plan = p_plan[
            (p_plan["date"] >= start_date) &
            (p_plan["date"] <= end_date)
        ].copy()

        # 🔥 3. 미달을 첫날로 이동
        if shortage > 0 and not p_plan.empty:
            first_date = p_plan.iloc[0]["date"]
            p_plan.loc[p_plan["date"] == first_date, "qty"] += shortage
            shortage = 0  # 핵심

        # 🔥 4. 계산 (미달 다시 반영 금지)
        for _, row in p_plan.iterrows():
            date = row["date"]
            plan_qty = row["qty"]

            if stock >= plan_qty:
                stock -= plan_qty
            else:
                stock = -(plan_qty - stock)

            print(f"제품: {p}, 날짜: {date.date()}, 계획: {plan_qty}, 재고: {stock}")

            result.append({
                "product": p,
                "date": date,
                "plan_qty": plan_qty,
                "stock": stock
            })

    return pd.DataFrame(result)

def get_user_date_range():
    start = input("시작 날짜 입력 (YYYY-MM-DD): ")
    end = input("종료 날짜 입력 (YYYY-MM-DD): ")

    start = pd.to_datetime(start)
    end = pd.to_datetime(end)

    return start, end
def adjust_plan_with_range(p_plan, stock, start_date,end_date):
    # 🔥 기간 이후만 필터
    p_plan = p_plan[ (p_plan["date"] >= start_date) &
        (p_plan["date"] <= end_date)].copy()

    # 🔥 미달이 있으면 첫날로 몰기
    if stock < 0 and not p_plan.empty:
        first_date = p_plan.iloc[0]["date"]
        p_plan.loc[p_plan["date"] == first_date, "qty"] += abs(stock)
        stock = 0

    return p_plan, stock
# -----------------------------
# 7. 실행
# -----------------------------
def main():
    df, dates, products = extract_values()

    print("값 shape:", df.shape)

    color_map = extract_colors(df.shape[0], df.shape[1])

    merged = build_dataframe(df, dates, products, color_map)

    print("추출 데이터:", len(merged))

    plan, green, purple = split_data(merged)

    final = calculate(plan, green)

    final.to_excel("mrp_result.xlsx", index=False)

    print("완료: mrp_result.xlsx")


if __name__ == "__main__":
    main()