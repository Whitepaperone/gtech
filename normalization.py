from datetime import datetime, date
from typing import List, Dict, Optional
from tkinter import Tk, filedialog

import pandas as pd
from openpyxl import load_workbook, utils

root = Tk()
root.withdraw()

PLAN_FILE = filedialog.askopenfilename(
    title="생산계획 엑셀 선택",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)
if not PLAN_FILE:
    raise SystemExit("생산계획 엑셀을 선택하지 않았습니다.")

MES_FILE = filedialog.askopenfilename(
    title="MES 작업지시 엑셀 선택",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)
if not MES_FILE:
    raise SystemExit("MES 작업지시 엑셀을 선택하지 않았습니다.")

OUTPUT_FILE = "./비교결과.xlsx"


# =========================
# 생산계획 시트 공통 구조
# =========================
HEADER_DATE_ROW = 33
HEADER_KIND_ROW = 34
DATA_START_ROW = 35
DATE_COL_START = 54   # BB열 부근
DATE_COL_END = utils.column_index_from_string("DL")  # DL열 부근

SHEET_CONFIG = {
    "완성공정(실적)": {
        "process": "완성",
        "model_col": 2,   # B
        "part_col": 6,    # F
    },
    "TANK공정(실적)": {
        "process": "TANK",
        "model_col": 2,   # B
        "part_col": 5,    # E (필요시 조정)
    },
    "CORE공정(실적)": {
        "process": "CORE",
        "model_col": 2,   # B
        "part_col": 4,    # D (필요시 조정)
    },
}


# =========================
# 유틸
# =========================
def normalize_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def is_date_value(v) -> bool:
    return isinstance(v, (datetime, date))


def normalize_kind(kind_text: str) -> Optional[str]:
    t = normalize_text(kind_text).replace("\n", "")
    if t in ("미달", "계획", "실적"):
        return t
    return None


def is_skip_row(model: str, part_no: str) -> bool:
    if not model or not part_no:
        return True

    text = f"{model} {part_no}"
    skip_keywords = ["합계", "소계", "TOTAL", "누계"]
    return any(k in text for k in skip_keywords)


# =========================
# 생산계획 정규화
# =========================
def build_date_columns(ws) -> List[Dict]:
    cols = []

    for c in range(DATE_COL_START, DATE_COL_END + 1):
        date_val = ws.cell(HEADER_DATE_ROW, c).value
        kind_val = ws.cell(HEADER_KIND_ROW, c).value

        if not is_date_value(date_val):
            continue

        kind = normalize_kind(kind_val)
        if kind is None:
            continue

        cols.append({
            "col": c,
            "date": pd.to_datetime(date_val).date(),
            "kind": kind,
        })

    return cols


def extract_plan_sheet(ws, config: Dict) -> pd.DataFrame:
    date_cols = build_date_columns(ws)
    if not date_cols:
        raise RuntimeError(f"{ws.title}: 날짜/구분 컬럼을 찾지 못했습니다.")

    model_col = config["model_col"]
    part_col = config["part_col"]
    process_name = config["process"]

    records = []
    current_model = ""

    for r in range(DATA_START_ROW, ws.max_row + 1):
        model_val = normalize_text(ws.cell(r, model_col).value)
        part_no = normalize_text(ws.cell(r, part_col).value)

        if model_val:
            current_model = model_val

        model = current_model

        if is_skip_row(model, part_no):
            continue

        for dc in date_cols:
            qty = ws.cell(r, dc["col"]).value
            if qty is None or qty == "":
                continue

            try:
                qty_num = float(qty)
            except Exception:
                continue

            if qty_num == 0:
                continue

            records.append({
                "시트명": ws.title,
                "공정": process_name,
                "모델": model,
                "품번": part_no,
                "날짜": dc["date"],
                "구분": dc["kind"],   # 미달 / 계획 / 실적
                "수량": qty_num,
                "원본행": r,
                "원본열": dc["col"],
            })

    df = pd.DataFrame(records)
    if df.empty:
        return df

    return df[[
        "시트명", "공정", "모델", "품번",
        "날짜", "구분", "수량", "원본행", "원본열"
    ]]


def merge_plan(df: pd.DataFrame) -> pd.DataFrame:
    """
    TANK와 CORE는 같은 품번 + 같은 날짜의 '계획'은 합침.
    단, 미달/실적은 합치지 않음.
    """
    if df.empty:
        return df

    target_df = df[df["공정"].isin(["TANK", "CORE"])].copy()
    other_df = df[~df["공정"].isin(["TANK", "CORE"])].copy()

    target_plan = target_df[target_df["구분"] == "계획"].copy()
    target_non_plan = target_df[target_df["구분"] != "계획"].copy()

    if not target_plan.empty:
        target_plan = (
            target_plan
            .groupby(
                ["시트명", "공정", "모델", "품번", "날짜", "구분"],
                as_index=False
            )
            .agg({
                "수량": "sum",
                "원본행": "min",
                "원본열": "min",
            })
        )

    merged = pd.concat([other_df, target_non_plan, target_plan], ignore_index=True)
    merged = merged.sort_values(
        ["공정", "모델", "품번", "날짜", "구분"]
    ).reset_index(drop=True)

    return merged

def extract_plan_all(plan_file: str) -> pd.DataFrame:
    wb = load_workbook(plan_file, data_only=True)
    frames = []

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"[건너뜀] 시트 없음: {sheet_name}")
            continue

        ws = wb[sheet_name]
        df_sheet = extract_plan_sheet(ws, config)
        print(f"[계획 추출] {sheet_name}: {len(df_sheet)}건")
        frames.append(df_sheet)

    if not frames:
        return pd.DataFrame()

    df = pd.concat(frames, ignore_index=True)
    df = merge_plan(df)
    return df


# =========================
# MES 정규화
# =========================
WORKSHOP_MAP = {
    "용접": "TANK",
    "CLINCHING": "TANK",
    "TANK조립&Leak Test": "TANK",
    "CORE조립" : "CORE",
    "출하-악세서리": "완성",
    "완성조립": "완성",
}


def map_process_from_workshop(name: str) -> Optional[str]:
    t = normalize_text(name)
    return WORKSHOP_MAP.get(t)


def extract_mes(mes_file: str) -> pd.DataFrame:
    # 차트시트가 있어도 Sheet만 읽도록 지정
    df = pd.read_excel(mes_file, sheet_name="Sheet")

    needed = ["계획일", "작업장명", "작업지시상태", "품번", "지시량", "실적량"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise RuntimeError(f"MES 파일에 필요한 컬럼이 없습니다: {missing}")

    df = df[needed].copy()
    df["공정"] = df["작업장명"].map(map_process_from_workshop)

    # 비교 대상 공정만
    df = df[df["공정"].notna()].copy()

    # 종료 제외
    df = df[df["작업지시상태"] != "종료"].copy()

    # 날짜 정리
    df["날짜"] = pd.to_datetime(df["계획일"]).dt.date

    # 수치형
    df["지시량"] = pd.to_numeric(df["지시량"], errors="coerce").fillna(0)
    df["실적량"] = pd.to_numeric(df["실적량"], errors="coerce").fillna(0)

    # 비교용 집계
    grouped = (
        df.groupby(["공정", "품번", "날짜"], as_index=False)
          .agg({
              "지시량": "sum",
              "실적량": "sum",
          })
    )

    return grouped.sort_values(["공정", "품번", "날짜"]).reset_index(drop=True)

def apply_latest_midal_carry(plan_base: pd.DataFrame) -> pd.DataFrame:
    """
    최신 미달량을 다음 계획일들에 carry 적용.
    공정+품번별로 날짜순으로 처리.
    - 최신미달 > 0 : 부족 → 다음 계획수량에 더함
    - 최신미달 < 0 : 초과 → 다음 계획수량에서 차감
    """
    if plan_base.empty:
        return plan_base

    df = plan_base.copy()
    df["조정계획수량"] = df["계획수량"]

    result_rows = []

    for (process, part_no), g in df.groupby(["공정", "품번"], sort=False):
        g = g.sort_values("날짜").copy()

        latest_midal = 0
        latest_midal_date = None

        non_zero_midal = g[g["최신미달수량"] != 0]
        if not non_zero_midal.empty:
            latest_row = non_zero_midal.sort_values("최신미달일").tail(1).iloc[0]
            latest_midal = float(latest_row["최신미달수량"])
            latest_midal_date = latest_row["최신미달일"]

        carry = latest_midal

        adjusted_list = []
        for _, row in g.iterrows():
            row = row.copy()
            plan_qty = float(row["계획수량"])

            # 최신 미달 이후 날짜에만 carry 적용
            if latest_midal_date is not None and row["날짜"] >= latest_midal_date:
                if carry > 0:
                    # 부족분은 계획에 추가
                    plan_qty = plan_qty + carry
                    carry = 0

                elif carry < 0:
                    # 초과 생산은 다음 계획에서 차감
                    need = plan_qty
                    consume = min(need, abs(carry))
                    plan_qty = need - consume
                    carry = carry + consume  # carry는 음수이므로 0에 가까워짐

            row["조정계획수량"] = plan_qty
            row["잔여carry"] = carry
            adjusted_list.append(row)

        result_rows.append(pd.DataFrame(adjusted_list))

    out = pd.concat(result_rows, ignore_index=True)
    return out.sort_values(["공정", "품번", "날짜"]).reset_index(drop=True)
# =========================
# 비교용 계획 집계
# =========================
def add_midal_delta(plan_pivot: pd.DataFrame) -> pd.DataFrame:
    """
    같은 공정+품번 기준으로 날짜순 미달 증감 계산
    - 양수: 새 미달 발생
    - 0: 기존 미달 유지
    - 음수: 미달 해소
    """
    if plan_pivot.empty:
        plan_pivot["전일미달수량"] = 0
        plan_pivot["미달증감"] = 0
        return plan_pivot

    df = plan_pivot.copy()
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.normalize()

    df = df.sort_values(["공정", "품번", "날짜"]).reset_index(drop=True)

    df["전일미달수량"] = (
        df.groupby(["공정", "품번"])["미달수량"]
          .shift(1)
          .fillna(0)
    )

    df["미달증감"] = df["미달수량"] - df["전일미달수량"]

    return df


def build_plan_compare_base(plan_df: pd.DataFrame) -> pd.DataFrame:
    """
    날짜별 비교용 집계
    - 공정 + 날짜 + 품번 기준
    - 미달 / 계획 / 실적 각각 합계
    - 잔량판단합계 = 미달 + 계획 + 실적

    해석:
    - 작업지시가 현재 MES에서 안 보이더라도
      미달+계획+실적 == 0 이면
      '작업지시가 있었는데 완료되어 MES에서 사라진 것'으로 추정
    """
    if plan_df.empty:
        return pd.DataFrame(columns=[
            "공정", "날짜", "품번",
            "미달수량", "계획수량", "실적수량", "잔량판단합계"
        ])

    df = plan_df.copy()
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.date
    df["품번"] = df["품번"].astype(str).str.strip()
    df["구분"] = df["구분"].astype(str).str.strip()
    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0)

    grouped = (
        df.groupby(["공정", "날짜", "품번", "구분"], as_index=False)
          .agg(수량=("수량", "sum"))
    )

    pivot = grouped.pivot_table(
        index=["공정", "날짜", "품번"],
        columns="구분",
        values="수량",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    for col in ["미달", "계획", "실적"]:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot = pivot.rename(columns={
        "미달": "미달수량",
        "계획": "계획수량",
        "실적": "실적수량",
    })

    pivot["잔량판단합계"] = (
        pivot["미달수량"] + pivot["계획수량"] + pivot["실적수량"]
    )

    return pivot.sort_values(["공정", "날짜", "품번"]).reset_index(drop=True)

# =========================
# 비교
# =========================




def compare_plan_vs_mes_detail(sheet1_df: pd.DataFrame, sheet2_df: pd.DataFrame) -> pd.DataFrame:
    """
    1번 시트(계획 원본)를 그대로 사용하면서
    같은 공정+품번+날짜 기준으로 MES 지시량/실적량을 붙여 비교.

    추가 판단:
    - MES 작업지시수량 == 0 이어도
      같은 키의 (미달 + 계획 + 실적) == 0 이면
      '완료후MES소멸추정' 으로 판정
    """

    plan = sheet1_df.copy()
    mes = sheet2_df.copy()

    # ---------------------------
    # 1) 키/형식 정리
    # ---------------------------
    for col in ["공정", "품번"]:
        plan[col] = (
            plan[col].astype(str)
            .str.strip()
            .str.replace(r"\s+", " ", regex=True)
            .str.upper()
        )
        mes[col] = (
            mes[col].astype(str)
            .str.strip()
            .str.replace(r"\s+", " ", regex=True)
            .str.upper()
        )

    plan["날짜"] = pd.to_datetime(plan["날짜"], errors="coerce").dt.normalize()
    mes["날짜"] = pd.to_datetime(mes["날짜"], errors="coerce").dt.normalize()

    plan["구분"] = plan["구분"].astype(str).str.strip()
    plan["수량"] = pd.to_numeric(plan["수량"], errors="coerce").fillna(0)
    mes["지시량"] = pd.to_numeric(mes["지시량"], errors="coerce").fillna(0)
    mes["실적량"] = pd.to_numeric(mes["실적량"], errors="coerce").fillna(0)

    # ---------------------------
    # 2) 계획 원본에서 날짜별 요약 생성
    #    (미달/계획/실적 합계)
    # ---------------------------
    plan_summary = (
        plan.groupby(["공정", "품번", "날짜", "구분"], as_index=False)
            .agg(수량=("수량", "sum"))
    )

    plan_pivot = plan_summary.pivot_table(
        index=["공정", "품번", "날짜"],
        columns="구분",
        values="수량",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    for col in ["미달", "계획", "실적"]:
        if col not in plan_pivot.columns:
            plan_pivot[col] = 0

    plan_pivot = plan_pivot.rename(columns={
        "미달": "미달수량",
        "계획": "계획수량_합계",
        "실적": "실적수량_합계",
    })

    plan_pivot["잔량판단합계"] = (
        plan_pivot["미달수량"] +
        plan_pivot["계획수량_합계"] +
        plan_pivot["실적수량_합계"]
    )
    plan_pivot = add_midal_delta(plan_pivot)

    # ---------------------------
    # 3) MES 집계
    # ---------------------------
    mes_grouped = (
        mes.groupby(["공정", "품번", "날짜"], as_index=False)
           .agg(
               작업지시수량=("지시량", "sum"),
               MES실적량_원본=("실적량", "sum")
           )
    )

    # ---------------------------
    # 4) 1번 시트 원본행에 MES 집계 붙이기
    # ---------------------------
    result = plan.merge(
        mes_grouped,
        on=["공정", "품번", "날짜"],
        how="left"
    )

    result["날짜"] = pd.to_datetime(result["날짜"], errors="coerce").dt.normalize()
    plan_pivot["날짜"] = pd.to_datetime(plan_pivot["날짜"], errors="coerce").dt.normalize()

    result = result.merge(
    plan_pivot[
        ["공정", "품번", "날짜",
         "미달수량", "계획수량_합계", "실적수량_합계",
         "잔량판단합계", "전일미달수량", "미달증감"]
    ],
    on=["공정", "품번", "날짜"],
    how="left"
)

    for col in ["작업지시수량", "MES실적량_원본", "미달수량", "계획수량_합계", "실적수량_합계", "잔량판단합계", "전일미달수량", "미달증감"]:
        result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0)

    # ---------------------------
    # 5) 표시 컬럼 생성
    # ---------------------------
    # 계획행일 때만 작업지시수량 표시
    result["작업지시수량"] = result.apply(
        lambda x: x["작업지시수량"] if x["구분"] == "계획" else 0,
        axis=1
    )

    # 실적행은 계획시트 실적 수량을 표시
    result["MES실적량"] = result.apply(
        lambda x: x["수량"] if x["구분"] == "실적" else 0,
        axis=1
    )

    # 계획행 차이 = 계획행 수량 - 현재 MES 작업지시수량
    result["차이"] = result.apply(
        lambda x: x["계획수량_합계"] - x["작업지시수량"] if x["구분"] == "계획" else 0,
        axis=1
    )

    # ---------------------------
    # 6) 판정
    # ---------------------------
    def judge(row):
        
        plan_qty = float(row.get("계획수량_합계", 0))
        work_qty = float(row.get("작업지시수량", 0))
        actual_qty = float(row.get("실적수량_합계", 0))
        midal_qty = float(row.get("미달수량", 0))

        diff_qty = plan_qty - work_qty

        # 1) 당일 작업지시가 있으면 먼저 작업지시 기준 판정
        if work_qty > 0:
            if diff_qty == 0:
                return "일치"
            elif diff_qty > 0:
                return "해당날짜_작업지시수량부족"
            else:
                return "해당날짜_작업지시수량과다"

        # 2) 작업지시가 없을 때만 완료 추정 판단
        # 실적이 계획 이상이면 완료
        if actual_qty >= plan_qty and plan_qty > 0:
            return "완료후MES소멸추정"

        # 미달 음수는 선생산으로 보고 완료 처리
        if midal_qty < 0:
            return "완료후MES소멸추정"

        # 미달 양수는 진짜 미달
        if midal_qty > 0:
            return "신규미달_작업지시확인필요"

        # 아무 근거 없으면 작업지시 없음
        return "해당날짜_작업지시없음"

    result["판정"] = result.apply(judge, axis=1)

    # ---------------------------
    # 7) 계획에 없고 MES에만 있는 행 찾기
    # ---------------------------
    plan_keys = (
        plan[plan["구분"] == "계획"][["공정", "품번", "날짜"]]
        .drop_duplicates()
        .assign(계획존재=1)
    )

    mes_only = mes_grouped.merge(
        plan_keys,
        on=["공정", "품번", "날짜"],
        how="left"
    )

    mes_only = mes_only[mes_only["계획존재"].isna()].copy()

    if not mes_only.empty:
        mes_only["시트명"] = "MES추가"
        mes_only["모델"] = ""
        mes_only["구분"] = "계획없음"
        mes_only["수량"] = 0
        mes_only["미달수량"] = 0
        mes_only["계획수량_합계"] = 0
        mes_only["실적수량_합계"] = 0
        mes_only["잔량판단합계"] = 0
        mes_only["MES실적량"] = 0
        mes_only["차이"] = 0 - mes_only["작업지시수량"]
        mes_only["판정"] = "해당날짜_계획없는데작업지시있음"

        mes_only = mes_only[
            ["시트명", "공정", "모델", "품번", "날짜", "구분", "수량",
             "미달수량", "계획수량_합계", "실적수량_합계", "잔량판단합계",
             "작업지시수량", "MES실적량", "차이", "판정"]
        ]
    else:
        mes_only = pd.DataFrame(columns=[
            "시트명", "공정", "모델", "품번", "날짜", "구분", "수량",
            "미달수량", "계획수량_합계", "실적수량_합계", "잔량판단합계",
            "작업지시수량", "MES실적량", "차이", "판정"
        ])

    # ---------------------------
    # 8) 최종 컬럼 정리
    # ---------------------------
    result = result.drop(columns=["MES실적량_원본"], errors="ignore")

    cols = [
        "시트명", "공정", "모델", "품번", "날짜", "구분", "수량",
        "미달수량", "계획수량_합계", "실적수량_합계", "잔량판단합계",
        "작업지시수량", "MES실적량", "차이", "판정"
    ]
    result = result[[c for c in cols if c in result.columns]]

    final_result = pd.concat([result, mes_only], ignore_index=True)

    final_result["날짜"] = pd.to_datetime(final_result["날짜"], errors="coerce")
    final_result = final_result.sort_values(["공정", "품번", "날짜", "구분"]).reset_index(drop=True)
    final_result["날짜"] = final_result["날짜"].dt.strftime("%Y-%m-%d")

    return final_result
# =========================
# 저장
# =========================
def save_results(plan_df, mes_df, plan_base_df, compare_df, output_file: str):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        plan_df.to_excel(writer, sheet_name="1_계획정규화", index=False)
        mes_df.to_excel(writer, sheet_name="2_MES정규화", index=False)
        plan_base_df.to_excel(writer, sheet_name="3_계획집계", index=False)
        compare_df.to_excel(writer, sheet_name="4_비교결과", index=False)

        if not compare_df.empty:
            summary = (
                compare_df.groupby("판정", as_index=False)
                          .size()
                          .rename(columns={"size": "건수"})
            )
            summary.to_excel(writer, sheet_name="5_요약", index=False)

def main():
    plan_df = extract_plan_all(PLAN_FILE)
    print(f"[계획 정규화] {len(plan_df)}건")

    mes_df = extract_mes(MES_FILE)
    print(f"[MES 정규화] {len(mes_df)}건")

    plan_base_df = build_plan_compare_base(plan_df)
    print(f"[계획 비교기준] {len(plan_base_df)}건")

    compare_df = compare_plan_vs_mes_detail(plan_df, mes_df)
    print(f"[비교 결과] {len(compare_df)}건")

    save_results(plan_df, mes_df, plan_base_df, compare_df, OUTPUT_FILE)
    print(f"저장 완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()