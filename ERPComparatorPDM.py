from __future__ import annotations

import argparse
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from CommonUtils import create_progress_window
from ERPComparator import (
    HEADER_FILL,
    MISSING_ERP_FILL,
    MISSING_QTY_FILL,
    OUTPUT_HEADERS,
    OUTPUT_SHEET_NAME,
    BomUploadRow,
    split_path,
    unit_for_part,
)
from ERPPartsDb import (
    DEFAULT_ERP_DB,
    load_erp_part_numbers_from_db,
    normalize_part_lookup,
    refresh_erp_db_from_excel,
)
from PdmBomParser import PdmBomItem, parse_pdm_bom


def numeric_level(item: PdmBomItem) -> int | None:
    try:
        return int(float(item.level))
    except (TypeError, ValueError):
        return None


def build_pdm_bom_upload_rows(
    items: list[PdmBomItem],
    erp_part_numbers: set[str] | None = None,
) -> list[BomUploadRow]:
    rows: list[BomUploadRow] = []

    for item in items:
        level = numeric_level(item)
        path_parts = split_path(item.path)
        if level is not None and level <= 0:
            continue
        if len(path_parts) < 2:
            continue

        child_part = str(item.part_number or path_parts[-1]).strip()
        if not child_part:
            continue

        quantity = float(item.quantity or 0)
        child_lookup = normalize_part_lookup(child_part)
        rows.append(
            BomUploadRow(
                parent_part=path_parts[-2],
                child_part=child_part,
                parent_qty=1,
                quantity=quantity,
                unit=unit_for_part(child_part),
                needs_quantity=quantity == 0,
                is_registered_in_erp=erp_part_numbers is None or child_lookup in erp_part_numbers,
            )
        )

    return rows


def write_bom_upload_file(output_file: str | Path, rows: list[BomUploadRow]) -> str:
    workbook = Workbook()
    ws = workbook.active
    ws.title = OUTPUT_SHEET_NAME

    for col_index, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = Font(name="맑은 고딕", size=10, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(rows, start=2):
        values = [row.parent_part, row.child_part, row.parent_qty, row.quantity, row.unit]
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="center")

        if row.needs_quantity:
            qty_cell = ws.cell(row=row_index, column=4)
            qty_cell.fill = MISSING_QTY_FILL
            qty_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")

        if not row.is_registered_in_erp:
            child_cell = ws.cell(row=row_index, column=2)
            child_cell.fill = MISSING_ERP_FILL
            child_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="9C0006")

    for col_index, width in enumerate([18, 18, 12, 12, 10], start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    workbook.save(output_file)
    workbook.close()
    return OUTPUT_SHEET_NAME


def default_output_path(input_file: str | Path) -> str:
    source = Path(input_file)
    return str(source.with_name(f"{source.stem}_PDM_BOM업로드.xlsx"))


def run(
    input_file: str | Path,
    output_file: str | Path,
    sheet_name: str | None = None,
    erp_file: str | Path | None = None,
    db_path: str | Path = DEFAULT_ERP_DB,
    refresh_erp: bool = False,
    progress=None,
) -> tuple[int, int, int, str]:
    if progress:
        progress(10, "PDM BOM List 추출 중...")
    items = parse_pdm_bom(input_file, sheet_name)

    if refresh_erp and not erp_file:
        raise ValueError("ERP DB를 갱신하려면 ERP 품목상품 엑셀 파일이 필요합니다.")

    if erp_file and refresh_erp:
        if progress:
            progress(20, "ERP DB 갱신 중...")
        refresh_erp_db_from_excel(erp_file, db_path, progress=progress)

    if progress:
        progress(50, "ERP DB 확인 중...")
    erp_part_numbers = load_erp_part_numbers_from_db(db_path)

    if progress:
        progress(65, "BOM 업로드 시트 데이터 생성 중...")
    rows = build_pdm_bom_upload_rows(items, erp_part_numbers)
    missing_count = sum(1 for row in rows if row.needs_quantity)
    missing_erp_count = sum(1 for row in rows if not row.is_registered_in_erp)

    if progress:
        progress(85, "결과 파일 저장 중...")
    result_sheet_name = write_bom_upload_file(output_file, rows)

    if progress:
        progress(100, "완료")
    return len(rows), missing_count, missing_erp_count, result_sheet_name


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="PDM BOM List를 ERP BOM 업로드 양식으로 변환합니다.")
    parser.add_argument("xlsx_path", nargs="?", help="PDM BOM List 엑셀 파일 경로")
    parser.add_argument("--sheet", help="읽을 시트명. 생략하면 첫 번째 활성 시트를 읽습니다.")
    parser.add_argument("--erp", help="ERP DB 갱신에 사용할 품목상품 엑셀 파일 경로")
    parser.add_argument("--erp-db", default=str(DEFAULT_ERP_DB), help="ERP 품목 DB 경로")
    parser.add_argument("--refresh-erp", action="store_true", help="ERP 엑셀을 다시 읽어서 DB를 갱신합니다.")
    parser.add_argument("--output", "-o", help="결과 엑셀 파일 경로")
    return parser


def select_pdm_file(parent) -> str:
    return filedialog.askopenfilename(
        parent=parent,
        title="PDM BOM List 엑셀 파일 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
    )


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if args.xlsx_path:
        input_file = args.xlsx_path
        output_file = args.output or default_output_path(input_file)
        row_count, missing_count, missing_erp_count, result_sheet_name = run(
            input_file,
            output_file,
            args.sheet,
            args.erp,
            args.erp_db,
            args.refresh_erp or bool(args.erp),
        )
        print(f"Saved: {output_file}")
        print(f"Sheet: {result_sheet_name}")
        print(f"Rows: {row_count}")
        print(f"Missing quantity rows: {missing_count}")
        print(f"Missing ERP part rows: {missing_erp_count}")
        return

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    try:
        input_file = select_pdm_file(root)
        if not input_file:
            return

        erp_file = None
        refresh_erp = False
        if not Path(DEFAULT_ERP_DB).exists():
            messagebox.showinfo(
                "ERP DB 생성",
                "ERP 품목 DB가 없어 처음 한 번 ERP 품목상품 엑셀 파일을 읽어옵니다.",
                parent=root,
            )
            erp_file = filedialog.askopenfilename(
                parent=root,
                title="ERP 품목상품 엑셀 파일 선택",
                filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
            )
            if not erp_file:
                return
            refresh_erp = True

        output_file = filedialog.asksaveasfilename(
            parent=root,
            title="PDM BOM 업로드 파일 저장",
            initialdir=str(Path(input_file).parent),
            initialfile=Path(default_output_path(input_file)).name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not output_file:
            return

        progress_win, update_progress = create_progress_window(root, "PDM BOM 업로드 시트 생성")
        try:
            row_count, missing_count, missing_erp_count, result_sheet_name = run(
                input_file,
                output_file,
                erp_file=erp_file,
                refresh_erp=refresh_erp,
                progress=update_progress,
            )
        finally:
            progress_win.destroy()

        messagebox.showinfo(
            "완료",
            f"결과 파일 생성 완료\n\n{output_file}\n\n"
            f"생성 시트: {result_sheet_name}\n"
            f"생성 행 수: {row_count}\n"
            f"수량 확인 필요: {missing_count}\n"
            f"ERP 미등록 품번: {missing_erp_count}",
            parent=root,
        )
    except Exception as exc:
        messagebox.showerror("오류", str(exc), parent=root)
        raise
    finally:
        root.destroy()


if __name__ == "__main__":
    main()
