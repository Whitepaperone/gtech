from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sqlite3
from typing import Any

from openpyxl import load_workbook

from ChangeListParser import cell_text

try:
    import pythoncom
    import win32com.client
except ImportError:  # pragma: no cover - fallback for PCs without Excel COM support.
    pythoncom = None
    win32com = None


DEFAULT_ERP_DB = Path(__file__).with_name("erp_parts.db")
ERP_DATA_START_ROW = 2
ERP_FIXED_COLUMNS = {
    "part_number": 2,  # B
    "category": 3,  # C
    "name": 4,  # D
    "spec": 5,  # E
    "supplier": 7,  # G
    "unit": 10,  # J
    "min_purchase_qty": 15,  # O
}


def normalize_part_lookup(value: Any) -> str:
    return (
        cell_text(value)
        .upper()
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )


def create_erp_db(db_path: str | Path = DEFAULT_ERP_DB) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS erp_parts (
                part_key TEXT PRIMARY KEY,
                part_number TEXT NOT NULL,
                category TEXT,
                name TEXT,
                spec TEXT,
                supplier TEXT,
                unit TEXT,
                min_purchase_qty TEXT,
                source_file TEXT,
                updated_at TEXT
            )
            """
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_erp_parts_part_number ON erp_parts(part_number)")


def fixed_cell_value(row_values: tuple[Any, ...], field_name: str) -> str:
    column = ERP_FIXED_COLUMNS.get(field_name)
    if not column or len(row_values) < column:
        return ""
    return cell_text(row_values[column - 1])


def normalize_com_range_values(values: Any) -> list[tuple[Any, ...]]:
    if values is None:
        return []
    if not isinstance(values, tuple):
        return [(values,)]
    if not values:
        return []
    if not isinstance(values[0], tuple):
        return [values]
    return list(values)


def read_erp_rows_with_excel_com(erp_file: str | Path, progress=None) -> list[tuple[str, str, str, str, str, str, str, str, str, str]]:
    if pythoncom is None or win32com is None:
        raise RuntimeError("Excel COM을 사용할 수 없습니다.")

    source_file = str(Path(erp_file).resolve())
    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    pythoncom.CoInitialize()
    excel = None
    workbook = None

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(source_file, ReadOnly=True)
        ws = workbook.Worksheets(1)
        last_row = ws.Cells(ws.Rows.Count, ERP_FIXED_COLUMNS["part_number"]).End(-4162).Row

        if last_row < ERP_DATA_START_ROW:
            return []

        if progress:
            progress(20, f"Excel 빠른 읽기 중... {last_row:,}행")

        main_values = normalize_com_range_values(ws.Range(f"B{ERP_DATA_START_ROW}:E{last_row}").Value)
        supplier_values = normalize_com_range_values(ws.Range(f"G{ERP_DATA_START_ROW}:G{last_row}").Value)
        unit_values = normalize_com_range_values(ws.Range(f"J{ERP_DATA_START_ROW}:J{last_row}").Value)
        min_qty_values = normalize_com_range_values(ws.Range(f"O{ERP_DATA_START_ROW}:O{last_row}").Value)

        rows = []
        for index, main_row in enumerate(main_values):
            part_number = cell_text(main_row[0] if len(main_row) > 0 else "")
            part_key = normalize_part_lookup(part_number)
            if not part_key:
                continue

            rows.append(
                (
                    part_key,
                    part_number,
                    cell_text(main_row[1] if len(main_row) > 1 else ""),
                    cell_text(main_row[2] if len(main_row) > 2 else ""),
                    cell_text(main_row[3] if len(main_row) > 3 else ""),
                    cell_text(supplier_values[index][0] if index < len(supplier_values) else ""),
                    cell_text(unit_values[index][0] if index < len(unit_values) else ""),
                    cell_text(min_qty_values[index][0] if index < len(min_qty_values) else ""),
                    source_file,
                    updated_at,
                )
            )

        return rows
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def read_erp_rows_with_openpyxl(erp_file: str | Path, progress=None) -> list[tuple[str, str, str, str, str, str, str, str, str, str]]:
    workbook = load_workbook(erp_file, data_only=True, read_only=True)
    rows: list[tuple[str, str, str, str, str, str, str, str, str, str]] = []
    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws = workbook.worksheets[0]
    if progress:
        progress(20, f"ERP 시트 읽는 중... {ws.title}")

    for row_index, row_values in enumerate(
        ws.iter_rows(
            min_row=ERP_DATA_START_ROW,
            max_col=max(ERP_FIXED_COLUMNS.values()),
            values_only=True,
        ),
        start=ERP_DATA_START_ROW,
    ):
        if progress and row_index % 5000 == 0:
            progress(20, f"ERP 품번 읽는 중... {row_index:,}행")

        part_number = fixed_cell_value(row_values, "part_number")
        part_key = normalize_part_lookup(part_number)
        if not part_key:
            continue

        rows.append(
            (
                part_key,
                part_number,
                fixed_cell_value(row_values, "category"),
                fixed_cell_value(row_values, "name"),
                fixed_cell_value(row_values, "spec"),
                fixed_cell_value(row_values, "supplier"),
                fixed_cell_value(row_values, "unit"),
                fixed_cell_value(row_values, "min_purchase_qty"),
                str(Path(erp_file)),
                updated_at,
            )
        )

    workbook.close()
    return rows


def save_erp_db_from_excel(
    erp_file: str | Path,
    db_path: str | Path = DEFAULT_ERP_DB,
    replace_all: bool = True,
    progress=None,
) -> int:
    create_erp_db(db_path)
    try:
        rows = read_erp_rows_with_excel_com(erp_file, progress=progress)
    except Exception:
        rows = read_erp_rows_with_openpyxl(erp_file, progress=progress)

    if not rows:
        raise ValueError("ERP 품목상품 파일에서 품번/품목코드 열을 찾지 못했습니다.")

    with sqlite3.connect(db_path) as conn:
        if replace_all:
            conn.execute("DELETE FROM erp_parts")
        conn.executemany(
            """
            INSERT OR REPLACE INTO erp_parts (
                part_key,
                part_number,
                category,
                name,
                spec,
                supplier,
                unit,
                min_purchase_qty,
                source_file,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )

    return len(rows)


def refresh_erp_db_from_excel(
    erp_file: str | Path,
    db_path: str | Path = DEFAULT_ERP_DB,
    progress=None,
) -> int:
    return save_erp_db_from_excel(erp_file, db_path, replace_all=True, progress=progress)


def add_erp_parts_from_excel(
    erp_file: str | Path,
    db_path: str | Path = DEFAULT_ERP_DB,
    progress=None,
) -> int:
    return save_erp_db_from_excel(erp_file, db_path, replace_all=False, progress=progress)


def load_erp_part_numbers_from_db(db_path: str | Path = DEFAULT_ERP_DB) -> set[str]:
    if not Path(db_path).exists():
        raise FileNotFoundError(f"ERP DB가 없습니다: {db_path}")

    create_erp_db(db_path)
    with sqlite3.connect(db_path) as conn:
        rows = conn.execute("SELECT part_key FROM erp_parts").fetchall()

    part_numbers = {row[0] for row in rows if row[0]}
    if not part_numbers:
        raise ValueError("ERP DB에 저장된 품번이 없습니다. ERP 품목상품 파일로 DB를 갱신해주세요.")
    return part_numbers
