from __future__ import annotations

from datetime import date
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkcalendar import Calendar

from AppConstants import (
    DATA_START_ROW,
    DATE_HEADER_GAP_BREAK,
    HEADER_DATE_ROW,
    HEADER_KIND_ROW,
    ORDER_ZONE_FIXED_COLUMNS,
)
from CommonUtils import (
    build_merged_map,
    create_progress_window,
    normalize_kind,
    normalize_part_no,
    normalize_process_token,
    safe_float,
)
from ExtractPlan import (
    build_header_lookup,
    find_date_start_col,
    is_skip_summary_row,
    should_skip_row,
    starts_with_alpha_part_no,
)


OUTPUT_SUFFIX = "수주구간표시"
PRESERVE_UNTIL_COL = 54  # BB열까지는 원본 서식을 유지한다.
NEXT_MONTH_FILL = PatternFill("solid", fgColor="FFD966")
SECOND_NEXT_MONTH_FILL = PatternFill("solid", fgColor="92D050")
NO_FILL = PatternFill(fill_type=None)
CALC_KINDS = {"미달", "계획"}


def build_all_date_columns(ws) -> list[dict]:
    """날짜 영역의 미달/계획/실적 열을 모두 수집한다."""
    start_col = find_date_start_col(ws)
    if start_col is None:
        raise RuntimeError(f"{ws.title}: 날짜 시작 열을 찾지 못했습니다.")

    columns = []
    started = False
    invalid_streak = 0

    for col in range(start_col, ws.max_column + 1):
        date_value = ws.cell(HEADER_DATE_ROW, col).value
        kind = normalize_kind(ws.cell(HEADER_KIND_ROW, col).value)

        if pd.notna(pd.to_datetime(date_value, errors="coerce")) and date_value:
            started = True
            invalid_streak = 0
            if kind in {"미달", "계획", "실적"}:
                columns.append(
                    {
                        "col": col,
                        "date": pd.to_datetime(date_value).normalize(),
                        "kind": kind,
                    }
                )
            continue

        if started:
            invalid_streak += 1
            if invalid_streak >= DATE_HEADER_GAP_BREAK:
                break

    return columns


def pick_part_no(row_values: list, cols: dict) -> str:
    """기존 생산계획 구조의 품번 열 우선순위로 품번을 고른다."""
    for key in ("finish_part_col", "tank_part_col", "core_part_col", "customer_part_col"):
        col = cols.get(key)
        if col:
            part_no = normalize_part_no(row_values[col - 1])
            if part_no:
                return part_no
    return ""


def parse_sheet_records(ws, start_date) -> list[dict]:
    """계산 전 원본 엑셀을 컴퓨터가 읽기 좋은 record 목록으로 변환한다."""
    date_columns = build_all_date_columns(ws)
    cols = build_header_lookup(ws)
    merged_map = build_merged_map(ws)
    records = []

    for row_cells in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, max_col=ws.max_column):
        row_number = row_cells[0].row
        row_values = [cell.value for cell in row_cells]
        part_no = pick_part_no(row_values, cols)

        row_text = " ".join(
            text
            for text in (normalize_process_token(cell.value) for cell in row_cells)
            if text
        )
        first_col_text = normalize_process_token(merged_map.get((row_number, 1), row_values[0]))

        if not part_no:
            continue
        if should_skip_row(part_no, "", row_text, sheet_title=ws.title, product_key=part_no):
            continue
        if is_skip_summary_row(part_no, row_text, first_col_text):
            continue
        if not starts_with_alpha_part_no(part_no):
            continue

        sales = safe_float(row_values[ORDER_ZONE_FIXED_COLUMNS["sales_col"] - 1])
        stock = safe_float(row_values[ORDER_ZONE_FIXED_COLUMNS["stock_col"] - 1])
        current_order = safe_float(row_values[ORDER_ZONE_FIXED_COLUMNS["current_month_order_col"] - 1])
        next_order = safe_float(row_values[ORDER_ZONE_FIXED_COLUMNS["next_month_order_col"] - 1])
        second_next_order = safe_float(row_values[ORDER_ZONE_FIXED_COLUMNS["second_next_month_order_col"] - 1])

        for date_col in date_columns:
            qty = safe_float(row_values[date_col["col"] - 1])
            records.append(
                {
                    "시트명": ws.title,
                    "품번": part_no,
                    "원본행": row_number,
                    "원본열": date_col["col"],
                    "날짜": date_col["date"],
                    "구분": date_col["kind"],
                    "수량": qty,
                    "판매": sales,
                    "재고": stock,
                    "이번달수주": current_order,
                    "다음달수주": next_order,
                    "다다음달수주": second_next_order,
                    "시작일": pd.to_datetime(start_date).normalize(),
                }
            )

    return records


def parse_plan_records(plan_file: str, start_date, progress=None) -> pd.DataFrame:
    """값 계산용 workbook은 data_only=True로 열어 수식 결과값을 읽는다."""
    if progress:
        progress(10, "계산용 엑셀 값을 읽는 중...")

    workbook = load_workbook(plan_file, data_only=True)
    records = []
    sheet_names = workbook.sheetnames[:1]
    total = max(len(sheet_names), 1)

    for index, sheet_name in enumerate(sheet_names, start=1):
        if progress:
            progress(10 + int((index - 1) / total * 35), f"{sheet_name} 파싱 중...")
        records.extend(parse_sheet_records(workbook[sheet_name], start_date))
        if progress:
            progress(10 + int(index / total * 35), f"{sheet_name} 파싱 완료")

    return pd.DataFrame(records)


def first_calc_column(group: pd.DataFrame, start_date) -> int | None:
    """시작일의 미달 열을 우선 찾고, 없으면 시작일 이후 첫 날짜 열을 사용한다."""
    start_date = pd.to_datetime(start_date).normalize()
    candidates = group[group["날짜"] >= start_date].sort_values(["날짜", "원본열"])
    if candidates.empty:
        return None

    same_day_shortage = candidates[(candidates["날짜"] == start_date) & (candidates["구분"] == "미달")]
    if not same_day_shortage.empty:
        return int(same_day_shortage.iloc[0]["원본열"])

    return int(candidates.iloc[0]["원본열"])


def classify_month(cumulative: float, current_order: float, next_order: float, second_next_order: float) -> str:
    """누적값 기준으로 이번달/다음달/다다음달 수주 구간을 판정한다."""
    current_limit = current_order
    next_limit = current_order + next_order
    second_next_limit = current_order + next_order + second_next_order

    if cumulative <= current_limit:
        return "이번달"
    if cumulative <= next_limit:
        return "다음달"
    if cumulative <= second_next_limit:
        return "다다음달"
    return "수주초과"


def calculate_order_zones(records_df: pd.DataFrame, start_date) -> pd.DataFrame:
    """파싱된 records를 기준으로 품번별 누적 수량과 수주 구간을 계산한다."""
    if records_df.empty:
        return pd.DataFrame(columns=["시트명", "품번", "원본행", "원본열", "판정구간", "누적수량"])

    result_rows = []
    group_cols = ["시트명", "품번", "원본행"]

    for _, group in records_df.groupby(group_cols, sort=False):
        start_col = first_calc_column(group, start_date)
        if start_col is None:
            continue

        work = group[
            (group["원본열"] >= start_col)
            & (group["구분"].isin(CALC_KINDS))
            & (group["수량"] != 0)
        ].sort_values(["날짜", "원본열"])
        if work.empty:
            continue

        first = work.iloc[0]
        cumulative = safe_float(first["판매"]) + safe_float(first["재고"])
        current_order = safe_float(first["이번달수주"])
        next_order = safe_float(first["다음달수주"])
        second_next_order = safe_float(first["다다음달수주"])

        for row in work.itertuples(index=False):
            cumulative += safe_float(getattr(row, "수량"))
            result_rows.append(
                {
                    "시트명": getattr(row, "시트명"),
                    "품번": getattr(row, "품번"),
                    "원본행": int(getattr(row, "원본행")),
                    "원본열": int(getattr(row, "원본열")),
                    "판정구간": classify_month(cumulative, current_order, next_order, second_next_order),
                    "누적수량": cumulative,
                }
            )

    return pd.DataFrame(result_rows)


def clear_background_after_preserve_col(ws) -> None:
    """BB열 이후 날짜 영역의 기존 배경색을 제거해 새 색상이 잘 보이게 한다."""
    if ws.max_column <= PRESERVE_UNTIL_COL:
        return

    for row in ws.iter_rows(
        min_row=1,
        max_row=ws.max_row,
        min_col=PRESERVE_UNTIL_COL + 1,
        max_col=ws.max_column,
    ):
        for cell in row:
            cell.fill = NO_FILL


def color_original_workbook(input_file: str, output_file: str, results_df: pd.DataFrame) -> int:
    """계산 결과를 기준으로 저장용 workbook의 원본 셀 위치에 색을 칠한다."""
    workbook = load_workbook(input_file, data_only=False)
    first_sheet_name = workbook.worksheets[0].title
    clear_background_after_preserve_col(workbook[first_sheet_name])
    colored_count = 0

    for row in results_df.itertuples(index=False):
        sheet_name = getattr(row, "시트명")
        if sheet_name != first_sheet_name:
            continue

        original_col = int(getattr(row, "원본열"))
        if original_col <= PRESERVE_UNTIL_COL:
            continue

        cell = workbook[sheet_name].cell(row=int(getattr(row, "원본행")), column=original_col)
        zone = getattr(row, "판정구간")
        if zone == "다음달":
            cell.fill = NEXT_MONTH_FILL
            colored_count += 1
        elif zone == "다다음달":
            cell.fill = SECOND_NEXT_MONTH_FILL
            colored_count += 1

    workbook.save(output_file)
    return colored_count


def default_output_path(input_file: str) -> str:
    source = Path(input_file)
    return str(source.with_name(f"{source.stem}_{OUTPUT_SUFFIX}.xlsx"))


def select_start_date(parent) -> pd.Timestamp:
    """시작일을 선택한다. 기본값은 오늘 날짜다."""
    selected = {"date": pd.Timestamp(date.today()).normalize()}
    win = tk.Toplevel(parent)
    win.title("시작일 선택")
    win.geometry("320x310")
    win.resizable(False, False)
    win.attributes("-topmost", True)

    tk.Label(win, text="계산 시작일").pack(pady=(12, 4))
    calendar = Calendar(win, selectmode="day", date_pattern="yyyy-mm-dd")
    calendar.selection_set(date.today())
    calendar.pack(padx=12, pady=8)

    def confirm():
        selected["date"] = pd.to_datetime(calendar.get_date()).normalize()
        win.destroy()

    def use_today():
        selected["date"] = pd.Timestamp(date.today()).normalize()
        win.destroy()

    button_frame = tk.Frame(win)
    button_frame.pack(pady=8)
    tk.Button(button_frame, text="선택일 사용", width=12, command=confirm).pack(side="left", padx=4)
    tk.Button(button_frame, text="오늘 사용", width=12, command=use_today).pack(side="left", padx=4)

    win.protocol("WM_DELETE_WINDOW", use_today)
    win.grab_set()
    parent.wait_window(win)
    return selected["date"]


def run(input_file: str, output_file: str, start_date, progress=None) -> tuple[int, int]:
    records_df = parse_plan_records(input_file, start_date, progress=progress)
    if progress:
        progress(55, "수주 구간을 계산하는 중...")
    results_df = calculate_order_zones(records_df, start_date)
    if progress:
        progress(75, "원본 엑셀 셀에 색을 칠하는 중...")
    colored_count = color_original_workbook(input_file, output_file, results_df)
    if progress:
        progress(100, "완료")
    return len(results_df), colored_count


def main() -> None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    try:
        input_file = filedialog.askopenfilename(
            parent=root,
            title="생산계획 엑셀 선택",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")],
        )
        if not input_file:
            return

        start_date = select_start_date(root)
        output_file = filedialog.asksaveasfilename(
            parent=root,
            title="수주 구간 표시 파일 저장",
            initialdir=str(Path(input_file).parent),
            initialfile=Path(default_output_path(input_file)).name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not output_file:
            return

        progress_win, update_progress = create_progress_window(root, "수주 구간 표시")
        try:
            result_count, colored_count = run(input_file, output_file, start_date, progress=update_progress)
        finally:
            progress_win.destroy()

        messagebox.showinfo(
            "완료",
            f"저장 위치:\n{output_file}\n\n계산 건수: {result_count}\n색칠 건수: {colored_count}",
            parent=root,
        )
    except Exception as exc:
        messagebox.showerror("오류", str(exc), parent=root)
    finally:
        root.destroy()


if __name__ == "__main__":
    main()
