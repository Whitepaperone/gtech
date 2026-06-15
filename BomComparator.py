from __future__ import annotations

import argparse
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ChangeListParser import ChangeListItem, parse_change_list
from CommonUtils import create_progress_window, select_excel_file
from PdmBomParser import PdmBomItem, parse_pdm_bom


@dataclass
class AggregatedItem:
    path: str
    part_number: str
    quantity: float
    levels: str


@dataclass
class ComparisonRow:
    path: str
    part_number: str
    status: str
    change_level: str
    change_quantity: float | None
    pdm_level: str
    pdm_quantity: float | None
    note: str


HEADER_ROW = [
    "BOM Path",
    "Part Number",
    "Status",
    "Change List Level",
    "Change List Quantity",
    "PDM BOM Level",
    "PDM BOM Quantity",
    "Note",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MATCH_FILL = PatternFill("solid", fgColor="E2F0D9")
MISMATCH_FILL = PatternFill("solid", fgColor="FFC7CE")
MISSING_FILL = PatternFill("solid", fgColor="FFD966")
PART_MISSING_FILL = PatternFill("solid", fgColor="F4B183")


def normalize_part_number(value: str) -> str:
    return str(value or "").strip().upper().replace(" ", "").replace("_", "")


def normalize_path(value: str) -> str:
    parts = [normalize_part_number(part) for part in str(value or "").split(">")]
    return " > ".join(part for part in parts if part)


def normalize_quantity(value: float | None) -> float:
    if value is None:
        return 0
    return float(value)


def number_equal(left: float | None, right: float | None) -> bool:
    return abs(normalize_quantity(left) - normalize_quantity(right)) < 0.000001


def compact_values(values: Iterable[str]) -> str:
    unique = sorted({str(value).strip() for value in values if str(value).strip()})
    return ", ".join(unique)


def aggregate_change_items(items: list[ChangeListItem]) -> dict[str, AggregatedItem]:
    aggregated: dict[str, AggregatedItem] = {}
    source_levels: dict[str, list[str]] = {}

    for item in items:
        key = normalize_path(item.path) or normalize_part_number(item.part_number)
        if not key:
            continue

        if key not in aggregated:
            aggregated[key] = AggregatedItem(
                path=item.path,
                part_number=item.part_number,
                quantity=0,
                levels="",
            )
            source_levels[key] = []

        aggregated[key].quantity += normalize_quantity(item.quantity)
        source_levels[key].append(item.level)

    for key, item in aggregated.items():
        item.levels = compact_values(source_levels[key])

    return aggregated


def aggregate_pdm_items(items: list[PdmBomItem]) -> dict[str, AggregatedItem]:
    aggregated: dict[str, AggregatedItem] = {}
    source_levels: dict[str, list[str]] = {}

    for item in items:
        key = normalize_path(item.path) or normalize_part_number(item.part_number)
        if not key:
            continue

        if key not in aggregated:
            aggregated[key] = AggregatedItem(
                path=item.path,
                part_number=item.part_number,
                quantity=0,
                levels="",
            )
            source_levels[key] = []

        aggregated[key].quantity += normalize_quantity(item.quantity)
        source_levels[key].append(item.level)

    for key, item in aggregated.items():
        item.levels = compact_values(source_levels[key])

    return aggregated


def compare_items(
    change_items: list[ChangeListItem],
    pdm_items: list[PdmBomItem],
) -> list[ComparisonRow]:
    change_by_part = aggregate_change_items(change_items)
    pdm_by_part = aggregate_pdm_items(pdm_items)
    matched_change_keys: set[str] = set()
    matched_pdm_keys: set[str] = set()

    rows: list[ComparisonRow] = []
    for key in sorted(set(change_by_part) & set(pdm_by_part)):
        change = change_by_part[key]
        pdm = pdm_by_part[key]
        matched_change_keys.add(key)
        matched_pdm_keys.add(key)
        status = "OK" if number_equal(change.quantity, pdm.quantity) else "Quantity Mismatch"
        note = "" if status == "OK" else "Quantity differs."

        rows.append(
            ComparisonRow(
                path=change.path,
                part_number=change.part_number,
                status=status,
                change_level=change.levels,
                change_quantity=change.quantity,
                pdm_level=pdm.levels,
                pdm_quantity=pdm.quantity,
                note=note,
            )
        )

    unmatched_pdm_by_part: dict[str, list[str]] = {}
    for key, pdm in pdm_by_part.items():
        if key in matched_pdm_keys:
            continue
        unmatched_pdm_by_part.setdefault(normalize_part_number(pdm.part_number), []).append(key)

    for change_key in sorted(set(change_by_part) - matched_change_keys):
        change = change_by_part[change_key]
        pdm_candidates = unmatched_pdm_by_part.get(normalize_part_number(change.part_number), [])

        if pdm_candidates:
            pdm_key = pdm_candidates.pop(0)
            pdm = pdm_by_part[pdm_key]
            matched_change_keys.add(change_key)
            matched_pdm_keys.add(pdm_key)
            status = "Path Mismatch" if number_equal(change.quantity, pdm.quantity) else "Path/Quantity Mismatch"
            rows.append(
                ComparisonRow(
                    path=f"Change: {change.path}\nPDM: {pdm.path}",
                    part_number=change.part_number,
                    status=status,
                    change_level=change.levels,
                    change_quantity=change.quantity,
                    pdm_level=pdm.levels,
                    pdm_quantity=pdm.quantity,
                    note="Same part number exists in both files, but BOM path is different.",
                )
            )
            continue

        rows.append(
            ComparisonRow(
                path=change.path,
                part_number=change.part_number,
                status="Missing in PDM BOM",
                change_level=change.levels,
                change_quantity=change.quantity,
                pdm_level="",
                pdm_quantity=None,
                note="Latest part number from Change List is not in PDM BOM.",
            )
        )

    for pdm_key in sorted(set(pdm_by_part) - matched_pdm_keys):
        pdm = pdm_by_part[pdm_key]
        rows.append(
            ComparisonRow(
                path=pdm.path,
                part_number=pdm.part_number,
                status="Missing in Change List",
                change_level="",
                change_quantity=None,
                pdm_level=pdm.levels,
                pdm_quantity=pdm.quantity,
                note="Part number exists only in PDM BOM.",
            )
        )

    return rows


def add_parsed_items_sheet(workbook, title: str, items) -> None:
    ws = workbook.create_sheet(title)
    ws.append(["Level", "BOM Path", "Part Number", "Quantity"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        ws.append([item.level, item.path, item.part_number, item.quantity])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, width in enumerate([12, 64, 26, 18], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def write_comparison_excel(
    rows: list[ComparisonRow],
    output_path: str | Path,
    change_items: list[ChangeListItem] | None = None,
    pdm_items: list[PdmBomItem] | None = None,
) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "BOM Compare"

    ws.append(HEADER_ROW)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(
            [
                row.path,
                row.part_number,
                row.status,
                row.change_level,
                row.change_quantity,
                row.pdm_level,
                row.pdm_quantity,
                row.note,
            ]
        )

    for row_index in range(2, ws.max_row + 1):
        status = ws.cell(row=row_index, column=3).value
        change_quantity = ws.cell(row=row_index, column=5).value
        pdm_quantity = ws.cell(row=row_index, column=7).value

        if status == "OK":
            ws.cell(row=row_index, column=3).fill = MATCH_FILL
        elif status == "Quantity Mismatch":
            ws.cell(row=row_index, column=3).fill = MISMATCH_FILL
            ws.cell(row=row_index, column=5).fill = MISMATCH_FILL
            ws.cell(row=row_index, column=7).fill = MISMATCH_FILL
        elif status in {"Path Mismatch", "Path/Quantity Mismatch"}:
            ws.cell(row=row_index, column=1).fill = MISMATCH_FILL
            ws.cell(row=row_index, column=3).fill = MISMATCH_FILL
            if status == "Path/Quantity Mismatch":
                ws.cell(row=row_index, column=5).fill = MISMATCH_FILL
                ws.cell(row=row_index, column=7).fill = MISMATCH_FILL
        elif status == "Missing in PDM BOM":
            for column in range(1, len(HEADER_ROW) + 1):
                ws.cell(row=row_index, column=column).fill = MISSING_FILL
            ws.cell(row=row_index, column=6).fill = PART_MISSING_FILL
            ws.cell(row=row_index, column=7).fill = PART_MISSING_FILL
        elif status == "Missing in Change List":
            for column in range(1, len(HEADER_ROW) + 1):
                ws.cell(row=row_index, column=column).fill = MISSING_FILL
            ws.cell(row=row_index, column=4).fill = PART_MISSING_FILL
            ws.cell(row=row_index, column=5).fill = PART_MISSING_FILL

        if change_quantity is not None:
            ws.cell(row=row_index, column=5).number_format = "0.######"
        if pdm_quantity is not None:
            ws.cell(row=row_index, column=7).number_format = "0.######"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = [64, 26, 24, 20, 22, 18, 20, 58]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    if change_items is not None:
        add_parsed_items_sheet(workbook, "Parsed Change List", change_items)
    if pdm_items is not None:
        add_parsed_items_sheet(workbook, "Parsed PDM BOM", pdm_items)

    workbook.save(output_path)


def choose_change_file(parent=None) -> str:
    if parent is not None:
        return select_excel_file("Select Change List Excel file", parent=parent)

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        return select_excel_file("Select Change List Excel file", parent=root)
    finally:
        root.destroy()


def choose_pdm_file(parent=None) -> str:
    if parent is not None:
        return select_excel_file("Select PDM BOM List Excel file", parent=parent)

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        return select_excel_file("Select PDM BOM List Excel file", parent=root)
    finally:
        root.destroy()


def choose_output_file(change_file: str | Path, parent=None) -> str:
    source_path = Path(change_file)
    if parent is not None:
        return filedialog.asksaveasfilename(
            parent=parent,
            title="Save BOM comparison result",
            initialdir=str(source_path.parent),
            initialfile=f"{source_path.stem}_bom_compare.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        return filedialog.asksaveasfilename(
            parent=root,
            title="Save BOM comparison result",
            initialdir=str(source_path.parent),
            initialfile=f"{source_path.stem}_bom_compare.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
    finally:
        root.destroy()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compare Change List and PDM BOM List.")
    parser.add_argument("--change-list", help="Change List Excel file path")
    parser.add_argument("--pdm-bom", help="PDM BOM List Excel file path")
    parser.add_argument("--change-sheet", help="Change List sheet name")
    parser.add_argument("--pdm-sheet", help="PDM BOM List sheet name")
    parser.add_argument("--output", help="Comparison result Excel file path")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    progress_win = None

    try:
        change_file = args.change_list or choose_change_file(root)
        if not change_file:
            messagebox.showinfo("BOM Compare", "Change List file was not selected.", parent=root)
            return

        pdm_file = args.pdm_bom or choose_pdm_file(root)
        if not pdm_file:
            messagebox.showinfo("BOM Compare", "PDM BOM List file was not selected.", parent=root)
            return

        output_file = args.output or choose_output_file(change_file, root)
        if not output_file:
            messagebox.showinfo("BOM Compare", "Output file was not selected.", parent=root)
            return

        progress_win, update_progress = create_progress_window(root, "BOM Compare")
        update_progress(10, "Reading Change List...")
        change_items = parse_change_list(change_file, args.change_sheet)

        update_progress(35, "Reading PDM BOM List...")
        pdm_items = parse_pdm_bom(pdm_file, args.pdm_sheet)

        update_progress(60, "Comparing BOM data...")
        rows = compare_items(change_items, pdm_items)

        update_progress(85, "Writing result Excel file...")
        write_comparison_excel(rows, output_file, change_items, pdm_items)

        update_progress(100, "Complete.")
        progress_win.destroy()
        progress_win = None
        messagebox.showinfo("BOM Compare Complete", f"Comparison result was saved:\n{output_file}", parent=root)
    except Exception as exc:
        if progress_win is not None:
            progress_win.destroy()
        messagebox.showerror("BOM Compare Error", str(exc), parent=root)
    finally:
        root.destroy()


if __name__ == "__main__":
    main()
