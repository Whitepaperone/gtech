from __future__ import annotations

import argparse
import csv
import json
import re
import tkinter as tk
from dataclasses import asdict, dataclass
from pathlib import Path
from tkinter import filedialog
from typing import Any

from CommonUtils import select_excel_file

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled at runtime for users.
    load_workbook = None


@dataclass
class ChangeListItem:
    level: str
    path: str
    part_number: str
    quantity: float


@dataclass
class PartNumberCell:
    column: int
    value: str
    quantity: float


HEADER_ALIASES = {
    "level": {"level", "lvl", "레벨"},
    "part_number": {"partnumber", "partno", "partnum", "품번", "품목번호"},
    "quantity": {"quantity", "qunatity", "qty", "qty.", "수량"},
}

CHANGE_ACTIONS = {"C", "CHANGE"}
ADD_ACTIONS = {"A", "ADD"}
DELETE_ACTIONS = {"D", "DELETE", "DEL"}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s_\-./()'`’]+", "", text).lower()


def header_matches(header_name: str, normalized: str) -> bool:
    if normalized in HEADER_ALIASES[header_name]:
        return True
    if header_name == "part_number":
        return "part" in normalized and "number" in normalized
    if header_name == "quantity":
        return normalized in {"qty", "quantity", "qunatity"} or normalized.startswith("quant")
    return False


def normalize_action(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return text.upper()


def row_action(ws, row_index: int, columns: dict[str, list[int]]) -> str:
    first_part_column = min(columns["part_number"]) if columns["part_number"] else ws.max_column + 1
    for column in range(1, first_part_column):
        action = normalize_action(ws.cell(row=row_index, column=column).value)
        if action in CHANGE_ACTIONS | ADD_ACTIONS | DELETE_ACTIONS:
            return action
    return ""


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def quantity_value(value: Any) -> float:
    if value is None or str(value).strip() == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0


def has_cell_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def get_merged_cell_value(ws, row: int, column: int) -> Any:
    cell = ws.cell(row=row, column=column)
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def strip_revision_and_description(value: Any) -> str:
    text = cell_text(value)
    if not text:
        return ""
    text = re.split(r"\s*\(", text, maxsplit=1)[0].strip()
    text = text.rsplit("/", 1)[0].strip()
    return text


def find_root_part_number_above_header(ws, header_row: int, columns: dict[str, list[int]]) -> str:
    for part_column in reversed(columns["part_number"]):
        for row_index in range(header_row - 1, 0, -1):
            value = strip_revision_and_description(get_merged_cell_value(ws, row_index, part_column))
            if value:
                return value
    return ""


def collect_header_columns(ws, header_row: int) -> dict[str, list[int]]:
    columns = {"level": [], "part_number": [], "quantity": []}

    for cell in ws[header_row]:
        normalized = normalize_header(cell.value)
        for header_name in HEADER_ALIASES:
            if header_matches(header_name, normalized):
                columns[header_name].append(cell.column)

    if not columns["part_number"]:
        raise ValueError("Part Number 열을 찾지 못했습니다.")
    if not columns["quantity"]:
        raise ValueError("Quantity 열을 찾지 못했습니다.")

    return columns


def find_level_header_row(ws, max_scan_rows: int = 30) -> int | None:
    for row_index in range(1, min(ws.max_row, max_scan_rows) + 1):
        for cell in ws[row_index]:
            if header_matches("level", normalize_header(cell.value)):
                return row_index
    return None


def select_change_list_sheet(workbook, sheet_name: str | None = None):
    if sheet_name:
        ws = workbook[sheet_name]
        header_row = find_level_header_row(ws)
        if header_row is None:
            raise ValueError(f"'{sheet_name}' 시트에서 Level 헤더 셀을 찾을 수 없습니다.")
        return ws, header_row

    ws = workbook.active
    header_row = find_level_header_row(ws)
    if header_row is not None:
        return ws, header_row

    for candidate in workbook.worksheets:
        header_row = find_level_header_row(candidate)
        if header_row is not None:
            return candidate, header_row

    raise ValueError("헤더 행을 찾지 못했습니다. 모든 시트의 30행 안에서 Level 셀을 찾을 수 없습니다.")


def quantity_column_for_part(
    part_column: int,
    next_part_column: int | None,
    quantity_columns: list[int],
    fallback_index: int,
) -> int | None:
    grouped_quantity_columns = [
        col
        for col in quantity_columns
        if col > part_column and (next_part_column is None or col < next_part_column)
    ]
    if grouped_quantity_columns:
        return grouped_quantity_columns[0]

    if fallback_index < len(quantity_columns):
        return quantity_columns[fallback_index]

    return quantity_columns[-1] if quantity_columns else None


def extract_part_chain(ws, row_index: int, columns: dict[str, list[int]]) -> list[PartNumberCell]:
    chain: list[PartNumberCell] = []
    part_columns = columns["part_number"]

    for index, part_column in enumerate(part_columns):
        value = cell_text(ws.cell(row=row_index, column=part_column).value)
        if not value:
            continue

        next_part_column = part_columns[index + 1] if index + 1 < len(part_columns) else None
        qty_col = quantity_column_for_part(
            part_column,
            next_part_column,
            columns["quantity"],
            index,
        )
        qty = quantity_value(ws.cell(row=row_index, column=qty_col).value) if qty_col else 0
        chain.append(PartNumberCell(column=part_column, value=value, quantity=qty))

    return chain


def extract_change_part_chain(ws, row_index: int, columns: dict[str, list[int]]) -> list[PartNumberCell]:
    chain: list[PartNumberCell] = []
    part_columns = columns["part_number"]
    previous_part_number = ""

    for index, part_column in enumerate(part_columns):
        value = cell_text(ws.cell(row=row_index, column=part_column).value)
        next_part_column = part_columns[index + 1] if index + 1 < len(part_columns) else None
        qty_col = quantity_column_for_part(
            part_column,
            next_part_column,
            columns["quantity"],
            index,
        )
        raw_qty = ws.cell(row=row_index, column=qty_col).value if qty_col else None
        qty = quantity_value(raw_qty) if qty_col else 0

        if value:
            previous_part_number = value
            chain.append(PartNumberCell(column=part_column, value=value, quantity=qty))
            continue

        if previous_part_number and qty_col and has_cell_value(raw_qty):
            chain.append(PartNumberCell(column=part_column, value=previous_part_number, quantity=qty))

    return chain


def row_level(ws, row_index: int, columns: dict[str, list[int]]) -> str:
    for level_column in columns["level"]:
        value = cell_text(ws.cell(row=row_index, column=level_column).value)
        if value and normalize_action(value) not in CHANGE_ACTIONS | ADD_ACTIONS | DELETE_ACTIONS:
            return value
    return ""


def level_number(value: str) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def normalize_part_key(value: str) -> str:
    return str(value or "").strip().upper().replace(" ", "").replace("_", "")


def build_path(parent_parts_by_level: dict[int, str], level: int, part_number: str) -> str:
    parts = [parent_parts_by_level[index] for index in sorted(parent_parts_by_level) if index < level]
    parts.append(part_number)
    return " > ".join(part for part in parts if part)


def prune_stack(parent_parts_by_level: dict[int, str], level: int) -> None:
    for existing_level in list(parent_parts_by_level):
        if existing_level >= level:
            del parent_parts_by_level[existing_level]


def scoped_alias_key(parent_path: str, part_number: str) -> str:
    return f"{parent_path}|{normalize_part_key(part_number)}"


def resolve_root(alias_by_part_number: dict[str, str], alias_key: str) -> str:
    visited: list[str] = []
    current = alias_key

    while current in alias_by_part_number and alias_by_part_number[current] != current:
        visited.append(current)
        current = alias_by_part_number[current]

    for item in visited:
        alias_by_part_number[item] = current

    return current


def parse_change_list(xlsx_path: str | Path, sheet_name: str | None = None) -> list[ChangeListItem]:
    if load_workbook is None:
        raise RuntimeError("openpyxl이 필요합니다. `pip install openpyxl` 후 다시 실행해주세요.")

    workbook = load_workbook(xlsx_path, data_only=True)
    ws, header_row = select_change_list_sheet(workbook, sheet_name)
    columns = collect_header_columns(ws, header_row)

    active_items: dict[str, ChangeListItem] = {}
    alias_by_part_number: dict[str, str] = {}
    parent_parts_by_level: dict[int, str] = {}
    synthetic_root_part_number = find_root_part_number_above_header(ws, header_row, columns)
    has_seen_level_zero = False
    has_added_synthetic_root = False

    for row_index in range(header_row + 1, ws.max_row + 1):
        action = row_action(ws, row_index, columns)
        chain = (
            extract_change_part_chain(ws, row_index, columns)
            if action in CHANGE_ACTIONS
            else extract_part_chain(ws, row_index, columns)
        )
        if not chain:
            continue

        first_part = chain[0]
        latest_part = chain[-1]
        level = row_level(ws, row_index, columns)
        numeric_level = level_number(level)
        if numeric_level is None:
            if action in DELETE_ACTIONS:
                delete_part_numbers = {normalize_part_key(part.value) for part in chain}
                for key, item in list(active_items.items()):
                    if normalize_part_key(item.part_number) in delete_part_numbers:
                        active_items.pop(key, None)
            continue

        if numeric_level == 0:
            has_seen_level_zero = True

        if (
            numeric_level > 0
            and not has_seen_level_zero
            and not has_added_synthetic_root
            and synthetic_root_part_number
        ):
            active_items[scoped_alias_key("", synthetic_root_part_number)] = ChangeListItem(
                level="0",
                path=synthetic_root_part_number,
                part_number=synthetic_root_part_number,
                quantity=1,
            )
            parent_parts_by_level[0] = synthetic_root_part_number
            has_added_synthetic_root = True

        prune_stack(parent_parts_by_level, numeric_level)
        parent_path = " > ".join(
            parent_parts_by_level[index] for index in sorted(parent_parts_by_level) if index < numeric_level
        )
        first_alias_key = scoped_alias_key(parent_path, first_part.value)
        latest_alias_key = scoped_alias_key(parent_path, latest_part.value)
        root_part_number = resolve_root(alias_by_part_number, first_alias_key)
        latest_path = build_path(parent_parts_by_level, numeric_level, latest_part.value)

        if action in DELETE_ACTIONS:
            active_items.pop(root_part_number, None)
            alias_by_part_number[first_alias_key] = root_part_number
            continue

        if action in CHANGE_ACTIONS:
            current = active_items.get(root_part_number)
            active_items[root_part_number] = ChangeListItem(
                level=current.level if current and current.level else level,
                path=latest_path,
                part_number=latest_part.value,
                quantity=latest_part.quantity,
            )
            alias_by_part_number[first_alias_key] = root_part_number
            alias_by_part_number[latest_alias_key] = root_part_number
            parent_parts_by_level[numeric_level] = latest_part.value
            continue

        if action in ADD_ACTIONS:
            root_part_number = latest_alias_key
            alias_by_part_number[root_part_number] = root_part_number
            active_items[root_part_number] = ChangeListItem(
                level=level,
                path=latest_path,
                part_number=latest_part.value,
                quantity=latest_part.quantity,
            )
            parent_parts_by_level[numeric_level] = latest_part.value
            continue

        active_items[root_part_number] = ChangeListItem(
            level=level,
            path=latest_path,
            part_number=latest_part.value,
            quantity=latest_part.quantity,
        )
        alias_by_part_number[first_alias_key] = root_part_number
        alias_by_part_number[latest_alias_key] = root_part_number
        parent_parts_by_level[numeric_level] = latest_part.value

    return list(active_items.values())


def write_csv(items: list[ChangeListItem], output_path: str | Path) -> None:
    with open(output_path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["level", "path", "part_number", "quantity"],
        )
        writer.writeheader()
        writer.writerows(asdict(item) for item in items)


def write_json(items: list[ChangeListItem], output_path: str | Path) -> None:
    with open(output_path, "w", encoding="utf-8") as file:
        json.dump([asdict(item) for item in items], file, ensure_ascii=False, indent=2)


def save_items(items: list[ChangeListItem], output_path: str | Path) -> None:
    path = Path(output_path)
    suffix = path.suffix.lower()

    if suffix == ".json":
        write_json(items, path)
        return

    if suffix == ".csv" or suffix == "":
        if suffix == "":
            path = path.with_suffix(".csv")
        write_csv(items, path)
        return

    raise ValueError("Output file must be .csv or .json.")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="취탈리스트에서 최종 Level, Part Number, Quantity 목록을 추출합니다.")
    parser.add_argument("xlsx_path", nargs="?", help="취탈리스트 엑셀 파일 경로")
    parser.add_argument("--sheet", help="읽을 시트명. 생략하면 첫 번째 활성 시트를 읽습니다.")
    parser.add_argument("--csv", help="CSV 출력 경로")
    parser.add_argument("--json", help="JSON 출력 경로")
    return parser


def choose_change_list_file() -> str:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        return select_excel_file("취탈리스트 엑셀 파일 선택", parent=root)
    finally:
        root.destroy()


def choose_output_file(source_file: str | Path) -> str:
    source_path = Path(source_file)
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        return filedialog.asksaveasfilename(
            parent=root,
            title="Save parsed change list",
            initialdir=str(source_path.parent),
            initialfile=f"{source_path.stem}_parsed.csv",
            defaultextension=".csv",
            filetypes=[
                ("CSV files", "*.csv"),
                ("JSON files", "*.json"),
            ],
        )
    finally:
        root.destroy()


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    xlsx_path = args.xlsx_path or choose_change_list_file()
    if not xlsx_path:
        print("엑셀 파일을 선택하지 않았습니다.")
        return

    items = parse_change_list(xlsx_path, args.sheet)

    if args.csv:
        write_csv(items, args.csv)
    if args.json:
        write_json(items, args.json)

    if not args.csv and not args.json:
        output_path = choose_output_file(xlsx_path)
        if not output_path:
            print("저장 파일을 선택하지 않았습니다.")
            return
        save_items(items, output_path)
        print(f"Saved parsed change list: {output_path}")


if __name__ == "__main__":
    main()
