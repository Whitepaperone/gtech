from tkinter import messagebox
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook

from AppConstants import COA_EXCLUDE_CUSTOMERS, FINISH_PLAN_SHEET_NAMES
from CommonUtils import (
    create_progress_window,
    normalize_part_no,
    normalize_process_token,
    select_excel_save_file,
)
from ExtractPlan import extract_finish_plan_sheet as extract_finish_plan_sheet_from_module


# =========================
# 설정
# =========================
SHEET_NAME = FINISH_PLAN_SHEET_NAMES
EXCLUDE_CUSTOMERS = COA_EXCLUDE_CUSTOMERS


def extract_plan_file(plan_file: str, progress=None) -> pd.DataFrame:
    if progress:
        progress(0, "생산계획 파일 열기 중...")
    
    wb = load_workbook(plan_file, data_only=True)
    if progress:
            progress(35, "생산계획 파일 열기 완료")
    frames = []
    total = len(SHEET_NAME)

    for i, sheet_name in enumerate(SHEET_NAME, start=1):
        if sheet_name not in wb.sheetnames:
            if progress:
                progress(35 + int(i / total * 50), f"{sheet_name} 시트 없음, 건너뜀")
            continue
        if progress:
            progress(35 + int((i - 1) / total * 50), f"{sheet_name} 추출 중...")

        ws = wb[sheet_name]
        df = extract_finish_plan_sheet_from_module(ws)

        frames.append(df)

        if progress:
            progress(35 + int(i / total * 50), f"{sheet_name} 추출 완료: {len(df)}건")

    if progress:
        progress(90, "생산계획 데이터 정리 중...")
    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


# =========================
# 수주 파일 추출
# =========================
def extract_order_file(order_file: str) -> pd.DataFrame:
    df = pd.read_excel(order_file, header=0)

    required_cols = ["수주번호", "날짜", "납기", "출고(고객)사", "품번", "품명", "규격", "수량", "할당수량", "잔여량", "비고"]

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise RuntimeError(f"수주 파일에서 컬럼을 찾지 못했습니다: {missing}")

    df = df[required_cols].copy()
    df = df[
        ~df["출고(고객)사"].apply(lambda x: normalize_process_token(x) in EXCLUDE_CUSTOMERS)
    ].copy()

    df["품번"] = df["품번"].apply(normalize_part_no)
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.normalize()
    df["납기"] = pd.to_datetime(df["납기"], errors="coerce").dt.normalize()

    for col in ["수량", "할당수량", "잔여량"]:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(",", "", regex=False)
            .replace("nan", "0")
        )
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df = df[df["잔여량"] > 0].copy()
    df = df[df["품번"] != ""].copy()
    df = df[df["납기"].notna()].copy()

    return df

# =========================
# 품번 매핑
# =========================
def apply_part_mapping(order_df: pd.DataFrame, mapping_file: str = None) -> pd.DataFrame:
    order_df = order_df.copy()

    # 수주 원본 품번 보존
    order_df["수주품번"] = order_df["품번"]
    order_df["비교품번"] = order_df["품번"]
    order_df["품번매핑여부"] = "N"

    if not mapping_file:
        return order_df

    map_df = pd.read_excel(mapping_file, header=0)
    map_df.columns = [normalize_process_token(c) for c in map_df.columns]

    # 매핑표 헤더 표준화
    rename_map = {
        "수주품번": "수주품번",
        "변경품번": "비교품번",
        "계획품번": "비교품번",
    }

    map_df = map_df.rename(columns=rename_map)

    if "수주품번" not in map_df.columns or "비교품번" not in map_df.columns:
        raise RuntimeError(
            f"품번매핑 파일에 '수주 품번'과 '변경 품번' 컬럼이 필요합니다. 현재 컬럼: {list(map_df.columns)}"
        )

    map_df = map_df[["수주품번", "비교품번"]].copy()

    map_df["수주품번"] = map_df["수주품번"].apply(normalize_part_no)
    map_df["비교품번"] = map_df["비교품번"].apply(normalize_part_no)

    map_df = map_df[
        (map_df["수주품번"] != "") &
        (map_df["비교품번"] != "")
    ].copy()

    map_df = map_df.drop_duplicates(subset=["수주품번"], keep="last")

    mapping_dict = dict(zip(map_df["수주품번"], map_df["비교품번"]))

    order_df["비교품번"] = order_df["수주품번"].map(mapping_dict).fillna(order_df["수주품번"])

    order_df["품번매핑여부"] = order_df.apply(
        lambda r: "Y" if r["비교품번"] != r["수주품번"] else "N",
        axis=1
    )

    return order_df

# =========================
# 수주잔량 vs 생산계획 비교
# =========================
def compare_order_balance_vs_plan(plan_df: pd.DataFrame, order_df: pd.DataFrame, horizon_weeks: int = 6):
    if plan_df.empty:
        raise RuntimeError("생산계획 데이터가 없습니다.")

    if order_df.empty:
        raise RuntimeError("잔여량이 있는 수주 데이터가 없습니다.")

    plan_df = plan_df.copy()
    order_df = order_df.copy()

    today = pd.Timestamp.today().normalize()
    horizon_end = today + pd.Timedelta(weeks=horizon_weeks)

    plan_df["품번"] = plan_df["품번"].apply(normalize_part_no)
    plan_df["날짜"] = pd.to_datetime(plan_df["날짜"], errors="coerce").dt.normalize()
    plan_df["계획수량"] = pd.to_numeric(plan_df["계획수량"], errors="coerce").fillna(0)

    plan_df = plan_df[
        (plan_df["품번"] != "") &
        (plan_df["날짜"].notna()) &
        (plan_df["계획수량"] > 0)
    ].copy()

    plan_start_date = plan_df["날짜"].min()

    # 1) 과거 납기 잔량
    past_order_df = order_df[
        order_df["납기"] < plan_start_date
    ].copy()

    past_order_df["판정"] = "계획시작일 이전 납기 / 재고 또는 미출고 확인 필요"
    past_order_df["오늘날짜"] = today
    past_order_df["계획시작일"] = plan_start_date

    # 2) 장기 납기 잔량
    future_order_df = order_df[
        order_df["납기"] > horizon_end
    ].copy()

    future_order_df["판정"] = f"{horizon_weeks}주 이후 납기 / 확인 제외"
    future_order_df["오늘날짜"] = today
    future_order_df["확인종료일"] = horizon_end

    # 3) 실제 비교 대상: 오늘 ~ 오늘+6주
    target_order_df = order_df[
        (order_df["납기"] >= plan_start_date) &
        (order_df["납기"] <= horizon_end)
    ].copy()

    plan_group = (
        plan_df
        .groupby(["품번", "날짜"], as_index=False)["계획수량"]
        .sum()
        .sort_values(["품번", "날짜"])
    )

    plan_group["계획잔량"] = plan_group["계획수량"]

    results = []

    target_order_df = target_order_df.sort_values(["품번", "납기", "수주번호"])

    for _, order in target_order_df.iterrows():
        part_no = order["비교품번"]
        due_date = order["납기"]
        need_qty = float(order["잔여량"])
        remain_need = need_qty

        matched_plan_qty = 0.0
        matched_plan_dates = []

        candidate_idx = plan_group[
            (plan_group["품번"] == part_no) &
            (plan_group["날짜"] <= due_date) &
            (plan_group["계획잔량"] > 0)
        ].index

        for idx in candidate_idx:
            available = float(plan_group.at[idx, "계획잔량"])
            use_qty = min(available, remain_need)

            plan_group.at[idx, "계획잔량"] = available - use_qty
            remain_need -= use_qty
            matched_plan_qty += use_qty

            matched_plan_dates.append(
                f"{plan_group.at[idx, '날짜'].strftime('%Y-%m-%d')}({use_qty:g})"
            )

            if remain_need <= 0:
                break

        shortage_qty = max(remain_need, 0)

        if matched_plan_qty == 0:
            judgment = "계획 없음"
        elif shortage_qty > 0:
            judgment = "계획 부족"
        else:
            judgment = "반영 완료"

        row = order.to_dict()
        row.update({
            "확인종료일": horizon_end,
            "납기전_반영계획수량": matched_plan_qty,
            "부족수량": shortage_qty,
            "반영계획일자": ", ".join(matched_plan_dates),
            "판정": judgment,
        })

        results.append(row)

    compare_df = pd.DataFrame(results)
    remain_plan_df = plan_group[plan_group["계획잔량"] > 0].copy()

    return compare_df, past_order_df, future_order_df, remain_plan_df


# =========================
# 실행부
# =========================
def main():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    plan_file = filedialog.askopenfilename(
        title="생산계획 엑셀 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )

    if not plan_file:
        messagebox.showinfo("알림", "생산계획 파일을 선택하지 않았습니다.")
        return

    order_file = filedialog.askopenfilename(
        title="수주 엑셀 선택",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
    )

    if not order_file:
        messagebox.showinfo("알림", "수주 파일을 선택하지 않았습니다.")
        return
    
    mapping_file = filedialog.askopenfilename(title="품번매핑 파일 선택")
    if not mapping_file:
        messagebox.showinfo("알림", "품번매핑 파일을 선택하지 않았습니다. 매핑 없이 진행합니다.")
        mapping_file = None


    save_file = select_excel_save_file("결과 파일 저장", plan_file, "COA수주_비교", root)
    if not save_file:
        messagebox.showinfo("알림", "저장할 파일을 선택하지 않았습니다. 결과 저장 없이 종료합니다.")
        return

   

    progress_win, progress = create_progress_window(root, "COA 수주잔량 비교 진행")

    try:
        progress(10, "수주 파일 읽는 중...")
        order_df = extract_order_file(order_file)

        progress(20, "품번 매핑 적용 중...")
        order_df = apply_part_mapping(order_df, mapping_file)
        mapped_count = order_df["품번매핑여부"].eq("Y").sum()
        progress(30, f"품번 매핑 적용 완료: {mapped_count}건 변경")

        plan_df = extract_plan_file(plan_file, progress=progress)

        progress(92, "수주잔량과 생산계획 비교 중...")
        compare_df, past_order_df, future_order_df, remain_plan_df = compare_order_balance_vs_plan(
            plan_df,
            order_df,
            horizon_weeks=6
        )

        progress(96, "엑셀 저장 중...")

        with pd.ExcelWriter(save_file, engine="openpyxl") as writer:
            compare_df.to_excel(writer, sheet_name="6주내_수주잔량_계획확인", index=False)
            past_order_df.to_excel(writer, sheet_name="과거납기_미출고잔량", index=False)
            future_order_df.to_excel(writer, sheet_name="6주이후_확인제외", index=False)
            remain_plan_df.to_excel(writer, sheet_name="계획잔량_참고", index=False)
            plan_df.to_excel(writer, sheet_name="추출된_생산계획", index=False)
            order_df.to_excel(writer, sheet_name="추출된_수주잔량", index=False)

        progress(100, "완료")
        progress_win.destroy()
        messagebox.showinfo("알림", f"결과 저장 완료: {save_file}")

    except Exception as e:
        progress_win.destroy()
        messagebox.showerror("오류", str(e))

if __name__ == "__main__":
    main()
