from openpyxl import load_workbook
import pandas as pd
from tkinter import Tk, filedialog, messagebox
from datetime import datetime

from CommonUtils import apply_quantities_by_part_left_to_right, load_actual_quantities_by_part

OUTPUT_FILE = "./output.xlsx"


def is_date_header_value(value):
    if pd.isna(value):
        return False

    if isinstance(value, (datetime, pd.Timestamp)):
        return True

    if isinstance(value, str):
        text = value.strip()
        if "/" in text or "-" in text:
            return not pd.isna(pd.to_datetime(text, errors="coerce"))
        return False

    # Plain production quantities like 5, 12, 100 are not date headers.
    # Real Excel serial dates are normally large numbers around 40000+.
    if isinstance(value, (int, float)) and value > 20000:
        return not pd.isna(pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce"))

    return False


def format_date_header(value):
    if isinstance(value, str):
        text = value.strip()
        if text:
            return text

    if isinstance(value, (datetime, pd.Timestamp)):
        dt = pd.to_datetime(value)
        if dt.year == datetime.now().year:
            return f"{dt.month}/{dt.day}"
        return f"{dt.year}/{dt.month}/{dt.day}"

    if isinstance(value, (int, float)) and value > 20000:
        dt = pd.to_datetime(value, unit="D", origin="1899-12-30")
        if dt.year == datetime.now().year:
            return f"{dt.month}/{dt.day}"
        return f"{dt.year}/{dt.month}/{dt.day}"

    return str(value).strip()


def find_date_columns(df):
    """
    날짜 열 자동 탐지
    - 날짜가 연속되지 않아도 실제 날짜 컬럼만 수집
    """
    for row_idx in range(0, 10):
        row = df.iloc[row_idx]
        date_cols = []

        for i, val in enumerate(row):
            if is_date_header_value(val):
                date_cols.append(i)

        if len(date_cols) >= 5:
            return row_idx, date_cols

    raise Exception("날짜 열을 찾을 수 없습니다.")


def find_data_start_row(df, start_col, end_col):
    """
    실제 데이터 시작 행 찾기
    """
    for i in range(0, len(df)):
        row = df.iloc[i, start_col:end_col]

        # 숫자 포함된 행 찾기
        numeric_count = sum(
            1 for x in row
            if isinstance(x, (int, float)) and not pd.isna(x)
        )

        if numeric_count >= 1:
            return i

    raise Exception("데이터 시작 행을 찾을 수 없습니다.")


def adjust_row(values):
    """
    재고 → 미래 계획 차감
    """
    stock = 0
    result = values.copy()

    # 첫날 음수 → 선생산량
    if result[0] < 0:
        stock = abs(result[0])
        result[0] = 0

    for i in range(1, len(result)):
        if stock <= 0:
            break

        val = result[i]

        if pd.isna(val):
            continue

        if val > 0:
            if val >= stock:
                result[i] -= stock
                stock = 0
            else:
                stock -= val
                result[i] = 0

    return result
def main():
    Tk().withdraw()

    file_path = filedialog.askopenfilename(
        title="엑셀 선택",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not file_path:
        return

    try:
        df = pd.read_excel(file_path, header=None)

        # 1️⃣ 날짜 열 찾기
        date_row_idx, date_cols = find_date_columns(df)

        # 2️⃣ 데이터 시작 행 찾기
        start_col = min(date_cols)
        end_col = max(date_cols) + 1

        start_row = find_data_start_row(df, start_col, end_col)

        today_grouped = load_actual_quantities_by_part(OUTPUT_FILE)

        product_col = start_col - 1
        customer_code_col = start_col - 2
        model_col = start_col - 3


        date_row = df.iloc[date_row_idx, date_cols].tolist()
        date_row = [
            format_date_header(d)
            for d in date_row
        ]
        date_internal_cols = [
            f"__date_{idx}"
            for idx in range(len(date_row))
        ]

        processed_rows = []

        # 3️⃣ 처리
        
        for i in range(start_row, len(df)):
            row_full = df.iloc[i]


            product = row_full[product_col]
            model= row_full[model_col]
            customer_code = row_full[customer_code_col]

            if pd.isna(product):
                continue

            row = row_full.iloc[date_cols].tolist()
            row = [pd.to_numeric(x, errors='coerce') for x in row]
            row = [0 if pd.isna(x) else x for x in row]

            adjusted = adjust_row(row)

            processed_rows.append([model, customer_code, product] + adjusted)


        # 저장
        # 🔥 DataFrame으로 변환
        result_df = pd.DataFrame(processed_rows)

        result_df.columns = ["모델", "고객사 품번", "품번"] + date_internal_cols

        result_df.iloc[:, 3:] = result_df.iloc[:, 3:].apply(pd.to_numeric, errors='coerce').fillna(0)

        agg_dict = {"모델": "first", "고객사 품번": lambda x: x.dropna().iloc[0] if not x.dropna().empty else ""}


        for col in result_df.columns[3:]:
            agg_dict[col] = "sum"

        # 🔥 품번 기준 그룹화 + 합계
        result_df["품번"] = result_df["품번"].astype(str).str.strip()
        result_df = result_df.groupby("품번", as_index=False).agg(agg_dict)
        result_df = apply_quantities_by_part_left_to_right(
            result_df,
            today_grouped,
            part_col="품번",
            value_start_idx=3,
        )
        result_df = result_df[(result_df.iloc[:, 3:] != 0).any(axis=1)]

        output_date_cols = []
        seen_dates = {}
        for date_label in date_row:
            seen_dates[date_label] = seen_dates.get(date_label, 0) + 1
            if seen_dates[date_label] == 1:
                output_date_cols.append(date_label)
            else:
                output_date_cols.append(f"{date_label}_{seen_dates[date_label]}")

        result_df = result_df.rename(
            columns=dict(zip(date_internal_cols, output_date_cols))
        )

        # 숫자 날짜 컬럼
        date_columns = result_df.columns[3:]

        # 행 합계 추가
        result_df["합계"] = result_df[date_columns].sum(axis=1)

        # 열 합계 행 추가
        total_row = {
            "모델": "합계",
            "고객사 품번": "",
            "품번": ""
        }

        for col in date_columns:
            total_row[col] = result_df[col].sum()

        total_row["합계"] = result_df["합계"].sum()

        result_df = pd.concat([result_df, pd.DataFrame([total_row])], ignore_index=True)

        output_path = file_path.replace(".xlsx", "_MRP.xlsx")

        result_df.to_excel(output_path, index=False)


        wb = load_workbook(output_path)
        ws = wb.active

        for col in range(3, len(result_df.columns) + 1):
            for row in range(1, len(result_df) + 1):  # ⭐ 헤더 포함
                cell = ws.cell(row=row, column=col)
                
                if isinstance(cell.value, datetime):
                    cell.number_format = 'yyyy-mm-dd'

        wb.save(output_path)

        messagebox.showinfo("완료", f"저장 완료:\n{output_path}")

    except Exception as e:
        messagebox.showerror("오류", str(e))


if __name__ == "__main__":
    main()
