from __future__ import annotations

import argparse
import csv
import json
import re
import tempfile
import tkinter as tk
from dataclasses import asdict, dataclass
from pathlib import Path
from tkinter import filedialog
from typing import Any

from CommonUtils import select_excel_file

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled at runtime for users.
    load_workbook = None


PRODUCT_ID_ROW = 10


@dataclass
class PdmBomItem:
    level: str
    path: str
    part_number: str
    quantity: float


HEADER_ALIASES = {
    "level": {"level", "lvl", "레벨"},
    "part_number": {"id", "idpartnumber", "partnumber", "partno", "partnum", "품번", "품목번호"},
    "quantity": {"quantity", "qunatity", "qty", "수량"},
}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s_\-./()]+", "", text).lower()


def header_matches(header_name: str, normalized: str) -> bool:
    if normalized in HEADER_ALIASES[header_name]:
        return True
    if header_name == "part_number":
        return "id" in normalized and "partnumber" in normalized
    return False


def normalize_token(value: Any) -> str:
    return normalize_header(value)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def quantity_value(value: Any) -> float:
    if value is None or str(value).strip() == "":
        return 0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0


def level_number(value: str) -> int | None:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def build_path(parent_parts_by_level: dict[int, str], level: int, part_number: str) -> str:
    parts = [parent_parts_by_level[index] for index in sorted(parent_parts_by_level) if index < level]
    parts.append(part_number)
    return " > ".join(part for part in parts if part)


def prune_stack(parent_parts_by_level: dict[int, str], level: int) -> None:
    for existing_level in list(parent_parts_by_level):
        if existing_level >= level:
            del parent_parts_by_level[existing_level]


def strip_revision_from_product_id(product_id: Any) -> str:
    text = cell_text(product_id)
    if not text:
        return ""
    return text.rsplit("/", 1)[0].strip()


def collect_header_columns(ws, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}

    for cell in ws[header_row]:
        normalized = normalize_header(cell.value)
        for header_name in HEADER_ALIASES:
            if header_matches(header_name, normalized) and header_name not in columns:
                columns[header_name] = cell.column

    required = ["level", "part_number", "quantity"]
    missing = [name for name in required if name not in columns]
    if missing:
        missing_text = ", ".join(missing)
        raise ValueError(
            f"Could not find required BOM header cell(s) on row {header_row}: {missing_text}"
        )

    return columns


def find_sub_assembly_row(ws) -> int:
    for row_index in range(1, ws.max_row + 1):
        for cell in ws[row_index]:
            if normalize_token(cell.value) == "subassemblypart":
                return row_index
    raise ValueError("Could not find 'Sub-Assembly/Part' row in PDM BOM.")


def find_bom_header_row(ws, start_row: int) -> int:
    for row_index in range(start_row + 1, ws.max_row + 1):
        for cell in ws[row_index]:
            if header_matches("level", normalize_header(cell.value)):
                return row_index
    raise ValueError("Could not find BOM table header row below 'Sub-Assembly/Part'.")


def find_product_id(ws) -> str:
    for cell in ws[PRODUCT_ID_ROW]:
        value = strip_revision_from_product_id(cell.value)
        if value:
            return value
    raise ValueError("Product ID was not found on row 10.")


def convert_xls_to_xlsx(xls_path: str | Path) -> Path:
    source_path = Path(xls_path).resolve()
    target_path = Path(tempfile.gettempdir()) / f"{source_path.stem}_converted.xlsx"

    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError(
            ".xls files require Excel conversion. Install pywin32 or save the BOM as .xlsx first."
        ) from exc

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = None
    try:
        workbook = excel.Workbooks.Open(str(source_path))
        workbook.SaveAs(str(target_path), FileFormat=51)
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        excel.Quit()

    return target_path


def workbook_path_for_openpyxl(input_path: str | Path) -> tuple[Path, bool]:
    path = Path(input_path)
    suffix = path.suffix.lower()

    if suffix == ".xls":
        return convert_xls_to_xlsx(path), True
    if suffix in {".xlsx", ".xlsm"}:
        return path, False

    raise ValueError("PDM BOM file must be .xls, .xlsx, or .xlsm.")


def parse_pdm_bom(xlsx_path: str | Path, sheet_name: str | None = None) -> list[PdmBomItem]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Run `pip install openpyxl` and try again.")

    workbook_path, should_delete = workbook_path_for_openpyxl(xlsx_path)

    workbook = None
    try:
        workbook = load_workbook(workbook_path, data_only=True)
        ws = workbook[sheet_name] if sheet_name else workbook.active

        sub_assembly_row = find_sub_assembly_row(ws)
        header_row = find_bom_header_row(ws, sub_assembly_row)
        columns = collect_header_columns(ws, header_row)
        product_part_number = find_product_id(ws)
        parent_parts_by_level: dict[int, str] = {0: product_part_number}
        items = [
            PdmBomItem(
                level="0",
                path=product_part_number,
                part_number=product_part_number,
                quantity=1,
            )
        ]

        for row_index in range(header_row + 1, ws.max_row + 1):
            part_number = cell_text(ws.cell(row=row_index, column=columns["part_number"]).value)
            if not part_number:
                continue

            level = cell_text(ws.cell(row=row_index, column=columns["level"]).value)
            numeric_level = level_number(level)
            if numeric_level is None:
                continue

            prune_stack(parent_parts_by_level, numeric_level)
            path = build_path(parent_parts_by_level, numeric_level, part_number)
            items.append(
                PdmBomItem(
                    level=level,
                    path=path,
                    part_number=part_number,
                    quantity=quantity_value(ws.cell(row=row_index, column=columns["quantity"]).value),
                )
            )
            parent_parts_by_level[numeric_level] = part_number

        return items
    finally:
        if workbook is not None:
            workbook.close()
        if should_delete and workbook_path.exists():
            workbook_path.unlink()


def write_csv(items: list[PdmBomItem], output_path: str | Path) -> None:
    with open(output_path, "w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["level", "path", "part_number", "quantity"],
        )
        writer.writeheader()
        writer.writerows(asdict(item) for item in items)


def write_json(items: list[PdmBomItem], output_path: str | Path) -> None:
    with open(output_path, "w", encoding="utf-8") as file:
        json.dump([asdict(item) for item in items], file, ensure_ascii=False, indent=2)


def save_items(items: list[PdmBomItem], output_path: str | Path) -> Path:
    path = Path(output_path)
    suffix = path.suffix.lower()

    if suffix == ".json":
        write_json(items, path)
        return path

    if suffix == "":
        path = path.with_suffix(".csv")

    if path.suffix.lower() == ".csv":
        write_csv(items, path)
        return path

    raise ValueError("Output file must be .csv or .json.")


def choose_pdm_bom_file() -> str:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        return select_excel_file("PDM BOM List 엑셀 파일 선택", parent=root)
    finally:
        root.destroy()


def choose_output_file(source_file: str | Path) -> str:
    source_path = Path(source_file)
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        return filedialog.asksaveasfilename(
            parent=root,
            title="Save parsed PDM BOM list",
            initialdir=str(source_path.parent),
            initialfile=f"{source_path.stem}_parsed.csv",
            defaultextension=".csv",
            filetypes=[
                ("CSV files", "*.csv"),
                ("JSON files", "*.json"),
            ],
        )
    finally:
        root.destroy()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="PDM BOM List를 비교용 목록으로 파싱합니다.")
    parser.add_argument("xlsx_path", nargs="?", help="PDM BOM List 엑셀 파일 경로")
    parser.add_argument("--sheet", help="읽을 시트명. 생략하면 첫 번째 활성 시트를 읽습니다.")
    parser.add_argument("--output", help="CSV 또는 JSON 출력 경로")
    parser.add_argument("--csv", help="CSV 출력 경로")
    parser.add_argument("--json", help="JSON 출력 경로")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    xlsx_path = args.xlsx_path or choose_pdm_bom_file()
    if not xlsx_path:
        print("엑셀 파일을 선택하지 않았습니다.")
        return

    items = parse_pdm_bom(xlsx_path, args.sheet)

    if args.output:
        saved_path = save_items(items, args.output)
        print(f"Saved parsed PDM BOM list: {saved_path}")
        return

    if args.csv:
        write_csv(items, args.csv)
        print(f"Saved parsed PDM BOM list: {args.csv}")
        return

    if args.json:
        write_json(items, args.json)
        print(f"Saved parsed PDM BOM list: {args.json}")
        return

    output_path = choose_output_file(xlsx_path)
    if not output_path:
        print("저장 파일을 선택하지 않았습니다.")
        return

    saved_path = save_items(items, output_path)
    print(f"Saved parsed PDM BOM list: {saved_path}")


if __name__ == "__main__":
    main()
