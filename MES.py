from typing import Optional
import tkinter as tk
from tkinter import filedialog, messagebox
from tkcalendar import Calendar

import pandas as pd
from openpyxl import load_workbook

from AppConstants import (
    FINAL_COMPARE_COLUMNS,
    MES_OUTPUT_FILE,
    MES_SHEET_CONFIG,
    WORKSHOP_TO_PROCESS,
)
from CommonUtils import (
    create_progress_window,
    filter_by_period,
    normalize_part_no,
    normalize_process_token,
    safe_float,
    select_excel_file,
)
from ExtractPlan import extract_plan_sheet as extract_plan_sheet_from_module

# =========================
# 상수
# =========================
OUTPUT_FILE = MES_OUTPUT_FILE
SHEET_CONFIG = MES_SHEET_CONFIG

# =========================
# 공통 유틸
# =========================
def select_date_range(parent):
    win = tk.Toplevel(parent)
    win.title("비교 기간 선택")
    win.geometry("620x330")
    win.resizable(False, False)

    selected = {"start": None, "end": None}

    tk.Label(win, text="시작일").grid(row=0, column=0, padx=10, pady=(10, 0))
    tk.Label(win, text="종료일").grid(row=0, column=1, padx=10, pady=(10, 0))

    cal_start = Calendar(win, selectmode="day", date_pattern="yyyy-mm-dd")
    cal_start.grid(row=1, column=0, padx=10, pady=10)

    cal_end = Calendar(win, selectmode="day", date_pattern="yyyy-mm-dd")
    cal_end.grid(row=1, column=1, padx=10, pady=10)

    def confirm():
        start = pd.to_datetime(cal_start.get_date()).normalize()
        end = pd.to_datetime(cal_end.get_date()).normalize()

        if start > end:
            messagebox.showerror("오류", "시작일이 종료일보다 늦습니다.", parent=win)
            return

        selected["start"] = start
        selected["end"] = end
        win.destroy()

    def no_filter():
        selected["start"] = None
        selected["end"] = None
        win.destroy()

    def on_close():
        selected["start"] = None
        selected["end"] = None
        win.destroy()

    tk.Button(win, text="확인", width=15, command=confirm).grid(row=2, column=0, pady=10)
    tk.Button(win, text="기간 필터 없이 진행", width=20, command=no_filter).grid(row=2, column=1, pady=10)

    win.protocol("WM_DELETE_WINDOW", on_close)

    # 창이 뒤로 숨는 것 방지
    win.update_idletasks()
    win.lift()
    win.attributes("-topmost", True)
    win.after(300, lambda: win.attributes("-topmost", False))
    win.focus_force()
    win.grab_set()

    parent.wait_window(win)

    return selected["start"], selected["end"]


def make_compare_key(process: str, workshop: str, team: str, part_no: str, day) -> str:
    """
    액세서리는 작업장명/작업반명/공정인쇄 차이로 비교가 자주 틀어지므로
    완성 + 출하-액세서리 는 공정 + 품번 + 날짜만으로 비교한다.
    나머지는 기존처럼 공정 + 작업장명 + 작업반명 + 품번 + 날짜 사용.
    """
    process = normalize_process_token(process)
    workshop = normalize_process_token(workshop)
    team = normalize_process_token(team)
    part_no = normalize_part_no(part_no)
    day = pd.to_datetime(day, errors="coerce").normalize()

    if process == "완성" and workshop == "출하-액세서리":
        return f"{process}|{part_no}|{day}"

    return f"{process}|{workshop}|{team}|{part_no}|{day}"


def normalize_today_actual_process(v: str) -> Optional[str]:
    t = normalize_process_token(v).replace(" ", "").upper()

    if t in ("완성조립", "완성"):
        return "완성"
    if t in ("TANK조립", "TANK", "용접"):
        return "TANK"
    if t in ("CORE조립", "CORE"):
        return "CORE"

    return None

def make_fifo_group_key(process: str, workshop: str, team: str, part_no: str) -> str:
    process = normalize_process_token(process)
    workshop = canonical_workshop_name(workshop)
    team = normalize_process_token(team)
    part_no = normalize_part_no(part_no)

    # 액세서리는 작업반 제외
    if process == "완성" and workshop == "출하-액세서리":
        return f"{process}|{part_no}"

    return f"{process}|{workshop}|{team}|{part_no}"
def canonical_workshop_name(name: str) -> str:
    return (
        normalize_process_token(name)
        .replace("악세서리", "액세서리")
        .replace("액세사리", "액세서리")
    )


def map_process_from_workshop(name: str) -> Optional[str]:
    return WORKSHOP_TO_PROCESS.get(normalize_process_token(name))


def extract_plan_all(plan_file: str, progress=None) -> pd.DataFrame:
    wb = load_workbook(plan_file, data_only=True)
    frames = []
    total = len(SHEET_CONFIG)

    for i, (sheet_name, config) in enumerate(SHEET_CONFIG.items(), start=1):
        if sheet_name not in wb.sheetnames:
            if progress:
                progress(20 + int(i / total * 25), f"{sheet_name} 시트 없음, 건너뜀")
            continue

        if progress:
            progress(20 + int((i - 1) / total * 25), f"{sheet_name} 계획 추출 중...")

        ws = wb[sheet_name]
        df_sheet = extract_plan_sheet_from_module(
            ws,
            config,
            check_red_font=CHECK_RED_FONT,
            progress=lambda pct, text, i=i, total=total: progress(
                20 + int(((i - 1) + pct / 100) / total * 25),
                text,
            ) if progress else None,
        )
        if progress:
            progress(20 + int(i / total * 25), f"{sheet_name} 계획 추출 완료: {len(df_sheet)}건")
        frames.append(df_sheet)

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


# =========================
# MES 정규화
# =========================
def extract_mes(mes_file: str) -> pd.DataFrame:
    df = pd.read_excel(mes_file, sheet_name="Sheet")

    needed = ["계획일", "작업장명", "작업반명", "작업지시상태", "품번", "지시량", "실적량"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise RuntimeError(f"MES 파일에 필요한 컬럼이 없습니다: {missing}")

    df = df[needed].copy()
    df["작업장명"] = df["작업장명"].map(canonical_workshop_name)
    df["공정"] = df["작업장명"].map(map_process_from_workshop)

    # 악세서리 계열 통일
    accessory_mask = df["작업장명"].astype(str).str.contains("액세서리", na=False)
    df.loc[accessory_mask, "작업장명"] = "출하-액세서리"
    df.loc[accessory_mask, "작업반명"] = "출하-액세서리"

    df = df[df["공정"].notna()].copy()
    df = df[df["작업지시상태"] != "종료"].copy()

    df["날짜"] = pd.to_datetime(df["계획일"], errors="coerce").dt.normalize()
    df["지시량"] = pd.to_numeric(df["지시량"], errors="coerce").fillna(0)
    df["실적량"] = pd.to_numeric(df["실적량"], errors="coerce").fillna(0)
    df["작업장명"] = df["작업장명"].apply(canonical_workshop_name)
    df["품번"] = df["품번"].apply(normalize_part_no)
    df["작업반명"] = df["작업반명"].fillna("").astype(str).str.strip()

    df["비교키"] = df.apply(
        lambda x: make_compare_key(
            x["공정"],
            x["작업장명"],
            x["작업반명"],
            x["품번"],
            x["날짜"]
        ),
        axis=1
    )

    grouped = (
        df.groupby(["비교키", "공정", "작업장명", "작업반명", "품번", "날짜"], as_index=False)
        .agg({"지시량": "sum", "실적량": "sum"})
    )

    return grouped.sort_values(
        ["공정", "작업장명", "작업반명", "품번", "날짜"]
    ).reset_index(drop=True)
# =========================
#  오늘 실적 정규화
# =========================

def extract_today_actual(today_actual_file: str) -> pd.DataFrame:
    if not today_actual_file:
        return pd.DataFrame(columns=["공정", "품번", "오늘실적수량"])

    df = pd.read_excel(today_actual_file)

    needed = ["공정", "품번", "실적수량"]
    missing = [c for c in needed if c not in df.columns]
    if missing:
        raise RuntimeError(f"오늘 실적 파일에 필요한 컬럼이 없습니다: {missing}")

    df = df[needed].copy()
    df["공정"] = df["공정"].apply(normalize_today_actual_process)
    df["품번"] = df["품번"].apply(normalize_part_no)
    df["실적수량"] = pd.to_numeric(df["실적수량"], errors="coerce").fillna(0)

    df = df[df["공정"].notna()].copy()

    grouped = (
        df.groupby(["공정", "품번"], as_index=False)
          .agg(오늘실적수량=("실적수량", "sum"))
    )

    return grouped
# =========================
# 계획 집계
# =========================
def build_plan_compare_base(plan_df: pd.DataFrame) -> pd.DataFrame:
    if plan_df.empty:
        return pd.DataFrame(columns=[
            "공정", "작업장명", "작업반명", "날짜", "품번",
            "미달수량", "계획수량", "실적수량", "잔량판단합계"
        ])

    df = plan_df.copy()
    df["날짜"] = pd.to_datetime(df["날짜"], errors="coerce").dt.normalize()
    df["품번"] = df["품번"].astype(str).str.strip()
    df["구분"] = df["구분"].astype(str).str.strip()
    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0)

    grouped = (
        df.groupby(["비교키", "공정", "작업장명", "작업반명", "날짜", "품번", "구분"], as_index=False)
          .agg(수량=("수량", "sum"))
    )

    pivot = grouped.pivot_table(
        index=["비교키", "공정", "작업장명", "작업반명", "날짜", "품번"],
        values="수량",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    for col in ["미달", "계획", "실적"]:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot = pivot.rename(columns={
        "미달": "미달수량",
        "계획": "계획수량",
        "실적": "실적수량",
    })
    pivot["잔량판단합계"] = pivot["미달수량"] + pivot["계획수량"] + pivot["실적수량"]

    return pivot.sort_values(
        ["공정", "작업장명", "작업반명", "품번", "날짜"]
    ).reset_index(drop=True)


# =========================
# 비교 로직
# =========================
def compare_plan_mes_with_fifo(plan_df: pd.DataFrame, mes_df: pd.DataFrame, today_actual_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    """
    비교 핵심 로직

    계산 기준:
    1. 같은 공정/작업장/작업반/품번 단위로 날짜순 계산
    2. 음수 미달 = 선생산 → 미래 계획에 먼저 사용
    3. 양수 미달 = 과거 계획 잔량 스냅샷
       - 이전 open plan 잔량이 없을 때만 backlog로 봄
    4. 계획표 실적은 가장 오래된 미완료 계획부터 FIFO 차감
    5. MES 작업지시가 0이어도 계획표 실적으로 계획이 소진되었으면 정상
    6. MES 작업지시가 없고 미달흐름이 맞으면 미달이월 정상으로 인정
    """

    if plan_df.empty:
        return pd.DataFrame(columns=FINAL_COMPARE_COLUMNS)

    plan = plan_df.copy()
    mes = mes_df.copy()

    key_cols = ["비교키", "공정", "작업장명", "작업반명", "품번", "날짜"]

    # =========================
    # 기본 정규화
    # =========================
    plan["날짜"] = pd.to_datetime(plan["날짜"], errors="coerce").dt.normalize()
    mes["날짜"] = pd.to_datetime(mes["날짜"], errors="coerce").dt.normalize()

    plan["수량"] = pd.to_numeric(plan["수량"], errors="coerce").fillna(0)
    mes["지시량"] = pd.to_numeric(mes["지시량"], errors="coerce").fillna(0)
    mes["실적량"] = pd.to_numeric(mes["실적량"], errors="coerce").fillna(0)

    plan["품번"] = plan["품번"].apply(normalize_part_no)
    mes["품번"] = mes["품번"].apply(normalize_part_no)

    # =========================
    # 오늘 실적 맵
    # =========================
    today_actual_map = {}

    if today_actual_df is not None and not today_actual_df.empty:
        tmp = today_actual_df.copy()
        tmp["품번"] = tmp["품번"].apply(normalize_part_no)
        tmp["공정"] = tmp["공정"].astype(str).str.strip()

        today_actual_map = {
            (row["공정"], row["품번"]): safe_float(row["오늘실적수량"])
            for _, row in tmp.iterrows()
        }

    # =========================
    # 계획 피벗: 미달/계획/실적 분리
    # =========================
    pivot = (
        plan.groupby(key_cols + ["구분"], as_index=False)
            .agg(수량=("수량", "sum"))
            .pivot_table(
                index=key_cols,
                columns="구분",
                values="수량",
                aggfunc="sum",
                fill_value=0
            )
            .reset_index()
    )

    for col in ["계획", "미달", "실적"]:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot = pivot.rename(columns={
        "계획": "계획수량",
        "미달": "미달수량",
        "실적": "계획표실적수량",
    })

    # FIFO 계산 그룹키
    pivot["FIFO그룹키"] = pivot.apply(
        lambda x: make_fifo_group_key(
            x["공정"],
            x["작업장명"],
            x["작업반명"],
            x["품번"]
        ),
        axis=1
    )

    # 공정인쇄 표시용
    plan_info = (
        plan[plan["구분"] == "계획"]
        .groupby(key_cols, as_index=False)
        .agg(
            공정인쇄=(
                "공정인쇄",
                lambda s: " / ".join(sorted({x for x in s if normalize_process_token(x)}))
            )
        )
    )

    pivot = pivot.merge(plan_info, on=key_cols, how="left")
    pivot["공정인쇄"] = pivot["공정인쇄"].fillna("")

    # =========================
    # MES merge
    # =========================
    mes2 = mes.rename(columns={
        "지시량": "작업지시수량",
        "실적량": "MES실적량_원본"
    })

    pivot = pivot.merge(
        mes2[key_cols + ["작업지시수량", "MES실적량_원본"]],
        on=key_cols,
        how="left"
    )

    pivot["작업지시수량"] = pd.to_numeric(
        pivot["작업지시수량"], errors="coerce"
    ).fillna(0)

    pivot["MES실적량_원본"] = pd.to_numeric(
        pivot["MES실적량_원본"], errors="coerce"
    ).fillna(0)

    result_rows = []

    # =========================
    # 그룹별 FIFO 계산
    # =========================
    for _, g in pivot.groupby("FIFO그룹키", dropna=False, sort=False):
        g = g.sort_values("날짜").reset_index(drop=True)

        group_process = normalize_process_token(g.loc[0, "공정"])
        group_part = normalize_part_no(g.loc[0, "품번"])

        today_actual_remain = safe_float(
            today_actual_map.get((group_process, group_part), 0.0)
        )

        # 음수 미달, 즉 선생산 잔량
        prebuild_pool = 0

        # 양수 미달. 이전 계획 잔량이 없는 경우에만 backlog로 사용
        backlog_midal_pool = 0

        # 미완료 계획 FIFO 큐
        open_plans = []

        # 계획행별 계산결과 저장
        row_states = {}

        # =========================
        # 1차 패스: 선생산/미달/실적 배분
        # =========================
        for i, row in g.iterrows():
            plan_qty = safe_float(row["계획수량"])
            midal_qty = safe_float(row["미달수량"])
            plan_sheet_actual_qty = abs(safe_float(row["계획표실적수량"]))
            mes_actual_qty = abs(safe_float(row["MES실적량_원본"]))
            work_qty = safe_float(row["작업지시수량"])

            # 완전 0행 제거
            if (
                plan_qty == 0
                and midal_qty == 0
                and plan_sheet_actual_qty == 0
                and mes_actual_qty == 0
                and work_qty == 0
            ):
                continue

            # MES 실적에 오늘 실적이 이미 포함된 경우, 비교용으로 차감
            deduct_today = min(mes_actual_qty, today_actual_remain)
            mes_actual_for_compare = mes_actual_qty - deduct_today
            today_actual_remain -= deduct_today

            # 음수 미달 = 선생산
            if midal_qty < 0:
                prebuild_pool += abs(midal_qty)

            # 현재 미완료 계획 잔량
            open_remain_sum = sum(p["remain"] for p in open_plans)

            # 양수 미달은 과거 계획 잔량.
            # 이미 open plan 잔량이 있으면 중복 반영하지 않음.
            if midal_qty > 0 and open_remain_sum == 0:
                backlog_midal_pool = midal_qty

            used_prebuild_for_this_plan = 0

            # 계획 등록
            if plan_qty > 0:
                used_prebuild_for_this_plan = min(prebuild_pool, plan_qty)
                prebuild_pool -= used_prebuild_for_this_plan

                open_plan = {
                    "row_index": i,
                    "plan_qty": plan_qty,
                    "remain": plan_qty - used_prebuild_for_this_plan,
                    "used_prebuild": used_prebuild_for_this_plan,
                    "used_actual": 0,
                }

                open_plans.append(open_plan)

                row_states[i] = {
                    "row": row,
                    "plan_qty": plan_qty,
                    "midal_qty": midal_qty,
                    "current_midal": max(midal_qty, 0),
                    "plan_sheet_actual_qty": plan_sheet_actual_qty,
                    "mes_actual_qty": mes_actual_qty,
                    "mes_actual_for_compare": mes_actual_for_compare,
                    "work_qty": work_qty,
                    "used_prebuild": used_prebuild_for_this_plan,
                    "used_actual": 0,
                }

            # 실적 배분 시작
            remain_actual = plan_sheet_actual_qty

            # 이전 기간 미달 backlog가 있으면 먼저 차감
            if backlog_midal_pool > 0 and remain_actual > 0:
                used_for_backlog = min(backlog_midal_pool, remain_actual)
                backlog_midal_pool -= used_for_backlog
                remain_actual -= used_for_backlog

            # 남은 실적은 가장 오래된 계획부터 FIFO 차감
            for p in open_plans:
                if remain_actual <= 0:
                    break

                if p["remain"] <= 0:
                    continue

                used = min(p["remain"], remain_actual)

                p["remain"] -= used
                p["used_actual"] += used
                remain_actual -= used

                target_i = p["row_index"]

                if target_i in row_states:
                    row_states[target_i]["used_actual"] = p["used_actual"]

        # =========================
        # 2차 패스: 판정 생성
        # =========================
        for i, state in row_states.items():
            row = state["row"]

            plan_qty = state["plan_qty"]
            midal_qty = state["midal_qty"]
            current_midal = state["current_midal"]
            plan_sheet_actual_qty = state["plan_sheet_actual_qty"]
            mes_actual_qty = state["mes_actual_qty"]
            mes_actual_for_compare = state["mes_actual_for_compare"]
            work_qty = state["work_qty"]

            used_prebuild = state["used_prebuild"]
            plan_actual_for_today = state["used_actual"]

            completed_qty = used_prebuild + plan_actual_for_today

            # =========================
            # 미달 흐름 판정
            # =========================
            calculated_next_midal = midal_qty + plan_qty - plan_sheet_actual_qty

            future_midal_rows = g.loc[i + 1:].copy()
            future_midal_rows = future_midal_rows[
                future_midal_rows["미달수량"].apply(lambda x: safe_float(x) != 0)
            ]

            if not future_midal_rows.empty:
                next_midal = safe_float(future_midal_rows.iloc[0]["미달수량"])

                if abs(calculated_next_midal - next_midal) < 0.000001:
                    midal_judge = "미달흐름일치"
                else:
                    midal_judge = "미달흐름불일치"
            else:
                midal_judge = "미달흐름확인제외"

            midal_flow_ok = midal_judge in ("미달흐름일치", "미달흐름확인제외")

            # =========================
            # 작업지시 판정
            # =========================
            if work_qty == plan_qty:
                work_judge = "작업지시일치"

            elif work_qty == 0 and completed_qty >= plan_qty:
                work_judge = "작업지시일치후소멸"

            elif (
                work_qty == 0
                and plan_qty > 0
                and midal_qty > 0
                and plan_sheet_actual_qty > 0
                and midal_flow_ok
            ):
                work_judge = "작업지시일치후미달이월"

            elif work_qty < plan_qty:
                work_judge = "작업지시부족"

            else:
                work_judge = "작업지시과다"

            # =========================
            # 실적 판정
            # MES 작업지시가 소멸된 케이스는 MES 실적 비교 제외
            # =========================
            if work_judge in ("작업지시일치후소멸", "작업지시일치후미달이월"):
                actual_judge = "실적확인제외_MES소멸"
            else:
                if abs(plan_sheet_actual_qty - mes_actual_for_compare) < 0.000001:
                    actual_judge = "실적일치"
                else:
                    actual_judge = "실적불일치"

            # =========================
            # 최종 판정
            # =========================
            if (
                work_judge in ("작업지시일치", "작업지시일치후소멸", "작업지시일치후미달이월")
                and actual_judge in ("실적일치", "실적확인제외_MES소멸")
                and midal_judge in ("미달흐름일치", "미달흐름확인제외")
            ):
                judge = "일치"
            else:
                judge = f"{work_judge}/{actual_judge}/{midal_judge}"

            result_rows.append({
                "품번": normalize_part_no(row["품번"]),
                "날짜": row["날짜"],
                "공정": row["공정"],
                "작업장명": row["작업장명"],
                "작업반명": row["작업반명"],
                "공정인쇄": row["공정인쇄"],
                "계획수량": plan_qty,
                "미달수량": current_midal,
                "계획표실적수량": plan_sheet_actual_qty,
                "MES실적수량": mes_actual_qty,
                "전일마감기준실적수량": mes_actual_for_compare,
                "계획대비실적수량": plan_actual_for_today,
                "작업지시수량": work_qty,
                "판정": judge,
            })

    result = pd.DataFrame(result_rows)

    # =========================
    # 결과 0행 제거
    # =========================
    if not result.empty:
        qty_cols = [
            "계획수량",
            "미달수량",
            "계획표실적수량",
            "MES실적수량",
            "전일마감기준실적수량",
            "계획대비실적수량",
            "작업지시수량",
        ]

        for c in qty_cols:
            result[c] = pd.to_numeric(result[c], errors="coerce").fillna(0)

        result = result[result[qty_cols].abs().sum(axis=1) != 0].copy()

    # =========================
    # 계획 없는 MES 작업지시 처리
    # 단, 이후 양수 미달로 설명 가능하면 일치 처리
    # =========================
    plan_keys = (
        pivot[pivot["계획수량"] > 0][key_cols]
        .drop_duplicates()
        .assign(계획존재=1)
    )

    mes_only = mes2.merge(plan_keys, on=key_cols, how="left")
    mes_only = mes_only[mes_only["계획존재"].isna()].copy()

    future_midal_allowance = {}
    positive_midal = pivot[pivot["미달수량"] > 0].copy()

    for _, r in positive_midal.iterrows():
        gkey = make_fifo_group_key(
            r["공정"],
            r["작업장명"],
            r["작업반명"],
            r["품번"]
        )

        future_midal_allowance[gkey] = future_midal_allowance.get(gkey, 0) + safe_float(r["미달수량"])

    if not mes_only.empty:
        mes_only["공정인쇄"] = ""
        mes_only["계획수량"] = 0
        mes_only["미달수량"] = 0
        mes_only["계획표실적수량"] = 0

        mes_only["MES실적수량"] = pd.to_numeric(
            mes_only["MES실적량_원본"], errors="coerce"
        ).fillna(0).abs()

        mes_only["전일마감기준실적수량"] = mes_only["MES실적수량"]
        mes_only["계획대비실적수량"] = 0

        remain_future_midal = future_midal_allowance.copy()
        judge_list = []

        mes_only = mes_only.sort_values(["공정", "작업장명", "작업반명", "품번", "날짜"]).copy()

        for _, r in mes_only.iterrows():
            gkey = make_fifo_group_key(
                r["공정"],
                r["작업장명"],
                r["작업반명"],
                r["품번"]
            )

            work_qty = safe_float(r["작업지시수량"])
            remain_midal = remain_future_midal.get(gkey, 0)

            if remain_midal > 0:
                use_qty = min(remain_midal, work_qty)
                remain_future_midal[gkey] -= use_qty

                if work_qty <= remain_midal:
                    judge = "일치"
                else:
                    judge = "부분일치_초과작업지시존재"
            else:
                judge = "해당날짜_계획없는데작업지시있음"

            judge_list.append(judge)

        mes_only["판정"] = judge_list

        mes_only = mes_only[
            (
                mes_only["작업지시수량"].abs()
                + mes_only["MES실적수량"].abs()
            ) != 0
        ].copy()

        mes_only = mes_only[FINAL_COMPARE_COLUMNS]

        result = pd.concat([result, mes_only], ignore_index=True)

    # =========================
    # 최종 정리
    # =========================
    if result.empty:
        return pd.DataFrame(columns=FINAL_COMPARE_COLUMNS)

    result = result[FINAL_COMPARE_COLUMNS]
    result["날짜"] = pd.to_datetime(result["날짜"], errors="coerce").dt.strftime("%Y-%m-%d")

    return result.sort_values(
        ["공정", "작업장명", "작업반명", "품번", "날짜"]
    ).reset_index(drop=True)


# =========================
# 저장
# =========================
def save_results(plan_df, mes_df, plan_base_df, compare_df, output_file: str):
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        plan_df.to_excel(writer, sheet_name="1_계획정규화", index=False)
        mes_df.to_excel(writer, sheet_name="2_MES정규화", index=False)
        plan_base_df.to_excel(writer, sheet_name="3_계획집계", index=False)
        compare_df.to_excel(writer, sheet_name="4_비교결과", index=False)

        if not compare_df.empty:
            summary = (
                compare_df.groupby("판정", as_index=False)
                .size()
                .rename(columns={"size": "건수"})
            )
            summary.to_excel(writer, sheet_name="5_요약", index=False)


# =========================
# 실행
# =========================
def main():
    root = tk.Tk()
    root.withdraw()

    root.attributes("-topmost", True)
    root.lift()
    root.focus_force()

    plan_file = select_excel_file("생산계획 엑셀 선택", root)
    if not plan_file:
        messagebox.showinfo("알림", "생산계획 엑셀을 선택하지 않았습니다.")
        return

    mes_file = select_excel_file("MES 작업지시 엑셀 선택", root)
    if not mes_file:
        messagebox.showinfo("알림", "MES 작업지시 엑셀을 선택하지 않았습니다.")
        return
    
    START_DATE, END_DATE = select_date_range(root)

    root.update()  # 날짜 선택 창이 닫힌 후에 업데이트

    global CHECK_RED_FONT
    CHECK_RED_FONT = messagebox.askyesno("빨간 글씨 체크 제외", "계획표에서 빨간 글씨로 표시된 셀을 제외하시겠습니까?\n(수락하면 빨간 글씨는 수량 계산에서 제외됩니다)")
    today_result_file = messagebox.askyesno("실적 적용 여부", "output.xlsx 같은 실적 파일을 불러와서\n작업지시 수량에 반영하시겠습니까?")
    today_actual_file = None
    if today_result_file:
        today_actual_file = filedialog.askopenfilename(title="오늘 실적 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not today_actual_file:
            messagebox.showinfo("알림", "오늘 실적 파일을 선택하지 않았습니다. 계속 진행합니다.")
            today_actual_file = None

    progress_win, progress = create_progress_window(root, "MES 작업지시 비교 진행")

    try:
        progress(5, "오늘 실적 정규화 중...")
        today_actual_df = extract_today_actual(today_actual_file)
        progress(10, f"오늘 실적 정규화 완료: {len(today_actual_df)}건")

        progress(20, "생산계획 정규화 중...")
        plan_df = extract_plan_all(plan_file, progress=progress)
        plan_df = filter_by_period(plan_df, start_date=START_DATE, end_date=END_DATE)
        progress(50, f"계획 정규화 완료: {len(plan_df)}건")

        progress(55, "MES 작업지시 정규화 중...")
        mes_df = extract_mes(mes_file)
        mes_df = filter_by_period(mes_df, start_date=START_DATE, end_date=END_DATE)
        progress(65, f"MES 정규화 완료: {len(mes_df)}건")

        progress(72, "계획 비교기준 생성 중...")
        plan_base_df = build_plan_compare_base(plan_df)
        progress(78, f"계획 비교기준 생성 완료: {len(plan_base_df)}건")

        progress(84, "계획과 MES 비교 중...")
        compare_df = compare_plan_mes_with_fifo(plan_df, mes_df, today_actual_df)
        compare_df = filter_by_period(compare_df, start_date=START_DATE, end_date=END_DATE)
        progress(92, f"비교 결과 생성 완료: {len(compare_df)}건")

        progress(96, "엑셀 저장 중...")
        save_results(plan_df, mes_df, plan_base_df, compare_df, OUTPUT_FILE)
        progress(100, "완료")
        progress_win.destroy()
        messagebox.showinfo("완료", f"저장 완료: {OUTPUT_FILE}")

    except Exception as e:
        progress_win.destroy()
        messagebox.showerror("오류", str(e))


if __name__ == "__main__":
    main()
