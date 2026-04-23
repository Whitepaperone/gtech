# -*- coding: utf-8 -*-
"""
생산계획 3개 시트 + MES 작업지시 비교

비교 대상
- CORE공정(실적)
- TANK공정(실적)
- 완성공정(실적) + 액세서리 블록

핵심 룰
- 세 시트를 한 번에 읽음
- 헤더/날짜열/액세서리 시작행을 동적으로 찾음
- 공정 매핑은 사용자 지정 룰 적용
- 품번 "2", HH 시작, 개발/기타, 클린칭C 등은 스킵
- MES는 생산로트번호(YYYYMMDD)를 날짜로 사용
- 종료/강제종료도 포함
- 단, "강제종료 + 생산완료 0" 은 삭제성 지시로 따로 집계

주의
1) 계획 파일이 수식 기반인데 캐시값이 비어 있으면 openpyxl(data_only=True)로 숫자가 안 읽힙니다.
   - 가장 쉬운 방법: 계획 파일을 Excel에서 한 번 열고 '저장' 후 실행
   - 또는 LibreOffice가 설치되어 있으면 USE_LIBREOFFICE_RECALC=True 로 사용
2) 완성공정 시트는 일부 행에서 "분류"가 비어 있을 수 있어, 그 경우 team 비교를 비워두고 workplace(완성조립) 기준으로 비교합니다.
"""

import os
import re
import shutil
import subprocess
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from tkinter import filedialog

from openpyxl import Workbook, load_workbook


# =========================
# 사용자 설정
# =========================
PLAN_FILE = filedialog.askopenfilename(
        title="계획 파일 선택",
        filetypes=[("Excel files", "*.xlsx")]
    )
MES_FILE = filedialog.askopenfilename(
        title="MES 파일 선택",
        filetypes=[("Excel files", "*.xlsx")]
    )
OUTPUT_FILE = r"./plan_mes_compare_output.xlsx"

USE_LIBREOFFICE_RECALC = True   # True면 LibreOffice로 재계산 저장 시도
LIBREOFFICE_EXE = "soffice"      # Windows면 보통 "C:\\Program Files\\LibreOffice\\program\\soffice.exe"

PLAN_SHEETS = ["완성공정(실적)", "TANK공정(실적)", "CORE공정(실적)"]

VALID_MES_STATUSES = {"생성", "확정", "진행", "종료", "강제종료", "보류"}
ACTIVE_STATUSES = {"생성", "확정", "진행", "보류"}
CLOSED_STATUSES = {"종료", "강제종료"}


# =========================
# 공통 유틸
# =========================
def to_text(v):
    if v is None:
        return ""
    return str(v).strip()


def to_number(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0
    try:
        return float(s)
    except Exception:
        return 0


def normalize_text(v):
    return to_text(v).replace("\n", "").strip()


def normalize_process_text(v):
    return normalize_text(v).replace(" ", "").upper()


def normalize_part_no(v):
    return to_text(v).replace(" ", "").upper()


def parse_date_header(v, default_year=2026):
    if v is None:
        return None

    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    if not s:
        return None

    fmts = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass

    # 월/일만 있는 경우
    for fmt in ("%m/%d", "%m-%d"):
        try:
            d = datetime.strptime(s, fmt)
            return d.replace(year=default_year).date()
        except Exception:
            pass

    return None


def should_skip_part_no(part_no):
    pn = normalize_part_no(part_no)
    if pn == "":
        return True
    if pn == "2":
        return True
    if pn.startswith("HH"):
        return True
    return False


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 120) + 1):
        row_text = [to_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 100) + 1)]
        joined = "|".join(row_text)
        if "모델" in joined and "CORE" in joined and "완성" in joined and "품번" in joined:
            return r
    raise RuntimeError(f"[{ws.title}] 헤더 행을 찾지 못했습니다.")


def build_column_map(ws, header_row):
    """
    2줄 헤더(row, row+1) 합쳐서 컬럼을 찾음
    같은 이름이 여러 번 나오면 첫 번째 컬럼 우선
    """
    col_map = {}

    for c in range(1, ws.max_column + 1):
        top = to_text(ws.cell(header_row, c).value)
        bot = to_text(ws.cell(header_row + 1, c).value)
        combo = f"{top}|{bot}"

        def set_once(key):
            if key not in col_map:
                col_map[key] = c

        if top == "모델":
            set_once("model")
        elif "고객사" in combo and "품번" in combo:
            set_once("cust_part_no")
        elif "CORE" in combo and "품번" in combo:
            set_once("core_part_no")
        elif "TANK" in combo and "품번" in combo:
            set_once("tank_part_no")
        elif "완성" in combo and "품번" in combo:
            set_once("finish_part_no")
        elif "공정" in combo and "인쇄" in combo:
            set_once("process_print")
        elif "분류" in combo:
            set_once("finish_category")
        elif "악세" in combo and "품번" in combo:
            set_once("acc_set_part_no")
        elif top == "SET 품번":
            set_once("acc_set_part_no")
        elif top == "완성품번":
            set_once("finish_part_no")

    return col_map


def find_date_metric_columns(ws, header_row):
    cols = []
    for c in range(1, ws.max_column + 1):
        d = parse_date_header(ws.cell(header_row, c).value)
        metric = to_text(ws.cell(header_row + 1, c).value)
        if d and metric in {"계획", "실적", "미달"}:
            cols.append({"col": c, "date": d, "metric": metric})
    return cols


def find_accessory_start_row(ws):
    for r in range(1, ws.max_row + 1):
        row_vals = [to_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 30) + 1)]
        joined = "|".join(row_vals)
        if "액세서리" in joined and "HEAT SCREEN" in joined.upper():
            return r
    return None


def is_deleted_like(status, prod_done_qty):
    # 실수로 만들었다가 삭제한 건 대개 강제종료 + 생산완료 0 형태로 보이므로 별도 표기
    return status == "강제종료" and to_number(prod_done_qty) == 0


def maybe_recalc_with_libreoffice(src_path):
    if not USE_LIBREOFFICE_RECALC:
        return src_path

    src = Path(src_path)
    out_dir = src.parent / "_recalc_tmp"
    out_dir.mkdir(exist_ok=True)

    temp_copy = out_dir / src.name
    shutil.copy2(src, temp_copy)

    cmd = [
        LIBREOFFICE_EXE,
        "--headless",
        "--convert-to", "xlsx",
        "--outdir", str(out_dir),
        str(temp_copy),
    ]
    subprocess.run(cmd, check=True)

    recalc_path = out_dir / src.name
    if not recalc_path.exists():
        raise RuntimeError("LibreOffice 재계산 저장 실패")
    return str(recalc_path)


# =========================
# 공정 매핑
# =========================
def map_core_process(process_value):
    p = normalize_process_text(process_value)
    if p.startswith("AL"):
        return ("CORE조립", "CORE조립-AL")
    if p.startswith("CU"):
        return ("CORE조립", "CORE조립-CU")
    return (None, None)


def map_tank_process(process_value):
    p = normalize_process_text(process_value)

    if p == "수동".upper():
        return ("용접", "TANK용접 - 수동")
    if p in {"로봇".upper(), "ROBOT"}:
        return ("용접", "TANK용접 - ROBOT")
    if p == "CLINCHING":
        return ("CLINCHING", "CLINCHING")
    if p == "클린칭C".upper():
        return (None, None)   # 스킵
    if p == "TANK조립".upper():
        return ("TANK조립&Leak Test", "TANK조립&Leak Test")
    if p == "한국".upper():
        return ("용접", "TANK용접 - 한국RAD")
    if p in {"개발,기타".upper(), "개발".upper(), "기타".upper()}:
        return (None, None)

    return (None, None)


def map_finish_category(category_value):
    p = normalize_process_text(category_value)

    if p == "일반".upper():
        return ("완성조립", "완성조립공정 - 일반")
    if p == "G2":
        return ("완성조립", "완성조립공정 - G2")
    if p == "클라크".upper():
        return ("완성조립", "완성조립공정 - 클라크")
    if p == "특수품".upper():
        return ("완성조립", "완성조립공정 - 특수품")
    if p in {"개발,기타".upper(), "개발".upper(), "기타".upper()}:
        return (None, None)

    return (None, None)


# =========================
# 계획 추출
# =========================
def extract_plan_rows(plan_file):
    plan_file = maybe_recalc_with_libreoffice(plan_file)
    wb = load_workbook(plan_file, data_only=True)

    records = []

    for sheet_name in PLAN_SHEETS:
        ws = wb[sheet_name]
        header_row = find_header_row(ws)
        col_map = build_column_map(ws, header_row)
        date_cols = find_date_metric_columns(ws, header_row)
        acc_start_row = find_accessory_start_row(ws)

        if not date_cols:
            raise RuntimeError(f"[{sheet_name}] 날짜/계획/실적/미달 열을 찾지 못했습니다.")

        # 본체 구간
        start_row = header_row + 2
        end_row = ws.max_row
        if sheet_name == "완성공정(실적)" and acc_start_row:
            end_row = acc_start_row - 1

        for r in range(start_row, end_row + 1):
            if sheet_name == "CORE공정(실적)":
                part_no = to_text(ws.cell(r, col_map["core_part_no"]).value)
                process_value = to_text(ws.cell(r, col_map["process_print"]).value)
                workplace, team = map_core_process(process_value)
                source_group = "CORE"
                source_process_value = process_value

            elif sheet_name == "TANK공정(실적)":
                part_no = to_text(ws.cell(r, col_map["tank_part_no"]).value)
                process_value = to_text(ws.cell(r, col_map["process_print"]).value)
                workplace, team = map_tank_process(process_value)
                source_group = "TANK"
                source_process_value = process_value

            else:
                part_no = to_text(ws.cell(r, col_map["finish_part_no"]).value)
                category_col = col_map.get("finish_category")
                category_value = to_text(ws.cell(r, category_col).value) if category_col else ""

                workplace, team = map_finish_category(category_value)

                # 완성 시트는 상세행에 분류가 비어 있는 경우가 있음
                # 이 경우 workplace만 완성조립으로 맞추고 team은 공란 처리
                if not workplace and not team:
                    if normalize_process_text(category_value) in {"개발,기타".upper(), "개발".upper(), "기타".upper()}:
                        workplace, team = (None, None)
                    else:
                        workplace, team = ("완성조립", "")
                source_group = "완성"
                source_process_value = category_value

            if should_skip_part_no(part_no):
                continue
            if workplace is None:
                continue

            for dcol in date_cols:
                c = dcol["col"]
                metric = dcol["metric"]
                dt = dcol["date"]
                qty = to_number(ws.cell(r, c).value)

                if qty == 0:
                    continue

                records.append({
                    "plan_sheet": sheet_name,
                    "source_group": source_group,
                    "source_process_value": source_process_value,
                    "plan_date": dt.strftime("%Y-%m-%d"),
                    "part_no": normalize_part_no(part_no),
                    "mes_workplace": workplace,
                    "mes_team": team,
                    "metric": metric,
                    "qty": qty,
                    "plan_row_no": r,
                })

        # 액세서리 블록
        if sheet_name == "완성공정(실적)" and acc_start_row:
            acc_col_map = build_column_map(ws, acc_start_row)
            set_col = acc_col_map.get("acc_set_part_no")
            if set_col:
                for r in range(acc_start_row + 1, ws.max_row + 1):
                    set_part_no = to_text(ws.cell(r, set_col).value)
                    if should_skip_part_no(set_part_no):
                        continue

                    for dcol in date_cols:
                        c = dcol["col"]
                        metric = dcol["metric"]
                        dt = dcol["date"]
                        qty = to_number(ws.cell(r, c).value)

                        if qty == 0:
                            continue

                        records.append({
                            "plan_sheet": sheet_name,
                            "source_group": "액세서리",
                            "source_process_value": "출하-액세사리",
                            "plan_date": dt.strftime("%Y-%m-%d"),
                            "part_no": normalize_part_no(set_part_no),
                            "mes_workplace": "출하-액세사리",
                            "mes_team": "출하-액세사리",
                            "metric": metric,
                            "qty": qty,
                            "plan_row_no": r,
                        })

    if not records:
        raise RuntimeError(
            "계획 수량을 1건도 읽지 못했습니다. "
            "계획 파일이 수식 기반인데 캐시값이 비어 있을 가능성이 큽니다. "
            "계획 파일을 Excel에서 열어서 저장한 뒤 다시 실행해 주세요."
        )

    return records


# =========================
# MES 추출
# =========================
def extract_mes_rows(mes_file):
    wb = load_workbook(mes_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for r in range(2, ws.max_row + 1):
        status = to_text(ws.cell(r, 7).value)
        if status not in VALID_MES_STATUSES:
            continue

        workplace = to_text(ws.cell(r, 5).value)
        team = to_text(ws.cell(r, 6).value)
        part_no = normalize_part_no(ws.cell(r, 12).value)

        order_qty = to_number(ws.cell(r, 18).value)   # 지시량
        result_qty = to_number(ws.cell(r, 19).value)  # 실적량
        done_qty = to_number(ws.cell(r, 20).value)    # 실적량(생산완료)

        lot_raw = to_text(ws.cell(r, 22).value)       # 생산로트번호 YYYYMMDD
        mes_date = ""
        if re.fullmatch(r"\d{8}", lot_raw):
            mes_date = f"{lot_raw[:4]}-{lot_raw[4:6]}-{lot_raw[6:8]}"

        if not part_no or not mes_date:
            continue

        deleted_like = is_deleted_like(status, done_qty)

        records.append({
            "mes_date": mes_date,
            "part_no": part_no,
            "workplace": workplace,
            "team": team,
            "status": status,
            "order_qty": order_qty,
            "result_qty": result_qty,
            "done_qty": done_qty,
            "deleted_like_qty": order_qty if deleted_like else 0,
            "effective_order_qty": 0 if deleted_like else order_qty,
            "mes_row_no": r,
        })

    return records


# =========================
# 집계 / 비교
# =========================
def aggregate_plan(plan_rows):
    grouped = defaultdict(lambda: {
        "plan_qty": 0,
        "plan_actual_qty": 0,
        "plan_shortage_qty": 0,
        "plan_sheet": "",
        "source_group": "",
        "source_process_value": "",
        "plan_rows": [],
    })

    for rec in plan_rows:
        key = (rec["plan_date"], rec["part_no"], rec["mes_workplace"], rec["mes_team"])
        g = grouped[key]

        g["plan_sheet"] = rec["plan_sheet"]
        g["source_group"] = rec["source_group"]
        g["source_process_value"] = rec["source_process_value"]
        g["plan_rows"].append(rec["plan_row_no"])

        if rec["metric"] == "계획":
            g["plan_qty"] += rec["qty"]
        elif rec["metric"] == "실적":
            g["plan_actual_qty"] += rec["qty"]
        elif rec["metric"] == "미달":
            g["plan_shortage_qty"] += rec["qty"]

    return grouped


def aggregate_mes(mes_rows):
    strict = defaultdict(lambda: {
        "mes_order_qty_all": 0,
        "mes_effective_order_qty": 0,
        "mes_deleted_like_qty": 0,
        "mes_result_qty": 0,
        "mes_done_qty": 0,
        "mes_active_qty": 0,
        "mes_closed_qty": 0,
        "mes_statuses": set(),
        "mes_rows": [],
    })
    loose = defaultdict(lambda: {
        "mes_order_qty_all": 0,
        "mes_effective_order_qty": 0,
        "mes_deleted_like_qty": 0,
        "mes_result_qty": 0,
        "mes_done_qty": 0,
        "mes_active_qty": 0,
        "mes_closed_qty": 0,
        "mes_statuses": set(),
        "mes_rows": [],
    })

    for rec in mes_rows:
        strict_key = (rec["mes_date"], rec["part_no"], rec["workplace"], rec["team"])
        loose_key = (rec["mes_date"], rec["part_no"], rec["workplace"])

        for bucket, key in ((strict, strict_key), (loose, loose_key)):
            g = bucket[key]
            g["mes_order_qty_all"] += rec["order_qty"]
            g["mes_effective_order_qty"] += rec["effective_order_qty"]
            g["mes_deleted_like_qty"] += rec["deleted_like_qty"]
            g["mes_result_qty"] += rec["result_qty"]
            g["mes_done_qty"] += rec["done_qty"]
            if rec["status"] in ACTIVE_STATUSES:
                g["mes_active_qty"] += rec["order_qty"]
            if rec["status"] in CLOSED_STATUSES:
                g["mes_closed_qty"] += rec["order_qty"]
            g["mes_statuses"].add(rec["status"])
            g["mes_rows"].append(rec["mes_row_no"])

    return strict, loose


def build_compare_rows(plan_agg, mes_strict, mes_loose):
    result = []

    for key, p in plan_agg.items():
        plan_date, part_no, mes_workplace, mes_team = key

        if mes_team:
            match_mode = "strict"
            m = mes_strict.get((plan_date, part_no, mes_workplace, mes_team), None)
        else:
            match_mode = "workplace_only"
            m = mes_loose.get((plan_date, part_no, mes_workplace), None)

        if m is None:
            m = {
                "mes_order_qty_all": 0,
                "mes_effective_order_qty": 0,
                "mes_deleted_like_qty": 0,
                "mes_result_qty": 0,
                "mes_done_qty": 0,
                "mes_active_qty": 0,
                "mes_closed_qty": 0,
                "mes_statuses": set(),
                "mes_rows": [],
            }

        diff_plan_vs_effective_order = p["plan_qty"] - m["mes_effective_order_qty"]
        diff_plan_vs_done = p["plan_qty"] - m["mes_done_qty"]

        if diff_plan_vs_effective_order == 0:
            judge = "OK"
        elif diff_plan_vs_effective_order > 0:
            judge = "작업지시 부족"
        else:
            judge = "작업지시 과다"

        caution = []
        if m["mes_deleted_like_qty"] > 0:
            caution.append("강제종료+생산완료0 존재")
        if match_mode == "workplace_only":
            caution.append("완성 team 미지정 비교")

        result.append({
            "plan_date": plan_date,
            "part_no": part_no,
            "mes_workplace": mes_workplace,
            "mes_team": mes_team,
            "match_mode": match_mode,
            "plan_sheet": p["plan_sheet"],
            "source_group": p["source_group"],
            "source_process_value": p["source_process_value"],

            "plan_qty": p["plan_qty"],
            "plan_actual_qty": p["plan_actual_qty"],
            "plan_shortage_qty": p["plan_shortage_qty"],

            "mes_order_qty_all": m["mes_order_qty_all"],
            "mes_effective_order_qty": m["mes_effective_order_qty"],
            "mes_deleted_like_qty": m["mes_deleted_like_qty"],
            "mes_done_qty": m["mes_done_qty"],
            "mes_active_qty": m["mes_active_qty"],
            "mes_closed_qty": m["mes_closed_qty"],
            "mes_statuses": ",".join(sorted(m["mes_statuses"])),

            "diff_plan_vs_effective_order": diff_plan_vs_effective_order,
            "diff_plan_vs_done": diff_plan_vs_done,
            "judge": judge,
            "caution": " / ".join(caution),

            "plan_rows": ",".join(map(str, sorted(set(p["plan_rows"])))),
            "mes_rows": ",".join(map(str, sorted(set(m["mes_rows"])))),
        })

    result.sort(key=lambda x: (x["plan_date"], x["source_group"], x["part_no"]))
    return result


# =========================
# 저장
# =========================
def autosize_columns(ws, max_width=30):
    for col_cells in ws.columns:
        length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            v = "" if cell.value is None else str(cell.value)
            length = max(length, len(v))
        ws.column_dimensions[col_letter].width = min(length + 2, max_width)


def write_sheet(ws, rows):
    if not rows:
        ws.append(["no_data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws.freeze_panes = "A2"
    autosize_columns(ws)


def save_output(plan_rows, mes_rows, compare_rows, output_file):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "plan_raw"
    ws2 = wb.create_sheet("mes_raw")
    ws3 = wb.create_sheet("compare")

    write_sheet(ws1, plan_rows)
    write_sheet(ws2, mes_rows)
    write_sheet(ws3, compare_rows)

    wb.save(output_file)


# =========================
# 실행
# =========================
def main():
    if not PLAN_FILE or not MES_FILE:
        return
    plan_rows = extract_plan_rows(PLAN_FILE)
    mes_rows = extract_mes_rows(MES_FILE)

    plan_agg = aggregate_plan(plan_rows)
    mes_strict, mes_loose = aggregate_mes(mes_rows)
    compare_rows = build_compare_rows(plan_agg, mes_strict, mes_loose)

    save_output(plan_rows, mes_rows, compare_rows, OUTPUT_FILE)

    print(f"plan_rows   : {len(plan_rows)}")
    print(f"mes_rows    : {len(mes_rows)}")
    print(f"compare_rows: {len(compare_rows)}")
    print(f"saved       : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
