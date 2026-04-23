from openpyxl import load_workbook
import pandas as pd
from tkinter import Tk, filedialog, messagebox
from datetime import datetime
import os

OUTPUT_FILE = "./output.xlsx"


def norm(v):
    if v is None:
        return ""
    return str(v).strip().replace(" ", "").replace("\n", "").replace("\r", "")


def to_num(v):
    try:
        if v is None or v == "":
            return 0
        return float(v)
    except:
        return 0


def is_date_like(v):
    if isinstance(v, (datetime, pd.Timestamp)):
        return True
    try:
        dt = pd.to_datetime(v, errors="raise")
        return pd.notna(dt)
    except:
        return False


def build_sheet_cache(ws):
    """
    병합셀 캐시 생성
    - (row, col) -> 실제 값이 들어있는 좌상단 셀 좌표
    """
    merged_anchor_map = {}

    for merged in ws.merged_cells.ranges:
        anchor = (merged.min_row, merged.min_col)
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                merged_anchor_map[(r, c)] = anchor

    return {
        "merged_anchor_map": merged_anchor_map
    }


def get_real_cell(ws, cache, row, col):
    """
    병합셀인 경우 실제 값을 가진 좌상단 셀 반환
    """
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell

    anchor = cache["merged_anchor_map"].get((row, col))
    if anchor:
        return ws.cell(row=anchor[0], column=anchor[1])

    return cell


def get_merged_value(ws, cache, row, col):
    return get_real_cell(ws, cache, row, col).value


def is_red_font(cell):
    """
    빨간 글씨 여부 판단
    """
    try:
        color = cell.font.color
        if color is None:
            return False

        if color.type == "rgb" and color.rgb:
            rgb = str(color.rgb).upper()
            if rgb in ("FFFF0000", "00FF0000"):
                return True

        if color.type == "indexed" and color.indexed is not None:
            if color.indexed == 10:
                return True

        return False
    except:
        return False


def get_qty_value(ws, cache, row, col):
    """
    수량 읽기
    - 완성공정(실적) 시트에서 빨간색 폰트면 0 처리
    """
    cell = get_real_cell(ws, cache, row, col)

    if is_red_font(cell):
        return 0

    return to_num(cell.value)


def should_skip_row(product, process_name):
    p = norm(product).upper()
    proc_raw = norm(process_name)

    if p == "2":
        return True
    if p.startswith("HH"):
        return True

    if "클린칭C" in proc_raw:
        return True
    if "개발,기타" in proc_raw:
        return True

    return False


def map_workcenter_and_team(process_name):
    raw = norm(process_name)
    p = raw.upper().replace(" ", "")

    if "AL" == p or "AL" in p:
        return "CORE조립", "CORE조립-AL"

    if "CU" == p or "CU" in p:
        return "CORE조립", "CORE조립-CU"

    if "수동" in raw:
        return "용접", "TANK용접 - 수동"

    if "로봇" in p:
        return "용접", "TANK용접 - ROBOT"

    if "클린칭" in p:
        return "CLINCHING", "CLINCHING"

    if "TANK조립" in raw:
        return "TANK조립&Leak Test", "TANK조립&Leak Test"

    if "한국" in raw:
        return "용접", "TANK용접 - 한국RAD"

    if "일반" in raw:
        return "완성조립", "완성조립공정 - 일반"

    if "G2" in p:
        return "완성조립", "완성조립공정 - G2"

    if "클라크" in raw:
        return "완성조립", "완성조립공정 - 클라크"

    if "특수품" in raw:
        return "완성조립", "완성조립공정 - 특수품"

    return None, None


def find_plan_and_shortage_columns(ws, cache, process_col):
    shortage_col = None
    plan_cols = []
    date_labels = []

    for c in range(process_col + 1, ws.max_column + 1):
        found = False

        for r in range(1, 5):
            top = get_merged_value(ws, cache, r, c)
            if not is_date_like(top):
                continue

            dt = pd.to_datetime(top).strftime("%Y/%m/%d")

            candidates = []
            if r + 1 <= 5:
                candidates.append(norm(get_merged_value(ws, cache, r + 1, c)))
            if r + 2 <= 5:
                candidates.append(norm(get_merged_value(ws, cache, r + 2, c)))
            candidates.append(norm(get_merged_value(ws, cache, r, c)))

            if "미달" in candidates and shortage_col is None:
                shortage_col = c
                found = True
                break

            if "계획" in candidates:
                plan_cols.append(c)
                date_labels.append(dt)
                found = True
                break

        if not found:
            continue

    if not plan_cols:
        debug = []
        for c in range(max(1, process_col - 1), min(ws.max_column, process_col + 10) + 1):
            vals = []
            for r in range(1, 6):
                vals.append(f"r{r}='{norm(get_merged_value(ws, cache, r, c))}'")
            debug.append(f"col {c}: " + ", ".join(vals))
        raise Exception(f"[{ws.title}] 계획 열을 찾지 못했습니다.\n" + "\n".join(debug))

    return shortage_col, plan_cols, date_labels


def adjust_by_shortage(plan_values, shortage_value):
    result = plan_values[:]
    stock = abs(to_num(shortage_value)) if to_num(shortage_value) < 0 else 0

    if stock <= 0:
        return result

    for i in range(len(result)):
        if stock <= 0:
            break

        v = to_num(result[i])
        if v <= 0:
            result[i] = 0
            continue

        if v >= stock:
            result[i] = v - stock
            stock = 0
        else:
            stock -= v
            result[i] = 0

    return result


def apply_today_result(row, today_dict, date_start_idx):
    product = norm(row["품번"])
    if product not in today_dict:
        return row

    remain = today_dict[product]
    vals = row.iloc[date_start_idx:].tolist()

    for i in range(len(vals)):
        if remain <= 0:
            break

        v = to_num(vals[i])
        if v <= 0:
            vals[i] = 0
            continue

        if v >= remain:
            vals[i] = v - remain
            remain = 0
        else:
            remain -= v
            vals[i] = 0

    row.iloc[date_start_idx:] = vals
    return row


def find_structure(ws, cache):
    process_col = None
    process_row = None

    for r in range(1, 5):
        for c in range(1, ws.max_column + 1):
            v = norm(get_merged_value(ws, cache, r, c))
            if v == "공정인쇄":
                process_col = c
                process_row = r
                break
        if process_col is not None:
            break

    if process_col is None:
        debug = []
        for c in range(1, min(ws.max_column, 10) + 1):
            vals = []
            for r in range(1, 5):
                vals.append(f"r{r}='{norm(get_merged_value(ws, cache, r, c))}'")
            debug.append(f"col {c}: " + ", ".join(vals))
        raise Exception(f"[{ws.title}] 공정인쇄열을 못찾음\n" + "\n".join(debug))

    model_col = process_col - 2
    product_col = process_col - 1

    shortage_col, plan_cols, date_labels = find_plan_and_shortage_columns(ws, cache, process_col)

    data_start_row = None
    for r in range(process_row + 1, ws.max_row + 1):
        product = norm(get_merged_value(ws, cache, r, product_col))
        process_name = norm(get_merged_value(ws, cache, r, process_col))

        numeric_count = 0
        for c in plan_cols[:3]:
            if to_num(get_merged_value(ws, cache, r, c)) != 0:
                numeric_count += 1

        if product or process_name or numeric_count > 0:
            data_start_row = r
            break

    if data_start_row is None:
        data_start_row = process_row + 1

    return {
        "model_col": model_col,
        "product_col": product_col,
        "process_col": process_col,
        "shortage_col": shortage_col,
        "plan_cols": plan_cols,
        "date_labels": date_labels,
        "data_start_row": data_start_row,
    }


def read_sheet(ws):
    cache = build_sheet_cache(ws)
    s = find_structure(ws, cache)

    model_col = s["model_col"]
    product_col = s["product_col"]
    process_col = s["process_col"]
    shortage_col = s["shortage_col"]
    plan_cols = s["plan_cols"]
    date_labels = s["date_labels"]
    data_start_row = s["data_start_row"]

    rows = []


    for r in range(data_start_row, ws.max_row + 1):
        model = norm(get_merged_value(ws, cache, r, model_col))
        product = norm(get_merged_value(ws, cache, r, product_col))
        process_name = norm(get_merged_value(ws, cache, r, process_col))

        if not product:
            continue

        if should_skip_row(product, process_name):
            continue

        workcenter, team = map_workcenter_and_team(process_name)
        if not workcenter or not team:
            continue

        shortage = get_merged_value(ws, cache, r, shortage_col) if shortage_col else 0
        plan_values = [
            get_qty_value(ws, cache, r, c)
            for c in plan_cols
        ]
        adjusted = adjust_by_shortage(plan_values, shortage)

        if sum(adjusted) == 0:
            continue

        row_data = [workcenter, team, model, product, process_name] + adjusted

        expected_len = 5 + len(date_labels)
        if len(row_data) != expected_len:
            raise Exception(
                f"[{ws.title}] 행 {r} 컬럼 수 불일치: expected={expected_len}, actual={len(row_data)}"
            )

        rows.append(row_data)

    if not rows:
        return None

    columns = ["작업장명", "작업반명", "모델", "품번", "공정 인쇄"] + date_labels
    return pd.DataFrame(rows, columns=columns)


def main():
    Tk().withdraw()

    file_path = filedialog.askopenfilename(
        title="엑셀 선택",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        return

    try:
        wb = load_workbook(file_path, data_only=True)

        today_grouped = {}
        if os.path.exists(OUTPUT_FILE):
            today_df = pd.read_excel(OUTPUT_FILE)
            if {"품번", "실적수량"}.issubset(today_df.columns):
                today_df = today_df[["품번", "실적수량"]].copy()
                today_df["품번"] = today_df["품번"].astype(str).str.strip()
                today_grouped = today_df.groupby("품번")["실적수량"].sum().to_dict()

        result_list = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df = read_sheet(ws)

            if df is None or df.empty:
                continue

            date_start_idx = 5

            agg = {
                "작업장명": "first",
                "작업반명": "first",
                "모델": "first",
                "공정 인쇄": "first",
            }
            for col in df.columns[date_start_idx:]:
                agg[col] = "sum"

            df["품번"] = df["품번"].astype(str).str.strip()
            df = df.groupby("품번", as_index=False).agg(agg)

            if today_grouped:
                df = df.apply(
                    lambda row: apply_today_result(row, today_grouped, date_start_idx),
                    axis=1
                )

            df = df[(df.iloc[:, date_start_idx:] != 0).any(axis=1)]

            if not df.empty:
                result_list.append(df)

        if not result_list:
            raise Exception("조건에 맞는 데이터가 없습니다.")

        final_df = pd.concat(result_list, ignore_index=True)

        output_path = file_path.replace(".xlsx", "_MES작업지시.xlsx")
        final_df.to_excel(output_path, index=False)

        messagebox.showinfo("완료", f"저장 완료:\n{output_path}")

    except Exception as e:
        messagebox.showerror("오류", str(e))


if __name__ == "__main__":
    main()